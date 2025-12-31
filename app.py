"""
Ứng dụng Quản lý Dân cư Tổ dân phố
Backend: Flask + PostgreSQL (psycopg2)
"""

from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify, send_from_directory
import psycopg2
from psycopg2 import Error
from functools import wraps
import os
import math
from datetime import datetime
from io import BytesIO
from werkzeug.utils import secure_filename
import uuid
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from flask import Response
import io
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ========== KHỞI TẠO FLASK APP ==========
app = Flask(__name__, 
            template_folder='Interface/templates',
            static_folder='Interface/static')

app.secret_key = 'quanlydancu_secret_key_2025'

# ========== CẤU HÌNH UPLOAD FILE ==========
UPLOAD_FOLDER = 'file'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'doc', 'docx', 'xls', 'xlsx', 'txt', 'zip', 'rar'}
MAX_FILE_SIZE = 16 * 1024 * 1024  # 16MB

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

# Tạo thư mục file nếu chưa tồn tại
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
# ==================LOAD FONT=====================
pdfmetrics.registerFont(
    TTFont("NotoSans", "asset/fonts/NotoSans.ttf")
)


# ========== CẤU HÌNH DATABASE ==========
DB_CONFIG = {
    'database': 'KTPM',
    'user': 'postgres',
    'password': 'admin',
    'host': 'localhost',
    'port': '5432'

}

# ========== HÀM KẾT NỐI DATABASE ==========
def get_db_connection():
    """
    Tạo kết nối đến PostgreSQL database
    Returns: connection object hoặc None nếu lỗi
    """
    try:
        connection = psycopg2.connect(
            database=DB_CONFIG['database'],
            user=DB_CONFIG['user'],
            password=DB_CONFIG['password'],
            host=DB_CONFIG['host'],
            port=DB_CONFIG['port']
        )
        return connection
    except Error as e:
        print(f"Lỗi kết nối database: {e}")
        return None


def execute_query(query, params=None, fetch_one=False, fetch_all=False):
    """
    Thực thi câu lệnh SQL
    Args:
        query: Câu lệnh SQL
        params: Tham số truyền vào (tuple)
        fetch_one: Lấy 1 bản ghi
        fetch_all: Lấy tất cả bản ghi
    Returns: Kết quả query hoặc None
    """
    connection = get_db_connection()
    if not connection:
        return None
    
    try:
        cursor = connection.cursor()
        cursor.execute(query, params)
        
        if fetch_one:
            result = cursor.fetchone()
        elif fetch_all:
            result = cursor.fetchall()
        else:
            connection.commit()
            # Trả về rowcount (số bản ghi bị ảnh hưởng)
            result = cursor.rowcount if cursor.rowcount >= 0 else 1
        
        cursor.close()
        connection.close()
        return result
    
    except Error as e:
        print(f"Lỗi thực thi query: {e}")
        if connection:
            connection.rollback()
            connection.close()
        return None


# ========== DECORATOR KIỂM TRA ĐĂNG NHẬP ==========
def login_required(f):
    """Decorator yêu cầu đăng nhập"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            flash('Vui lòng đăng nhập để truy cập!', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def role_required(allowed_roles):
    """
    Decorator kiểm tra vai trò người dùng (RBAC)
    
    Args:
        allowed_roles (list): Danh sách vai trò được phép truy cập
                             VD: ['CanBo', 'QuanLy']
    
    Returns:
        Redirect về dashboard với thông báo lỗi nếu không đủ quyền
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # Kiểm tra đăng nhập
            if 'user' not in session:
                flash('Vui lòng đăng nhập để truy cập!', 'warning')
                return redirect(url_for('login'))
            
            # Kiểm tra vai trò
            user_role = session['user'].get('vaitro')
            
            if user_role not in allowed_roles:
                flash('Bạn không có quyền truy cập chức năng này!', 'danger')
                return redirect(url_for('dashboard'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator


# ========== HELPER FUNCTIONS ==========

def tao_thongbao_canhan(cccd, noidung, loai='General', maphananh=None, mavande=None, mathongbao=None):
    """
    Tạo thông báo cá nhân cho người dùng
    
    Args:
        cccd: CCCD người nhận thông báo
        noidung: Nội dung thông báo
        loai: Loại thông báo (General, PhanAnh, VanDe, Chat, System, LichSu)
        maphananh: Mã phản ánh liên quan (nếu có)
        mavande: Mã vấn đề liên quan (nếu có)
        mathongbao: Mã thông báo chung liên quan (nếu có)
    
    Returns:
        mathongbao_nguoidung nếu thành công, None nếu lỗi
    """
    conn = get_db_connection()
    if not conn:
        return None
    
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO thongbao_nguoidung 
            (cccd, noidung, loai, maphananh, mavande, mathongbao)
            VALUES (%s, %s, %s, %s, %s, %s)
            RETURNING mathongbao_nguoidung
        """, (cccd, noidung, loai, maphananh, mavande, mathongbao))
        
        mathongbao_nguoidung = cur.fetchone()[0]
        conn.commit()
        cur.close()
        conn.close()
        
        return mathongbao_nguoidung
    
    except Exception as e:
        print(f"Lỗi tạo thông báo cá nhân: {e}")
        if conn:
            conn.rollback()
            conn.close()
        return None


def allowed_file(filename):
    """
    Kiểm tra file có phải định dạng cho phép không
    """
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def upload_file(file):
    """
    Upload file và lưu vào database
    
    Args:
        file: File object từ request.files
    
    Returns:
        matepdinhkem nếu thành công, None nếu lỗi
    """
    if not file or file.filename == '':
        return None
    
    if not allowed_file(file.filename):
        return None
    
    try:
        # Tạo tên file unique
        filename = secure_filename(file.filename)
        unique_filename = f"{uuid.uuid4().hex}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        
        # Lưu file
        file.save(filepath)
        
        # Lưu vào database
        conn = get_db_connection()
        if not conn:
            os.remove(filepath)  # Xóa file nếu không kết nối được DB
            return None
        
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO tepdinhkem (duongdan)
            VALUES (%s)
            RETURNING matepdinhkem
        """, (unique_filename,))  # Chỉ lưu tên file, không lưu đường dẫn đầy đủ
        
        matepdinhkem = cur.fetchone()[0]
        conn.commit()
        cur.close()
        conn.close()
        
        return matepdinhkem
    
    except Exception as e:
        print(f"Lỗi upload file: {e}")
        if os.path.exists(filepath):
            os.remove(filepath)
        if conn:
            conn.rollback()
            conn.close()
        return None


# ========== ROUTES ==========

@app.route('/file/<filename>')
@login_required
def serve_file(filename):
    """
    Serve file từ thư mục upload
    """
    try:
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename)
    except Exception as e:
        flash(f'Không tìm thấy file: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/image/<path:filename>')
def serve_image(filename):
    """Serve images from image folder"""
    from flask import send_from_directory
    return send_from_directory('image', filename)


@app.route('/')
def index():
    """Trang chủ - Redirect đến login hoặc dashboard"""
    if 'user' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


# ========== ĐĂNG KÝ TÀI KHOẢN CÔNG KHAI ==========

@app.route('/dang-ky', methods=['GET', 'POST'])
def public_register():
    """Trang đăng ký công khai cho người dân (không cần đăng nhập)"""
    
    if request.method == 'POST':
        # Lấy thông tin từ form
        cccd = request.form.get('cccd', '').strip()
        hoten = request.form.get('hoten', '').strip()
        ngaysinh = request.form.get('ngaysinh', '').strip()
        gioitinh = request.form.get('gioitinh', '').strip()
        sdt = request.form.get('sdt', '').strip()
        loaidangky = request.form.get('loaidangky', '').strip()
        tinh = request.form.get('tinh', '').strip()
        xaphuong = request.form.get('xaphuong', '').strip()
        diachi_chitiet = request.form.get('diachi_chitiet', '').strip()
        quoctich = request.form.get('quoctich', 'Việt Nam').strip()
        dantoc = request.form.get('dantoc', '').strip()
        ghichu = request.form.get('ghichu', '').strip()
        
        # Validation
        if not all([cccd, hoten, ngaysinh, gioitinh, sdt, loaidangky, tinh, xaphuong, diachi_chitiet]):
            flash('Vui lòng điền đầy đủ các thông tin bắt buộc!', 'danger')
            return redirect(url_for('public_register'))
        
        # Kiểm tra CCCD đã tồn tại trong hệ thống chưa
        check_user = execute_query("SELECT cccd FROM nguoidung WHERE cccd = %s", (cccd,), fetch_one=True)
        if check_user:
            flash('CCCD này đã được đăng ký trong hệ thống! Vui lòng đăng nhập hoặc liên hệ cán bộ.', 'warning')
            return redirect(url_for('login'))
        
        # Kiểm tra đơn đăng ký đã tồn tại chưa
        check_existing = execute_query(
            "SELECT madondangky, trangthai FROM dondangky WHERE cccd = %s",
            (cccd,),
            fetch_one=True
        )
        if check_existing:
            status = check_existing[1]
            if status == 'ChoDuyet':
                flash('Đơn đăng ký của bạn đang chờ duyệt. Vui lòng đợi cán bộ xử lý.', 'info')
            elif status == 'DaDuyet':
                flash('Đơn đăng ký của bạn đã được duyệt. Vui lòng đến cơ quan để nhận mật khẩu đăng nhập.', 'success')
            else:
                flash('Đơn đăng ký của bạn đã bị từ chối. Vui lòng liên hệ cán bộ để biết thêm chi tiết.', 'warning')
            return redirect(url_for('public_register'))
        
        # Tạo đơn đăng ký mới
        query = """
            INSERT INTO dondangky (cccd, hoten, ngaysinh, gioitinh, sdt, 
                                   loaidangky, tinh, xaphuong, diachi_chitiet, 
                                   quoctich, dantoc, ghichu, trangthai)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'ChoDuyet')
        """
        result = execute_query(query, (cccd, hoten, ngaysinh, gioitinh, sdt,
                                      loaidangky, tinh, xaphuong, diachi_chitiet,
                                      quoctich, dantoc, ghichu))
        
        if result:
            flash('Đăng ký thành công! Đơn của bạn đang chờ cán bộ duyệt. Vui lòng đợi thông báo.', 'success')
            return redirect(url_for('login'))
        else:
            flash('Lỗi khi tạo đơn đăng ký. Vui lòng thử lại!', 'danger')
            return redirect(url_for('public_register'))
    
    # GET - Hiển thị form đăng ký
    from datetime import datetime
    return render_template('dang_ky.html', now=datetime.now())


@app.route('/tra-cuu-don', methods=['GET', 'POST'])
def tra_cuu_don():
    """Tra cứu trạng thái đơn đăng ký bằng CCCD"""
    
    don_info = None
    
    if request.method == 'POST':
        cccd = request.form.get('cccd', '').strip()
        
        if not cccd:
            flash('Vui lòng nhập số CCCD!', 'danger')
            return redirect(url_for('tra_cuu_don'))
        
        # Kiểm tra trong bảng nguoidung trước
        check_user = execute_query(
            "SELECT cccd, name FROM nguoidung WHERE cccd = %s", 
            (cccd,), 
            fetch_one=True
        )
        
        if check_user:
            flash('CCCD này đã được kích hoạt trong hệ thống! Bạn có thể đăng nhập.', 'success')
            return render_template('tra_cuu_don.html', don_info=None, is_activated=True)
        
        # Tra cứu đơn đăng ký
        query = """
            SELECT 
                dd.madondangky,
                dd.cccd,
                dd.hoten,
                dd.ngaysinh,
                dd.gioitinh,
                dd.sdt,
                dd.loaidangky,
                dd.tinh,
                dd.xaphuong,
                dd.diachi_chitiet,
                dd.trangthai,
                dd.ngaytao,
                dd.ngayduyet,
                dd.matkhau_daxacnhan,
                dd.lydotuchoi,
                nd.name as nguoiduyet_ten
            FROM dondangky dd
            LEFT JOIN nguoidung nd ON dd.nguoiduyet_cccd = nd.cccd
            WHERE dd.cccd = %s
            ORDER BY dd.ngaytao DESC
            LIMIT 1
        """
        
        don_info = execute_query(query, (cccd,), fetch_one=True)
        
        if not don_info:
            flash('Không tìm thấy đơn đăng ký với số CCCD này!', 'warning')
    
    from datetime import datetime
    return render_template('tra_cuu_don.html', don_info=don_info, is_activated=False, now=datetime.now())


@app.route('/don-dang-ky')
@login_required
@role_required(['CanBo', 'QuanLy'])
def don_dang_ky_list():
    """Danh sách đơn đăng ký chờ duyệt (chỉ cán bộ/quản lý)"""
    
    # Lấy tham số lọc
    trangthai = request.args.get('trangthai', 'ChoDuyet')
    search = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    # Build query
    query = """
        SELECT 
            dd.madondangky,
            dd.cccd,
            dd.hoten,
            dd.ngaysinh,
            dd.gioitinh,
            dd.sdt,
            dd.loaidangky,
            dd.tinh,
            dd.xaphuong,
            dd.diachi_chitiet,
            dd.trangthai,
            dd.ngaytao,
            dd.ngayduyet,
            nd.name as nguoiduyet_ten,
            dd.matkhau_daxacnhan
        FROM dondangky dd
        LEFT JOIN nguoidung nd ON dd.nguoiduyet_cccd = nd.cccd
        WHERE 1=1
    """
    
    params = []
    
    if trangthai:
        query += " AND dd.trangthai = %s"
        params.append(trangthai)
    
    if search:
        query += " AND (dd.cccd ILIKE %s OR dd.hoten ILIKE %s OR dd.sdt ILIKE %s)"
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])
    
    # Count total
    count_query = f"SELECT COUNT(*) FROM ({query}) AS subquery"
    total_result = execute_query(count_query, tuple(params), fetch_one=True)
    total = total_result[0] if total_result else 0
    
    # Statistics
    stats = {
        'choduyet': execute_query("SELECT COUNT(*) FROM dondangky WHERE trangthai = 'ChoDuyet'", fetch_one=True)[0] or 0,
        'daduyet': execute_query("SELECT COUNT(*) FROM dondangky WHERE trangthai = 'DaDuyet'", fetch_one=True)[0] or 0,
        'tuchoi': execute_query("SELECT COUNT(*) FROM dondangky WHERE trangthai = 'TuChoi'", fetch_one=True)[0] or 0,
        'chua_giao_mk': execute_query("SELECT COUNT(*) FROM dondangky WHERE trangthai = 'DaDuyet' AND matkhau_daxacnhan = FALSE", fetch_one=True)[0] or 0
    }
    
    # Pagination
    query += " ORDER BY dd.ngaytao DESC"
    offset = (page - 1) * per_page
    query += f" LIMIT {per_page} OFFSET {offset}"
    
    records = execute_query(query, tuple(params), fetch_all=True)
    records = records if records else []
    
    total_pages = (total + per_page - 1) // per_page if total > 0 else 1
    
    from datetime import datetime
    return render_template('don_dang_ky_list.html',
                         records=records,
                         page=page,
                         total_pages=total_pages,
                         total=total,
                         stats=stats,
                         trangthai=trangthai,
                         search=search,
                         now=datetime.now())


@app.route('/don-dang-ky/<int:madondangky>')
@login_required
@role_required(['CanBo', 'QuanLy'])
def don_dang_ky_detail(madondangky):
    """Xem chi tiết đơn đăng ký"""
    
    query = """
        SELECT 
            dd.*,
            nd1.name as nguoiduyet_ten,
            nd2.name as nguoixacnhan_ten
        FROM dondangky dd
        LEFT JOIN nguoidung nd1 ON dd.nguoiduyet_cccd = nd1.cccd
        LEFT JOIN nguoidung nd2 ON dd.nguoixacnhan_cccd = nd2.cccd
        WHERE dd.madondangky = %s
    """
    record = execute_query(query, (madondangky,), fetch_one=True)
    
    if not record:
        flash('Không tìm thấy đơn đăng ký!', 'danger')
        return redirect(url_for('don_dang_ky_list'))
    
    from datetime import datetime
    return render_template('don_dang_ky_detail.html', record=record, now=datetime.now())


@app.route('/don-dang-ky/<int:madondangky>/duyet', methods=['POST'])
@login_required
@role_required(['CanBo', 'QuanLy'])
def don_dang_ky_duyet(madondangky):
    """Duyệt đơn đăng ký - Tạo tài khoản và mật khẩu tạm"""
    
    user_cccd = session['user']['cccd']
    
    # Lấy thông tin đơn
    query = "SELECT * FROM dondangky WHERE madondangky = %s AND trangthai = 'ChoDuyet'"
    don = execute_query(query, (madondangky,), fetch_one=True)
    
    if not don:
        flash('Không tìm thấy đơn hoặc đơn đã được xử lý!', 'danger')
        return redirect(url_for('don_dang_ky_list'))
    
    # Tạo mật khẩu tạm: pass + random(1-9999)
    import random
    matkhau_tam = f"pass{random.randint(1, 9999):04d}"
    
    connection = get_db_connection()
    if not connection:
        flash('Lỗi kết nối database!', 'danger')
        return redirect(url_for('don_dang_ky_detail', madondangky=madondangky))
    
    try:
        cursor = connection.cursor()
        
        # 1. Tạo tài khoản trong bảng nguoidung
        cursor.execute(
            """INSERT INTO nguoidung (cccd, name, ngaysinh, gioitinh, sdt, dantoc, 
                                      vaitro, user_name, matkhau)
               VALUES (%s, %s, %s, %s, %s, %s, 'NguoiDan', %s, %s)""",
            (don[1], don[2], don[3], don[4], don[5], don[11],  # cccd, hoten, ngaysinh, gioitinh, sdt, dantoc
             don[1], matkhau_tam)  # username = cccd, matkhau = matkhau_tam
        )
        
        # 2. Tạo địa chỉ
        cursor.execute(
            "INSERT INTO diachi (tinh, xaphuong, chitiet) VALUES (%s, %s, %s) RETURNING madiachi",
            (don[8], don[9], don[10])  # tinh, xaphuong, diachi_chitiet
        )
        madiachi = cursor.fetchone()[0]
        
        # 3. Tạo liên kết địa chỉ - người dùng
        loaidiachi = 'TamTru' if don[7] == 'TamTru' else 'CuTru'
        cursor.execute(
            """INSERT INTO diachinguoidung (madiachi, cccd, loaidiachi, thoidiemxacnhan)
               VALUES (%s, %s, %s, CURRENT_DATE)""",
            (madiachi, don[1], loaidiachi)
        )
        
        # 4. Cập nhật trạng thái đơn
        cursor.execute(
            """UPDATE dondangky 
               SET trangthai = 'DaDuyet', 
                   nguoiduyet_cccd = %s, 
                   ngayduyet = CURRENT_TIMESTAMP,
                   matkhau_tam = %s,
                   matkhau_daxacnhan = FALSE
               WHERE madondangky = %s""",
            (user_cccd, matkhau_tam, madondangky)
        )
        
        connection.commit()
        cursor.close()
        connection.close()
        
        # 5. Tạo thông báo cho người dân
        tao_thongbao_canhan(
            cccd=don[1],
            noidung=f"Đơn đăng ký của bạn đã được duyệt. Vui lòng đến trực tiếp để nhận mật khẩu đăng nhập lần đầu.",
            loai='System'
        )
        
        flash(f'Đã duyệt đơn thành công! Mật khẩu tạm: <strong>{matkhau_tam}</strong> - Vui lòng ghi nhớ để giao cho người dân.', 'success')
        return redirect(url_for('don_dang_ky_detail', madondangky=madondangky))
        
    except Error as e:
        connection.rollback()
        cursor.close()
        connection.close()
        flash(f'Lỗi khi duyệt đơn: {str(e)}', 'danger')
        return redirect(url_for('don_dang_ky_detail', madondangky=madondangky))


@app.route('/don-dang-ky/<int:madondangky>/tu-choi', methods=['POST'])
@login_required
@role_required(['CanBo', 'QuanLy'])
def don_dang_ky_tu_choi(madondangky):
    """Từ chối đơn đăng ký"""
    
    user_cccd = session['user']['cccd']
    lydotuchoi = request.form.get('lydotuchoi', '').strip()
    
    if not lydotuchoi:
        flash('Vui lòng nhập lý do từ chối!', 'danger')
        return redirect(url_for('don_dang_ky_detail', madondangky=madondangky))
    
    # Lấy thông tin đơn để gửi thông báo
    query = "SELECT cccd FROM dondangky WHERE madondangky = %s"
    don = execute_query(query, (madondangky,), fetch_one=True)
    
    if not don:
        flash('Không tìm thấy đơn!', 'danger')
        return redirect(url_for('don_dang_ky_list'))
    
    query = """
        UPDATE dondangky 
        SET trangthai = 'TuChoi',
            nguoiduyet_cccd = %s,
            ngayduyet = CURRENT_TIMESTAMP,
            lydotuchoi = %s
        WHERE madondangky = %s AND trangthai = 'ChoDuyet'
    """
    result = execute_query(query, (user_cccd, lydotuchoi, madondangky))
    
    if result:
        # Tạo thông báo cho người dân
        tao_thongbao_canhan(
            cccd=don[0],
            noidung=f"Đơn đăng ký của bạn đã bị từ chối. Lý do: {lydotuchoi}",
            loai='System'
        )
        flash('Đã từ chối đơn đăng ký!', 'info')
    else:
        flash('Lỗi khi từ chối đơn!', 'danger')
    
    return redirect(url_for('don_dang_ky_list'))


@app.route('/don-dang-ky/<int:madondangky>/xac-nhan-giao-matkhau', methods=['POST'])
@login_required
@role_required(['CanBo', 'QuanLy'])
def don_dang_ky_xacnhan_matkhau(madondangky):
    """Xác nhận đã giao mật khẩu cho người dân"""
    
    user_cccd = session['user']['cccd']
    
    query = """
        UPDATE dondangky 
        SET matkhau_daxacnhan = TRUE,
            nguoixacnhan_cccd = %s,
            ngayxacnhan = CURRENT_TIMESTAMP
        WHERE madondangky = %s AND trangthai = 'DaDuyet' AND matkhau_daxacnhan = FALSE
    """
    result = execute_query(query, (user_cccd, madondangky))
    
    if result:
        flash('Đã xác nhận giao mật khẩu cho người dân!', 'success')
    else:
        flash('Lỗi khi xác nhận!', 'danger')
    
    return redirect(url_for('don_dang_ky_detail', madondangky=madondangky))


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Trang đăng nhập bằng CCCD và mật khẩu"""
    if request.method == 'POST':
        cccd = request.form.get('cccd', '').strip()
        password = request.form.get('password', '').strip()
        
        if not cccd or not password:
            flash('Vui lòng nhập đầy đủ thông tin!', 'danger')
            return render_template('login.html')
        
        # Query kiểm tra đăng nhập bằng CCCD
        query = """
            SELECT cccd, name, vaitro, user_name, avatar_url 
            FROM nguoidung 
            WHERE cccd = %s AND matkhau = %s
        """
        user = execute_query(query, (cccd, password), fetch_one=True)
        
        if user:
            session['user'] = {
                'cccd': user[0].strip() if user[0] else '',
                'name': user[1],
                'vaitro': user[2],
                'user_name': user[3],
                'avatar_url': user[4] if user[4] else './image/default_avatar.png'
            }
            flash(f'Chào mừng {user[1]}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('CCCD hoặc mật khẩu không đúng!', 'danger')
    
    return render_template('login.html')


@app.route('/logout')
def logout():
    """Đăng xuất"""
    session.pop('user', None)
    flash('Đã đăng xuất thành công!', 'info')
    return redirect(url_for('login'))


@app.route('/profile')
@login_required
def profile():
    """Trang thông tin cá nhân"""
    user_cccd = session['user']['cccd']
    
    # Lấy thông tin đầy đủ của người dùng
    query = """
        SELECT 
            n.cccd,
            n.name,
            n.sdt,
            n.ngaysinh,
            n.gioitinh,
            n.dantoc,
            n.vaitro,
            n.user_name,
            n.baomatthongtin,
            d.tinh,
            d.xaphuong,
            d.chitiet,
            n.avatar_url
        FROM nguoidung n
        LEFT JOIN diachinguoidung dn ON n.cccd = dn.cccd AND dn.loaidiachi = 'CuTru'
        LEFT JOIN diachi d ON dn.madiachi = d.madiachi
        WHERE n.cccd = %s
    """
    user_info = execute_query(query, (user_cccd,), fetch_one=True)
    
    if not user_info:
        flash('Không tìm thấy thông tin người dùng!', 'danger')
        return redirect(url_for('dashboard'))
    
    # Lấy thông tin hộ khẩu nếu có
    query_hokhau = """
        SELECT 
            h.mahokhau,
            tv.quanhechuho,
            h.ngaycap,
            d.tinh as hk_tinh,
            d.xaphuong as hk_xaphuong,
            d.chitiet as hk_chitiet
        FROM thanhvienhokhau tv
        JOIN hokhau h ON tv.mahokhau = h.mahokhau
        LEFT JOIN diachi d ON h.madiachi = d.madiachi
        WHERE tv.cccd = %s AND tv.ngayketthuc IS NULL
    """
    hokhau_info = execute_query(query_hokhau, (user_cccd,), fetch_one=True)
    
    return render_template('profile.html', user_info=user_info, hokhau_info=hokhau_info)


@app.route('/profile/edit', methods=['GET', 'POST'])
@login_required
def profile_edit():
    """Chỉnh sửa thông tin cá nhân"""
    user_cccd = session['user']['cccd']
    
    if request.method == 'POST':
        sdt = request.form.get('sdt', '').strip()
        user_name = request.form.get('user_name', '').strip()
        
        # Validate
        if not sdt or not user_name:
            flash('Vui lòng nhập đầy đủ thông tin!', 'warning')
            return redirect(url_for('profile_edit'))
        
        # Xử lý upload ảnh đại diện
        avatar_url = None
        if 'avatar' in request.files:
            file = request.files['avatar']
            if file and file.filename != '':
                # Kiểm tra định dạng file
                allowed_extensions = {'png', 'jpg', 'jpeg', 'gif'}
                file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
                
                if file_ext in allowed_extensions:
                    # Lưu file với tên cccd.png
                    import os
                    filename = f"{user_cccd}.png"
                    filepath = os.path.join('image', filename)
                    file.save(filepath)
                    avatar_url = f"./image/{filename}"
                else:
                    flash('Chỉ chấp nhận file ảnh (PNG, JPG, JPEG, GIF)!', 'warning')
                    return redirect(url_for('profile_edit'))
        
        # Update thông tin
        if avatar_url:
            query = """
                UPDATE nguoidung
                SET sdt = %s, user_name = %s, avatar_url = %s
                WHERE cccd = %s
            """
            execute_query(query, (sdt, user_name, avatar_url, user_cccd))
            # Cập nhật session với avatar mới
            session['user']['avatar_url'] = avatar_url
        else:
            query = """
                UPDATE nguoidung
                SET sdt = %s, user_name = %s
                WHERE cccd = %s
            """
            execute_query(query, (sdt, user_name, user_cccd))
        
        # Cập nhật session
        session['user']['sdt'] = sdt
        session['user']['user_name'] = user_name
        
        flash('Đã cập nhật thông tin thành công!', 'success')
        return redirect(url_for('profile'))
    
    # GET - hiển thị form
    query = """
        SELECT cccd, name, sdt, user_name, ngaysinh, gioitinh, dantoc, vaitro, avatar_url
        FROM nguoidung
        WHERE cccd = %s
    """
    user_info = execute_query(query, (user_cccd,), fetch_one=True)
    
    return render_template('profile_edit.html', user_info=user_info)


@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    """Trang cài đặt - Đổi mật khẩu và cài đặt quyền riêng tư"""
    user_cccd = session['user']['cccd']
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'change_password':
            old_password = request.form.get('old_password', '').strip()
            new_password = request.form.get('new_password', '').strip()
            confirm_password = request.form.get('confirm_password', '').strip()
            
            # Validate
            if not old_password or not new_password or not confirm_password:
                flash('Vui lòng nhập đầy đủ thông tin!', 'warning')
                return redirect(url_for('settings'))
            
            if new_password != confirm_password:
                flash('Mật khẩu mới không khớp!', 'danger')
                return redirect(url_for('settings'))
            
            # Kiểm tra mật khẩu cũ
            query_check = "SELECT matkhau FROM nguoidung WHERE cccd = %s"
            result = execute_query(query_check, (user_cccd,), fetch_one=True)
            
            if not result or result[0] != old_password:
                flash('Mật khẩu cũ không đúng!', 'danger')
                return redirect(url_for('settings'))
            
            # Cập nhật mật khẩu mới
            query_update = "UPDATE nguoidung SET matkhau = %s WHERE cccd = %s"
            execute_query(query_update, (new_password, user_cccd))
            
            flash('Đã đổi mật khẩu thành công!', 'success')
            return redirect(url_for('settings'))
        
        elif action == 'update_privacy':
            baomatthongtin = request.form.get('baomatthongtin') == 'on'
            
            query_update = "UPDATE nguoidung SET baomatthongtin = %s WHERE cccd = %s"
            execute_query(query_update, (baomatthongtin, user_cccd))
            
            flash('Đã cập nhật cài đặt quyền riêng tư!', 'success')
            return redirect(url_for('settings'))
    
    # GET - hiển thị trang cài đặt
    query = "SELECT baomatthongtin FROM nguoidung WHERE cccd = %s"
    user_settings = execute_query(query, (user_cccd,), fetch_one=True)
    
    return render_template('settings.html', baomatthongtin=user_settings[0] if user_settings else True)


@app.route('/dashboard')
@login_required
def dashboard():
    """Trang Dashboard - Thống kê và danh sách hộ khẩu với phân trang"""
    
    # ========== CẤU HÌNH PHÂN TRANG ==========
    per_page = 20  # Số bản ghi mỗi trang
    
    # Lấy số trang từ URL, mặc định là 1
    try:
        page = int(request.args.get('page', 1))
        if page < 1:
            page = 1
    except (ValueError, TypeError):
        page = 1
    
    # Tính offset
    offset = (page - 1) * per_page
    
    # ========== THỐNG KÊ ==========
    # Thống kê tổng số nhân khẩu
    query_nhankhau = "SELECT COUNT(*) FROM nguoidung"
    total_nhankhau = execute_query(query_nhankhau, fetch_one=True)
    total_nhankhau = total_nhankhau[0] if total_nhankhau else 0
    
    # Thống kê tổng số hộ khẩu (dùng cho pagination)
    query_hokhau = "SELECT COUNT(*) FROM hokhau"
    total_hokhau = execute_query(query_hokhau, fetch_one=True)
    total_hokhau = total_hokhau[0] if total_hokhau else 0
    
    # Thống kê tổng số phản ánh
    query_phananh = "SELECT COUNT(*) FROM phananh"
    total_phananh = execute_query(query_phananh, fetch_one=True)
    total_phananh = total_phananh[0] if total_phananh else 0
    
    # Thống kê phản ánh chưa xử lý
    query_chuaxuly = "SELECT COUNT(*) FROM phananh WHERE trangthaiphananh = 'ChuaXuLy'"
    total_chuaxuly = execute_query(query_chuaxuly, fetch_one=True)
    total_chuaxuly = total_chuaxuly[0] if total_chuaxuly else 0
    
    # ========== PHÂN TRANG DANH SÁCH HỘ KHẨU ==========
    # Tính tổng số trang
    import math
    total_pages = math.ceil(total_hokhau / per_page) if total_hokhau > 0 else 1
    
    # Đảm bảo page không vượt quá total_pages
    if page > total_pages:
        page = total_pages
    
    # Query danh sách hộ khẩu với LIMIT và OFFSET
    query_dshokhau = """
        SELECT 
            h.mahokhau,
            d.tinh,
            d.xaphuong,
            d.chitiet,
            h.ngaycap,
            h.ghichu,
            (SELECT COUNT(*) FROM thanhvienhokhau tv WHERE tv.mahokhau = h.mahokhau AND tv.ngayketthuc IS NULL) as so_thanh_vien,
            (SELECT n.name FROM thanhvienhokhau tv 
             JOIN nguoidung n ON tv.cccd = n.cccd 
             WHERE tv.mahokhau = h.mahokhau AND tv.quanhechuho = 'ChuHo' AND tv.ngayketthuc IS NULL LIMIT 1) as chu_ho
        FROM hokhau h
        LEFT JOIN diachi d ON h.madiachi = d.madiachi
        ORDER BY h.mahokhau DESC
        LIMIT %s OFFSET %s
    """
    ds_hokhau = execute_query(query_dshokhau, (per_page, offset), fetch_all=True)
    ds_hokhau = ds_hokhau if ds_hokhau else []
    
    return render_template('dashboard.html',
                         total_nhankhau=total_nhankhau,
                         total_hokhau=total_hokhau,
                         total_phananh=total_phananh,
                         total_chuaxuly=total_chuaxuly,
                         ds_hokhau=ds_hokhau,
                         page=page,
                         per_page=per_page,
                         total_pages=total_pages)


@app.route('/hokhau')
@login_required
@role_required(['CanBo', 'QuanLy'])
def hokhau_list():
    """Danh sách hộ khẩu chi tiết với phân trang, tìm kiếm và lọc"""
    
    user_role = session['user'].get('vaitro')
    user_cccd = session['user'].get('cccd')
    
    # ========== CẤU HÌNH PHÂN TRANG ==========
    per_page = 20  # Số bản ghi mỗi trang
    
    # Lấy số trang từ URL, mặc định là 1
    try:
        page = int(request.args.get('page', 1))
        if page < 1:
            page = 1
    except (ValueError, TypeError):
        page = 1
    
    # Tính offset
    offset = (page - 1) * per_page
    
    # ========== LẤY THAM SỐ TÌM KIẾM VÀ LỌC ==========
    search = request.args.get('search', '').strip()
    xaphuong = request.args.get('xaphuong', '').strip()
    
    # ========== XÂY DỰNG ĐIỀU KIỆN WHERE ==========
    where_conditions = []
    params = []
    
    # Nếu là Người dân, chỉ xem hộ khẩu của mình
    if user_role == 'NguoiDan':
        where_conditions.append("h.mahokhau IN (SELECT mahokhau FROM thanhvienhokhau WHERE cccd = %s)")
        params.append(user_cccd)
    
    if search:
        # Xóa tiền tố "HK" nếu user nhập "HK1" -> tìm kiếm "1" trong DB
        search_for_id = search.replace('HK', '').replace('hk', '')
        # Tìm kiếm theo: mã hộ khẩu, địa chỉ chi tiết, xã/phường
        where_conditions.append("(h.mahokhau::TEXT LIKE %s OR d.chitiet ILIKE %s OR d.xaphuong ILIKE %s)")
        search_pattern = f"%{search_for_id}%"
        params.extend([search_pattern, search_pattern, search_pattern])
    
    # xaphuong là tùy chọn, không bắt buộc
    if xaphuong:
        where_conditions.append("d.xaphuong ILIKE %s")
        params.append(f"%{xaphuong}%")
    
    where_clause = " AND ".join(where_conditions) if where_conditions else "1=1"
    
    # ========== ĐẾM TỔNG SỐ HỘ KHẨU ==========
    query_count = f"""
        SELECT COUNT(*) 
        FROM hokhau h
        LEFT JOIN diachi d ON h.madiachi = d.madiachi
        WHERE {where_clause}
    """
    total_count = execute_query(query_count, tuple(params), fetch_one=True)
    total_count = total_count[0] if total_count else 0
    
    # Tính tổng số trang
    import math
    total_pages = math.ceil(total_count / per_page) if total_count > 0 else 1
    
    # Đảm bảo page không vượt quá total_pages
    if page > total_pages and total_pages > 0:
        page = total_pages
    
    # ========== QUERY DANH SÁCH HỘ KHẨU VỚI PHÂN TRANG ==========
    query = f"""
        SELECT 
            h.mahokhau,
            d.tinh,
            d.xaphuong,
            d.chitiet,
            h.ngaycap,
            h.ghichu,
            (SELECT COUNT(*) FROM thanhvienhokhau tv WHERE tv.mahokhau = h.mahokhau AND tv.ngayketthuc IS NULL) as so_thanh_vien,
            (SELECT n.name FROM thanhvienhokhau tv 
             JOIN nguoidung n ON tv.cccd = n.cccd 
             WHERE tv.mahokhau = h.mahokhau AND tv.quanhechuho = 'ChuHo' AND tv.ngayketthuc IS NULL LIMIT 1) as chu_ho
        FROM hokhau h
        LEFT JOIN diachi d ON h.madiachi = d.madiachi
        WHERE {where_clause}
        ORDER BY h.mahokhau DESC
        LIMIT %s OFFSET %s
    """
    params.extend([per_page, offset])
    ds_hokhau = execute_query(query, tuple(params), fetch_all=True)
    ds_hokhau = ds_hokhau if ds_hokhau else []
    
    return render_template('hokhau.html', 
                         ds_hokhau=ds_hokhau,
                         page=page,
                         per_page=per_page,
                         total_pages=total_pages,
                         total_count=total_count,
                         search=search,
                         xaphuong=xaphuong)


@app.route('/hokhau/add', methods=['GET', 'POST'])
@login_required
@role_required(['CanBo', 'QuanLy'])
def hokhau_add():
    """Thêm hộ khẩu mới"""
    
    if request.method == 'POST':
        # Lấy thông tin địa chỉ
        xaphuong = request.form.get('xaphuong', '').strip()
        diachichitiet = request.form.get('diachichitiet', '').strip()
        ghichu = request.form.get('ghichu', '').strip()
        
        # Lấy thông tin chủ hộ và thành viên
        chuho_cccd = request.form.get('chuho_cccd', '').strip()
        member_cccd_list = request.form.getlist('member_cccd[]')
        member_quanhe_list = request.form.getlist('member_quanhe[]')
        
        # Validate
        if not xaphuong or not diachichitiet:
            flash('Vui lòng nhập đầy đủ thông tin địa chỉ (Xã/Phường và Địa chỉ chi tiết)!', 'danger')
            return render_template('hokhau_add.html')
        
        if not chuho_cccd:
            flash('Vui lòng nhập CCCD chủ hộ!', 'danger')
            return render_template('hokhau_add.html')
        
        # Kiểm tra CCCD chủ hộ có tồn tại không
        check_chuho = "SELECT cccd, name FROM nguoidung WHERE cccd = %s"
        nguoi_chuho = execute_query(check_chuho, (chuho_cccd,), fetch_one=True)
        if not nguoi_chuho:
            flash(f'CCCD chủ hộ "{chuho_cccd}" không tồn tại trong hệ thống!', 'danger')
            return render_template('hokhau_add.html')
        
        # Kiểm tra các CCCD thành viên có tồn tại không
        valid_members = []
        for i, cccd_tv in enumerate(member_cccd_list):
            cccd_tv = cccd_tv.strip()
            if cccd_tv:  # Nếu có nhập CCCD
                if cccd_tv == chuho_cccd:
                    flash(f'Chủ hộ không thể là thành viên khác!', 'danger')
                    return render_template('hokhau_add.html')
                
                # Kiểm tra CCCD có tồn tại
                check_member = "SELECT cccd, name FROM nguoidung WHERE cccd = %s"
                nguoi_tv = execute_query(check_member, (cccd_tv,), fetch_one=True)
                if not nguoi_tv:
                    flash(f'CCCD thành viên "{cccd_tv}" không tồn tại trong hệ thống!', 'danger')
                    return render_template('hokhau_add.html')
                
                quanhe = member_quanhe_list[i] if i < len(member_quanhe_list) else 'Khac'
                if not quanhe:
                    flash(f'Vui lòng chọn quan hệ cho thành viên "{nguoi_tv[1]}"!', 'danger')
                    return render_template('hokhau_add.html')
                
                valid_members.append((cccd_tv, quanhe))
        
        # Bước 1: Insert địa chỉ
        insert_diachi = """
            INSERT INTO diachi (tinh, xaphuong, chitiet)
            VALUES (%s, %s, %s)
            RETURNING madiachi
        """
        connection = get_db_connection()
        if not connection:
            flash('Lỗi kết nối database!', 'danger')
            return render_template('hokhau_add.html')
        
        try:
            cursor = connection.cursor()
            
            # Insert địa chỉ
            cursor.execute(insert_diachi, ('Hà Nội', xaphuong, diachichitiet))
            madiachi = cursor.fetchone()[0]
            
            # Bước 2: Insert hộ khẩu
            insert_hokhau = """
                INSERT INTO hokhau (madiachi, ghichu)
                VALUES (%s, %s)
                RETURNING mahokhau
            """
            cursor.execute(insert_hokhau, (madiachi, ghichu or None))
            mahokhau = cursor.fetchone()[0]
            
            # Bước 3: Insert chủ hộ
            insert_chuho = """
                INSERT INTO thanhvienhokhau (mahokhau, cccd, quanhechuho, ngaybatdau)
                VALUES (%s, %s, %s, CURRENT_DATE)
            """
            cursor.execute(insert_chuho, (mahokhau, chuho_cccd, 'ChuHo'))
            
            # Bước 4: Insert thành viên khác (nếu có)
            for cccd_tv, quanhe in valid_members:
                insert_thanhvien = """
                    INSERT INTO thanhvienhokhau (mahokhau, cccd, quanhechuho, ngaybatdau)
                    VALUES (%s, %s, %s, CURRENT_DATE)
                """
                cursor.execute(insert_thanhvien, (mahokhau, cccd_tv, quanhe))
            
            connection.commit()
            cursor.close()
            connection.close()
            
            flash(f'Đã thêm hộ khẩu HK{mahokhau} (Chủ hộ: {nguoi_chuho[1]}) thành công!', 'success')
            return redirect(url_for('hokhau_detail', mahokhau=mahokhau))
            
        except Error as e:
            if connection:
                connection.rollback()
                connection.close()
            flash(f'Có lỗi xảy ra: {str(e)}', 'danger')
            return render_template('hokhau_add.html')
    
    return render_template('hokhau_add.html')


@app.route('/hokhau/<int:mahokhau>')
@login_required
def hokhau_detail(mahokhau):
    """Chi tiết một hộ khẩu"""
    
    user_role = session['user'].get('vaitro')
    user_cccd = session['user'].get('cccd')
    
    # Thông tin hộ khẩu
    query_hokhau = """
        SELECT 
            h.mahokhau,
            d.tinh,
            d.xaphuong,
            d.chitiet,
            h.ngaycap,
            h.ghichu
        FROM hokhau h
        LEFT JOIN diachi d ON h.madiachi = d.madiachi
        WHERE h.mahokhau = %s
    """
    hokhau = execute_query(query_hokhau, (mahokhau,), fetch_one=True)
    
    if not hokhau:
        flash('Không tìm thấy hộ khẩu!', 'danger')
        return redirect(url_for('hokhau_list'))
    
    # Nếu là Người dân, kiểm tra xem có phải thành viên trong hộ không
    if user_role == 'NguoiDan':
        query_check = "SELECT 1 FROM thanhvienhokhau WHERE mahokhau = %s AND cccd = %s"
        is_member = execute_query(query_check, (mahokhau, user_cccd), fetch_one=True)
        if not is_member:
            flash('Bạn không có quyền xem hộ khẩu này!', 'danger')
            return redirect(url_for('hokhau_list'))
    
    # Danh sách thành viên (chỉ lấy người còn trong hộ)
    query_thanhvien = """
        SELECT 
            n.cccd,
            n.name,
            n.ngaysinh,
            n.gioitinh,
            n.dantoc,
            n.sdt,
            tv.quanhechuho,
            tv.ngaybatdau
        FROM thanhvienhokhau tv
        JOIN nguoidung n ON tv.cccd = n.cccd
        WHERE tv.mahokhau = %s AND tv.ngayketthuc IS NULL
        ORDER BY 
            CASE WHEN tv.quanhechuho = 'ChuHo' THEN 0 ELSE 1 END,
            tv.ngaybatdau
    """
    thanhvien = execute_query(query_thanhvien, (mahokhau,), fetch_all=True)
    thanhvien = thanhvien if thanhvien else []
    
    return render_template('hokhau_detail.html', hokhau=hokhau, thanhvien=thanhvien)


@app.route('/hokhau/edit/<int:mahokhau>', methods=['GET', 'POST'])
@role_required(['CanBo', 'QuanLy'])
@login_required
def hokhau_edit(mahokhau):
    """Sửa thông tin hộ khẩu"""
    
    if request.method == 'POST':
        # Lấy dữ liệu từ form
        xaphuong = request.form.get('xaphuong', '').strip()
        chitiet = request.form.get('chitiet', '').strip()
        ghichu = request.form.get('ghichu', '').strip()
        
        # Validate
        if not xaphuong or not chitiet:
            flash('Vui lòng nhập đầy đủ thông tin xã/phường và địa chỉ chi tiết!', 'danger')
            return redirect(url_for('hokhau_edit', mahokhau=mahokhau))
        
        # Lấy madiachi của hộ khẩu
        query_madiachi = "SELECT madiachi FROM hokhau WHERE mahokhau = %s"
        result = execute_query(query_madiachi, (mahokhau,), fetch_one=True)
        
        if not result:
            flash('Không tìm thấy hộ khẩu!', 'danger')
            return redirect(url_for('hokhau_list'))
        
        madiachi = result[0]
        
        # Update địa chỉ
        update_diachi = """
            UPDATE diachi 
            SET xaphuong = %s, chitiet = %s
            WHERE madiachi = %s
        """
        execute_query(update_diachi, (xaphuong, chitiet, madiachi))
        
        # Update ghi chú hộ khẩu
        update_hokhau = """
            UPDATE hokhau 
            SET ghichu = %s
            WHERE mahokhau = %s
        """
        result = execute_query(update_hokhau, (ghichu or None, mahokhau))
        
        if result:
            flash('Đã cập nhật thông tin hộ khẩu thành công!', 'success')
            return redirect(url_for('hokhau_detail', mahokhau=mahokhau))
        else:
            flash('Có lỗi xảy ra khi cập nhật!', 'danger')
    
    # GET: Load dữ liệu hiện tại
    query = """
        SELECT 
            h.mahokhau,
            d.tinh,
            d.xaphuong,
            d.chitiet,
            h.ngaycap,
            h.ghichu,
            d.madiachi
        FROM hokhau h
        LEFT JOIN diachi d ON h.madiachi = d.madiachi
        WHERE h.mahokhau = %s
    """
    hokhau = execute_query(query, (mahokhau,), fetch_one=True)
    
    if not hokhau:
        flash('Không tìm thấy hộ khẩu!', 'danger')
        return redirect(url_for('hokhau_list'))
    
    # Lấy danh sách thành viên
    query_thanhvien = """
        SELECT 
            n.cccd,
            n.name,
            n.ngaysinh,
            tv.quanhechuho
        FROM thanhvienhokhau tv
        JOIN nguoidung n ON tv.cccd = n.cccd
        WHERE tv.mahokhau = %s AND tv.ngayketthuc IS NULL
        ORDER BY 
            CASE WHEN tv.quanhechuho = 'ChuHo' THEN 0 ELSE 1 END,
            tv.ngaybatdau
    """
    thanhvien = execute_query(query_thanhvien, (mahokhau,), fetch_all=True)
    thanhvien = thanhvien if thanhvien else []
    
    return render_template('hokhau_edit.html', hokhau=hokhau, thanhvien=thanhvien)


@app.route('/hokhau/delete/<int:mahokhau>', methods=['POST'])
@role_required(['CanBo', 'QuanLy'])
@login_required
def hokhau_delete(mahokhau):
    """Xóa hộ khẩu"""
    
    # Lấy thông tin để hiển thị
    query_info = """
        SELECT h.mahokhau, d.chitiet, d.madiachi
        FROM hokhau h
        LEFT JOIN diachi d ON h.madiachi = d.madiachi
        WHERE h.mahokhau = %s
    """
    hokhau_info = execute_query(query_info, (mahokhau,), fetch_one=True)
    
    if not hokhau_info:
        flash('Không tìm thấy hộ khẩu!', 'danger')
        return redirect(url_for('hokhau_list'))
    
    madiachi = hokhau_info[2]
    
    # Xóa theo thứ tự: thanhvienhokhau -> hokhau -> diachi
    connection = get_db_connection()
    if not connection:
        flash('Lỗi kết nối database!', 'danger')
        return redirect(url_for('hokhau_list'))
    
    try:
        cursor = connection.cursor()
        
        # Xóa thành viên hộ khẩu
        cursor.execute("DELETE FROM thanhvienhokhau WHERE mahokhau = %s", (mahokhau,))
        
        # Xóa hộ khẩu
        cursor.execute("DELETE FROM hokhau WHERE mahokhau = %s", (mahokhau,))
        
        # Xóa địa chỉ
        if madiachi:
            cursor.execute("DELETE FROM diachi WHERE madiachi = %s", (madiachi,))
        
        connection.commit()
        cursor.close()
        connection.close()
        
        flash(f'Đã xóa hộ khẩu HK{mahokhau} thành công!', 'success')
        
    except Error as e:
        if connection:
            connection.rollback()
            connection.close()
        flash(f'Có lỗi xảy ra: {str(e)}', 'danger')
    
    return redirect(url_for('hokhau_list'))


@app.route('/thanhvien/chuyen-di/<int:mahokhau>/<string:cccd>', methods=['GET', 'POST'])
@login_required
def thanhvien_chuyen_di(mahokhau, cccd):
    """Chuyển đi/Qua đời thành viên"""
    
    if request.method == 'POST':
        ngaychuyen = request.form.get('ngaychuyen')
        lydochuyen = request.form.get('lydochuyen')
        noichuyenden = request.form.get('noichuyenden')
        ghichu = request.form.get('ghichu')
        
        try:
            # Update ngayketthuc trong thanhvienhokhau
            query = """
                UPDATE thanhvienhokhau 
                SET ngayketthuc = %s,
                    lydochuyen = %s,
                    noichuyenden = %s,
                    ghichu = %s
                WHERE mahokhau = %s AND cccd = %s AND ngayketthuc IS NULL
            """
            execute_query(query, (ngaychuyen, lydochuyen, noichuyenden, ghichu, mahokhau, cccd.strip()))
            
            # Kiểm tra xem hộ còn thành viên không
            query_count = """
                SELECT COUNT(*) FROM thanhvienhokhau 
                WHERE mahokhau = %s AND ngayketthuc IS NULL
            """
            count = execute_query(query_count, (mahokhau,), fetch_one=True)
            
            # Nếu hết thành viên, đánh dấu hộ là trống
            if count and count[0] == 0:
                query_mark_empty = """
                    UPDATE hokhau 
                    SET ghichu = CASE 
                        WHEN ghichu IS NULL OR ghichu = '' THEN '[HỘ TRỐNG - Không còn thành viên]'
                        ELSE ghichu || ' | [HỘ TRỐNG - Không còn thành viên]'
                    END
                    WHERE mahokhau = %s
                """
                execute_query(query_mark_empty, (mahokhau,))
                flash(f'Đã cập nhật thông tin chuyển đi cho thành viên {cccd}. Hộ khẩu đã trống!', 'warning')
            else:
                flash(f'Đã cập nhật thông tin chuyển đi cho thành viên {cccd}!', 'success')
            
            return redirect(url_for('hokhau_detail', mahokhau=mahokhau))
        except Exception as e:
            flash(f'Lỗi khi cập nhật: {str(e)}', 'danger')
    
    # GET - Load thông tin thành viên
    query = """
        SELECT n.cccd, n.name, n.ngaysinh, tv.quanhechuho, tv.ngaybatdau
        FROM nguoidung n
        JOIN thanhvienhokhau tv ON n.cccd = tv.cccd
        WHERE tv.mahokhau = %s AND n.cccd = %s AND tv.ngayketthuc IS NULL
    """
    thanhvien = execute_query(query, (mahokhau, cccd.strip()), fetch_one=True)
    
    if not thanhvien:
        flash('Không tìm thấy thành viên trong hộ khẩu!', 'danger')
        return redirect(url_for('hokhau_detail', mahokhau=mahokhau))
    
    # Lấy thông tin hộ khẩu
    query_hk = """
        SELECT hk.mahokhau, hk.ghichu, dc.chitiet, dc.xaphuong
        FROM hokhau hk
        LEFT JOIN diachi dc ON hk.madiachi = dc.madiachi
        WHERE hk.mahokhau = %s
    """
    hokhau = execute_query(query_hk, (mahokhau,), fetch_one=True)
    
    from datetime import datetime
    return render_template('thanhvien_chuyen_di.html', 
                         thanhvien=thanhvien, 
                         hokhau=hokhau, 
                         mahokhau=mahokhau,
                         now=datetime.now())


@app.route('/hokhau/doi-chu-ho/<int:mahokhau>', methods=['GET', 'POST'])
@role_required(['CanBo', 'QuanLy'])
@login_required
def hokhau_doi_chu_ho(mahokhau):
    """Thay đổi chủ hộ"""
    
    if request.method == 'POST':
        cccd_moi = request.form.get('cccd_moi')
        ngaythaydoi = request.form.get('ngaythaydoi')
        lydothaydoi = request.form.get('lydothaydoi')
        noidung = request.form.get('noidung')
        
        try:
            # Lấy thông tin chủ hộ hiện tại
            query_chuho_cu = """
                SELECT cccd FROM thanhvienhokhau 
                WHERE mahokhau = %s AND quanhechuho = 'Chủ hộ' AND ngayketthuc IS NULL
            """
            chuho_cu = execute_query(query_chuho_cu, (mahokhau,), fetch_one=True)
            
            if not chuho_cu:
                flash('Không tìm thấy chủ hộ hiện tại!', 'danger')
                return redirect(url_for('hokhau_detail', mahokhau=mahokhau))
            
            cccd_cu = chuho_cu[0].strip()
            cccd_moi = cccd_moi.strip()
            
            if cccd_cu == cccd_moi:
                flash('Chủ hộ mới phải khác chủ hộ hiện tại!', 'warning')
                return redirect(url_for('hokhau_doi_chu_ho', mahokhau=mahokhau))
            
            connection = get_db_connection()
            cursor = connection.cursor()
            
            # 1. Cập nhật quan hệ của chủ hộ cũ thành thành viên
            cursor.execute("""
                UPDATE thanhvienhokhau 
                SET quanhechuho = 'Thành viên'
                WHERE mahokhau = %s AND cccd = %s AND ngayketthuc IS NULL
            """, (mahokhau, cccd_cu))
            
            # 2. Cập nhật quan hệ của chủ hộ mới
            cursor.execute("""
                UPDATE thanhvienhokhau 
                SET quanhechuho = 'Chủ hộ'
                WHERE mahokhau = %s AND cccd = %s AND ngayketthuc IS NULL
            """, (mahokhau, cccd_moi))
            
            # 3. Lưu lịch sử thay đổi
            nguoithuchien = session.get('username', 'system')
            cursor.execute("""
                INSERT INTO lichsuthaydoichuho 
                (mahokhau, cccd_cu, cccd_moi, ngaythaydoi, lydothaydoi, noidung, nguoithuchien)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (mahokhau, cccd_cu, cccd_moi, ngaythaydoi, lydothaydoi, noidung, nguoithuchien))
            
            connection.commit()
            cursor.close()
            connection.close()
            
            flash(f'Đã thay đổi chủ hộ thành công!', 'success')
            return redirect(url_for('hokhau_detail', mahokhau=mahokhau))
            
        except Exception as e:
            if connection:
                connection.rollback()
                connection.close()
            flash(f'Lỗi khi thay đổi chủ hộ: {str(e)}', 'danger')
    
    # GET - Load thông tin hộ khẩu và danh sách thành viên
    query_hk = """
        SELECT hk.mahokhau, hk.ghichu, dc.chitiet, dc.xaphuong
        FROM hokhau hk
        LEFT JOIN diachi dc ON hk.madiachi = dc.madiachi
        WHERE hk.mahokhau = %s
    """
    hokhau = execute_query(query_hk, (mahokhau,), fetch_one=True)
    
    if not hokhau:
        flash('Không tìm thấy hộ khẩu!', 'danger')
        return redirect(url_for('hokhau_list'))
    
    # Lấy danh sách thành viên (bao gồm cả chủ hộ hiện tại)
    query_thanhvien = """
        SELECT 
            n.cccd,
            n.name,
            n.ngaysinh,
            n.gioitinh,
            tv.quanhechuho,
            tv.ngaybatdau
        FROM thanhvienhokhau tv
        JOIN nguoidung n ON tv.cccd = n.cccd
        WHERE tv.mahokhau = %s AND tv.ngayketthuc IS NULL
        ORDER BY 
            CASE WHEN tv.quanhechuho = 'Chủ hộ' THEN 0 ELSE 1 END,
            tv.ngaybatdau
    """
    thanhvien = execute_query(query_thanhvien, (mahokhau,), fetch_all=True)
    
    from datetime import datetime
    return render_template('hokhau_doi_chu_ho.html', 
                         hokhau=hokhau, 
                         thanhvien=thanhvien,
                         mahokhau=mahokhau,
                         now=datetime.now())


@role_required(['CanBo', 'QuanLy'])
@app.route('/hokhau/tach-ho/<int:mahokhau>', methods=['GET', 'POST'])
@login_required
def hokhau_tach_ho(mahokhau):
    """Tách hộ khẩu - tạo hộ mới từ hộ hiện tại"""
    
    if request.method == 'POST':
        thanhvien_tach = request.form.getlist('thanhvien_tach')  # List of CCCD
        cccd_chuho_moi = request.form.get('cccd_chuho_moi')
        ngaytach = request.form.get('ngaytach')
        
        # Thông tin địa chỉ mới
        xaphuong_moi = request.form.get('xaphuong_moi')
        chitiet_moi = request.form.get('chitiet_moi')
        ghichu_moi = request.form.get('ghichu_moi')
        lydotach = request.form.get('lydotach')
        
        # Validation
        if not thanhvien_tach or len(thanhvien_tach) == 0:
            flash('Vui lòng chọn ít nhất 1 thành viên để tách!', 'warning')
            return redirect(url_for('hokhau_tach_ho', mahokhau=mahokhau))
        
        if cccd_chuho_moi not in thanhvien_tach:
            flash('Chủ hộ mới phải nằm trong danh sách thành viên được tách!', 'warning')
            return redirect(url_for('hokhau_tach_ho', mahokhau=mahokhau))
        
        try:
            connection = get_db_connection()
            cursor = connection.cursor()
            
            # 1. Tạo địa chỉ mới
            cursor.execute("""
                INSERT INTO diachi (xaphuong, chitiet)
                VALUES (%s, %s)
                RETURNING madiachi
            """, (xaphuong_moi, chitiet_moi))
            madiachi_moi = cursor.fetchone()[0]
            
            # 2. Tạo hộ khẩu mới
            cursor.execute("""
                INSERT INTO hokhau (madiachi, ghichu)
                VALUES (%s, %s)
                RETURNING mahokhau
            """, (madiachi_moi, ghichu_moi or f'Tách từ HK{mahokhau}. Lý do: {lydotach}'))
            mahokhau_moi = cursor.fetchone()[0]
            
            # 3. Cập nhật ngayketthuc cho các thành viên tách ra trong hộ cũ
            for cccd in thanhvien_tach:
                cccd = cccd.strip()
                cursor.execute("""
                    UPDATE thanhvienhokhau
                    SET ngayketthuc = %s,
                        lydochuyen = %s,
                        ghichu = %s
                    WHERE mahokhau = %s AND cccd = %s AND ngayketthuc IS NULL
                """, (ngaytach, 'Tách hộ', f'Tách sang HK{mahokhau_moi}', mahokhau, cccd))
            
            # 4. Thêm các thành viên vào hộ mới
            for cccd in thanhvien_tach:
                cccd = cccd.strip()
                quanhe = 'Chủ hộ' if cccd == cccd_chuho_moi.strip() else 'Thành viên'
                
                cursor.execute("""
                    INSERT INTO thanhvienhokhau (mahokhau, cccd, quanhechuho, ngaybatdau)
                    VALUES (%s, %s, %s, %s)
                """, (mahokhau_moi, cccd, quanhe, ngaytach))
            
            connection.commit()
            cursor.close()
            connection.close()
            
            flash(f'Tách hộ thành công! Hộ khẩu mới: HK{mahokhau_moi}', 'success')
            return redirect(url_for('hokhau_detail', mahokhau=mahokhau_moi))
            
        except Exception as e:
            if connection:
                connection.rollback()
                connection.close()
            flash(f'Lỗi khi tách hộ: {str(e)}', 'danger')
            return redirect(url_for('hokhau_tach_ho', mahokhau=mahokhau))
    
    # GET - Load thông tin hộ khẩu và danh sách thành viên
    query_hk = """
        SELECT hk.mahokhau, hk.ghichu, dc.chitiet, dc.xaphuong
        FROM hokhau hk
        LEFT JOIN diachi dc ON hk.madiachi = dc.madiachi
        WHERE hk.mahokhau = %s
    """
    hokhau = execute_query(query_hk, (mahokhau,), fetch_one=True)
    
    if not hokhau:
        flash('Không tìm thấy hộ khẩu!', 'danger')
        return redirect(url_for('hokhau_list'))
    
    # Lấy danh sách thành viên (không bao gồm chủ hộ - không cho phép tách chủ hộ)
    query_thanhvien = """
        SELECT 
            n.cccd,
            n.name,
            n.ngaysinh,
            n.gioitinh,
            tv.quanhechuho,
            tv.ngaybatdau
        FROM thanhvienhokhau tv
        JOIN nguoidung n ON tv.cccd = n.cccd
        WHERE tv.mahokhau = %s AND tv.ngayketthuc IS NULL
        ORDER BY 
            CASE WHEN tv.quanhechuho = 'Chủ hộ' THEN 0 ELSE 1 END,
            tv.ngaybatdau
    """
    thanhvien = execute_query(query_thanhvien, (mahokhau,), fetch_all=True)
    
    # Đếm số thành viên (không bao gồm chủ hộ)
    thanhvien_khong_chuho = [tv for tv in thanhvien if tv[4] != 'Chủ hộ']
    
    if len(thanhvien_khong_chuho) == 0:
        flash('Hộ khẩu chỉ có chủ hộ, không thể tách!', 'warning')
        return redirect(url_for('hokhau_detail', mahokhau=mahokhau))
    
    from datetime import datetime
    return render_template('hokhau_tach_ho.html', 
                         hokhau=hokhau, 
                         thanhvien=thanhvien,
                         mahokhau=mahokhau,
                         now=datetime.now())


@app.route('/tam-vang/add', methods=['GET', 'POST'])
@login_required
def tam_vang_add():
    """Cấp giấy tạm vắng"""
    
    user_role = session['user'].get('vaitro')
    user_cccd = session['user'].get('cccd')
    
    if request.method == 'POST':
        cccd = request.form.get('cccd', '').strip()
        ngaybatdau = request.form.get('ngaybatdau')
        ngayketthuc = request.form.get('ngayketthuc')
        lydo = request.form.get('lydo')
        noiden = request.form.get('noiden')
        
        # Nếu là Người dân, chỉ đăng ký cho mình
        if user_role == 'NguoiDan' and cccd != user_cccd:
            flash('Bạn chỉ được đăng ký tạm vắng cho chính mình!', 'danger')
            return redirect(url_for('tam_vang_add'))
        
        if not cccd or not ngaybatdau:
            flash('Vui lòng điền đầy đủ thông tin bắt buộc!', 'warning')
            return redirect(url_for('tam_vang_add'))
        
        # Lấy địa chỉ thường trú hiện tại
        query_diachi = """
            SELECT dc.madiachi, dc.xaphuong, dc.chitiet
            FROM diachinguoidung dcnd
            JOIN diachi dc ON dcnd.madiachi = dc.madiachi
            WHERE dcnd.cccd = %s AND dcnd.loaidiachi = 'CuTru' 
            AND dcnd.thoidiemketthuc IS NULL
            LIMIT 1
        """
        diachi_cutru = execute_query(query_diachi, (cccd,), fetch_one=True)
        
        if not diachi_cutru:
            flash('Không tìm thấy địa chỉ cư trú của người này!', 'warning')
            return redirect(url_for('tam_vang_add'))
        
        # Kiểm tra đã có tạm vắng chưa
        query_check = """
            SELECT 1 FROM diachinguoidung 
            WHERE cccd = %s AND loaidiachi = 'TamVang' 
            AND (thoidiemketthuc IS NULL OR thoidiemketthuc >= CURRENT_DATE)
        """
        existing = execute_query(query_check, (cccd,), fetch_one=True)
        if existing:
            flash('Người này đang có giấy tạm vắng hiệu lực!', 'warning')
            return redirect(url_for('tam_vang_add'))
        
        madiachi_cutru = diachi_cutru[0]
        
        try:
            connection = get_db_connection()
            cursor = connection.cursor()
            
            # Tạo địa chỉ tạm vắng (nơi đến)
            cursor.execute("""
                INSERT INTO diachi (chitiet, xaphuong)
                VALUES (%s, %s)
                RETURNING madiachi
            """, (noiden or 'Chưa xác định', lydo or 'Chưa rõ'))
            madiachi_tamvang = cursor.fetchone()[0]
            
            # Insert vào diachinguoidung với loaidiachi='TamVang'
            cursor.execute("""
                INSERT INTO diachinguoidung (madiachi, cccd, loaidiachi, thoidiemxacnhan, thoidiemketthuc)
                VALUES (%s, %s, 'TamVang', %s, %s)
            """, (madiachi_tamvang, cccd, ngaybatdau, ngayketthuc or None))
            
            connection.commit()
            cursor.close()
            connection.close()
            
            flash(f'Đã cấp giấy tạm vắng thành công cho CCCD {cccd}!', 'success')
            return redirect(url_for('tam_vang_tru_list'))
            
        except Exception as e:
            if connection:
                connection.rollback()
                connection.close()
            flash(f'Lỗi khi cấp giấy tạm vắng: {str(e)}', 'danger')
    
    # GET - Load danh sách người dân để chọn
    query_nguoidung = """
        SELECT cccd, name, ngaysinh, gioitinh
        FROM nguoidung
    """
    params = []
    
    # Nếu là Người dân, chỉ xem mình
    if user_role == 'NguoiDan':
        query_nguoidung += " WHERE cccd = %s"
        params.append(user_cccd)
    else:
        # Cán bộ/Quản lý xem người trong hộ khẩu
        query_nguoidung += """
            WHERE cccd IN (
                SELECT DISTINCT tv.cccd 
                FROM thanhvienhokhau tv
                WHERE tv.ngayketthuc IS NULL
            )
        """
    
    query_nguoidung += " ORDER BY name"
    nguoidung_list = execute_query(query_nguoidung, tuple(params), fetch_all=True)
    
    from datetime import datetime
    return render_template('tam_vang_add.html', 
                         nguoidung_list=nguoidung_list,
                         now=datetime.now(),
                         user_role=user_role,
                         user_cccd=user_cccd)


@app.route('/tam-vang-tru/pdf/<cccd>/<loai>')
@login_required
def tam_vang_tru_pdf(cccd, loai):
    connection = get_db_connection()
    cursor = connection.cursor()

    cursor.execute("""
        SELECT n.name, n.ngaysinh, n.sdt, n.dantoc, n.gioitinh
        FROM nguoidung n
        WHERE n.cccd = %s
    """, (cccd,))
    row = cursor.fetchone()

    if not row:
        return "Không tìm thấy công dân", 404

    name, ngaysinh, sdt, dantoc, gioitinh = row

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # ===== HEADER =====
    c.setFont("NotoSans", 12)
    c.drawCentredString(width / 2, height - 2*cm, "CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM")
    c.drawCentredString(width / 2, height - 2.8*cm, "Độc lập - Tự do - Hạnh phúc")
    c.line(5*cm, height - 3.1*cm, width - 5*cm, height - 3.1*cm)

    # ===== TITLE =====
    c.setFont("NotoSans", 16)
    c.drawCentredString(width / 2, height - 4.5*cm, f"GIẤY {loai.upper()}")

    # ===== CONTENT =====
    c.setFont("NotoSans", 12)
    y = height - 6*cm
    line_gap = 18

    c.drawString(3*cm, y, f"Họ và tên: {name}")
    y -= line_gap

    c.drawString(3*cm, y, f"Ngày sinh: {ngaysinh.strftime('%d/%m/%Y') if ngaysinh else ''}")
    y -= line_gap

    c.drawString(3*cm, y, f"Giới tính: {gioitinh}")
    y -= line_gap

    c.drawString(3*cm, y, f"Dân tộc: {dantoc}")
    y -= line_gap

    c.drawString(3*cm, y, f"Số CCCD: {cccd}")
    y -= line_gap

    c.drawString(3*cm, y, f"Số điện thoại: {sdt}")
    y -= line_gap * 2

    c.drawString(3*cm, y, f"Lý do {loai.lower()}: .........................................................")
    y -= line_gap

    c.drawString(3*cm, y, f"Thời gian: Từ ngày ....../....../...... đến ngày ....../....../......")
    y -= line_gap * 2

    c.drawString(3*cm, y, "Công dân cam kết những thông tin trên là đúng sự thật.")
    y -= line_gap * 2

    # ===== SIGNATURE =====
    c.drawString(width - 8*cm, y, "Ngày ..... tháng ..... năm .....")
    y -= line_gap * 2

    c.drawString(3*cm, y, "XÁC NHẬN CỦA CƠ QUAN")
    c.drawString(width - 7*cm, y, "NGƯỜI KHAI")
    y -= line_gap * 3

    c.drawString(width - 7*cm, y, name)

    # ===== FINISH =====
    c.showPage()
    c.save()
    buffer.seek(0)

    return Response(
        buffer,
        mimetype="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=giay_{loai}_{cccd}.pdf"
        }
    )
@app.route('/tam-tru/add', methods=['GET', 'POST'])
@login_required
def tam_tru_add():
    """Cấp giấy tạm trú"""
    
    user_role = session['user'].get('vaitro')
    user_cccd = session['user'].get('cccd')
    
    if request.method == 'POST':
        cccd = request.form.get('cccd', '').strip()
        ngaybatdau = request.form.get('ngaybatdau')
        ngayketthuc = request.form.get('ngayketthuc')
        lydo = request.form.get('lydo')
        diachi_tamtru = request.form.get('diachi_tamtru')
        xaphuong = request.form.get('xaphuong')
        
        if not cccd or not ngaybatdau or not diachi_tamtru or not xaphuong:
            flash('Vui lòng điền đầy đủ thông tin bắt buộc!', 'warning')
            return redirect(url_for('tam_tru_add'))
        
        # Kiểm tra đã có tạm trú chưa
        query_check = """
            SELECT 1 FROM diachinguoidung 
            WHERE cccd = %s AND loaidiachi = 'TamTru' 
            AND (thoidiemketthuc IS NULL OR thoidiemketthuc >= CURRENT_DATE)
        """
        existing = execute_query(query_check, (cccd,), fetch_one=True)
        if existing:
            flash('Người này đang có giấy tạm trú hiệu lực!', 'warning')
            return redirect(url_for('tam_tru_add'))
        
        try:
            connection = get_db_connection()
            cursor = connection.cursor()
            
            # Tạo địa chỉ tạm trú
            cursor.execute("""
                INSERT INTO diachi (chitiet, xaphuong)
                VALUES (%s, %s)
                RETURNING madiachi
            """, (diachi_tamtru, xaphuong))
            madiachi_tamtru = cursor.fetchone()[0]
            
            # Insert vào diachinguoidung với loaidiachi='TamTru'
            cursor.execute("""
                INSERT INTO diachinguoidung (madiachi, cccd, loaidiachi, thoidiemxacnhan, thoidiemketthuc)
                VALUES (%s, %s, 'TamTru', %s, %s)
            """, (madiachi_tamtru, cccd, ngaybatdau, ngayketthuc or None))
            
            connection.commit()
            cursor.close()
            connection.close()
            
            flash(f'Đã cấp giấy tạm trú thành công cho CCCD {cccd}!', 'success')
            return redirect(url_for('tam_vang_tru_list'))
            
        except Exception as e:
            if connection:
                connection.rollback()
                connection.close()
            flash(f'Lỗi khi cấp giấy tạm trú: {str(e)}', 'danger')
    
    # GET - Load danh sách người dân
    query_nguoidung = """
        SELECT cccd, name, ngaysinh, gioitinh
        FROM nguoidung
    """
    params = []
    
    # Chỉ hiển thị người trong hộ khẩu hiện tại
    query_nguoidung += """
        WHERE cccd IN (
            SELECT DISTINCT tv.cccd 
            FROM thanhvienhokhau tv
            WHERE tv.ngayketthuc IS NULL
        )
    """
    
    query_nguoidung += " ORDER BY name"
    nguoidung_list = execute_query(query_nguoidung, tuple(params), fetch_all=True)
    
    from datetime import datetime
    return render_template('tam_tru_add.html', 
                         nguoidung_list=nguoidung_list,
                         now=datetime.now())


@app.route('/tam-vang-tru/<string:cccd>/<string:loai>/gia-han', methods=['GET', 'POST'])
@login_required
@role_required(['CanBo', 'QuanLy'])
def tam_vang_tru_gia_han(cccd, loai):
    """Gia hạn tạm vắng/tạm trú"""
    
    if loai not in ['TamVang', 'TamTru']:
        flash('Loại không hợp lệ!', 'danger')
        return redirect(url_for('tam_vang_tru_list'))
    
    if request.method == 'POST':
        ngayketthuc_moi = request.form.get('ngayketthuc_moi')
        ghichu = request.form.get('ghichu')
        
        if not ngayketthuc_moi:
            flash('Vui lòng nhập ngày kết thúc mới!', 'warning')
            return redirect(url_for('tam_vang_tru_gia_han', cccd=cccd, loai=loai))
        
        try:
            query = """
                UPDATE diachinguoidung 
                SET thoidiemketthuc = %s
                WHERE cccd = %s AND loaidiachi = %s 
                AND (thoidiemketthuc IS NULL OR thoidiemketthuc >= CURRENT_DATE)
            """
            execute_query(query, (ngayketthuc_moi, cccd, loai))
            
            flash(f'Đã gia hạn {"tạm vắng" if loai == "TamVang" else "tạm trú"} thành công!', 'success')
            return redirect(url_for('tam_vang_tru_list'))
        except Exception as e:
            flash(f'Lỗi khi gia hạn: {str(e)}', 'danger')
    
    # GET - Load thông tin hiện tại
    query = """
        SELECT 
            n.cccd, n.name, n.ngaysinh,
            dcnd.loaidiachi, dcnd.thoidiemxacnhan, dcnd.thoidiemketthuc,
            dc.chitiet, dc.xaphuong
        FROM diachinguoidung dcnd
        JOIN nguoidung n ON dcnd.cccd = n.cccd
        JOIN diachi dc ON dcnd.madiachi = dc.madiachi
        WHERE dcnd.cccd = %s AND dcnd.loaidiachi = %s
        AND (dcnd.thoidiemketthuc IS NULL OR dcnd.thoidiemketthuc >= CURRENT_DATE)
        LIMIT 1
    """
    record = execute_query(query, (cccd, loai), fetch_one=True)
    
    if not record:
        flash(f'Không tìm thấy {"tạm vắng" if loai == "TamVang" else "tạm trú"} đang hiệu lực!', 'warning')
        return redirect(url_for('tam_vang_tru_list'))
    
    from datetime import datetime
    return render_template('tam_vang_tru_gia_han.html', 
                         record=record,
                         loai=loai,
                         now=datetime.now())


@app.route('/tam-vang-tru/<string:cccd>/<string:loai>/ket-thuc', methods=['POST'])
@login_required
@role_required(['CanBo', 'QuanLy'])
def tam_vang_tru_ket_thuc(cccd, loai):
    """Kết thúc sớm tạm vắng/tạm trú"""
    
    if loai not in ['TamVang', 'TamTru']:
        flash('Loại không hợp lệ!', 'danger')
        return redirect(url_for('tam_vang_tru_list'))
    
    ngayketthuc = request.form.get('ngayketthuc')
    
    if not ngayketthuc:
        flash('Vui lòng nhập ngày kết thúc!', 'warning')
        return redirect(url_for('tam_vang_tru_list'))
    
    try:
        query = """
            UPDATE diachinguoidung 
            SET thoidiemketthuc = %s
            WHERE cccd = %s AND loaidiachi = %s 
            AND (thoidiemketthuc IS NULL OR thoidiemketthuc >= CURRENT_DATE)
        """
        execute_query(query, (ngayketthuc, cccd, loai))
        
        flash(f'Đã kết thúc {"tạm vắng" if loai == "TamVang" else "tạm trú"} thành công!', 'success')
    except Exception as e:
        flash(f'Lỗi khi kết thúc: {str(e)}', 'danger')
    
    return redirect(url_for('tam_vang_tru_list'))


@app.route('/tam-vang-tru')
@login_required
def tam_vang_tru_list():
    """Danh sách tạm vắng/tạm trú"""
    
    user_role = session['user'].get('vaitro')
    user_cccd = session['user'].get('cccd')
    
    # Lấy filter params
    loai = request.args.get('loai', '')  # TamVang, TamTru, hoặc tất cả
    xaphuong = request.args.get('xaphuong', '')
    trangthai = request.args.get('trangthai', '')  # conhieuluc, hethang
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    # Build query
    query = """
        SELECT 
            n.cccd,
            n.name,
            n.ngaysinh,
            dcnd.loaidiachi,
            dcnd.thoidiemxacnhan,
            dcnd.thoidiemketthuc,
            dc.chitiet,
            dc.xaphuong
        FROM diachinguoidung dcnd
        JOIN nguoidung n ON dcnd.cccd = n.cccd
        JOIN diachi dc ON dcnd.madiachi = dc.madiachi
        WHERE dcnd.loaidiachi IN ('TamVang', 'TamTru')
    """
    
    params = []
    
    # Nếu là Người dân, chỉ xem của mình
    if user_role == 'NguoiDan':
        query += " AND dcnd.cccd = %s"
        params.append(user_cccd)
    
    if loai:
        query += " AND dcnd.loaidiachi = %s"
        params.append(loai)
    
    if xaphuong:
        query += " AND dc.xaphuong ILIKE %s"
        params.append(f'%{xaphuong}%')
    
    # Lọc trạng thái
    if trangthai == 'conhieuluc':
        query += " AND (dcnd.thoidiemketthuc IS NULL OR dcnd.thoidiemketthuc >= CURRENT_DATE)"
    elif trangthai == 'hethang':
        query += " AND dcnd.thoidiemketthuc < CURRENT_DATE"
    
    query += " ORDER BY dcnd.thoidiemxacnhan DESC"
    
    # Count total
    count_query = f"SELECT COUNT(*) FROM ({query}) AS subquery"
    total_result = execute_query(count_query, tuple(params), fetch_one=True)
    total = total_result[0] if total_result else 0
    
    # Query statistics (tối ưu hơn dùng Jinja filter)
    stats_query = """
        SELECT 
            dcnd.loaidiachi,
            COUNT(*) as so_luong,
            COUNT(CASE WHEN dcnd.thoidiemketthuc IS NULL OR dcnd.thoidiemketthuc >= CURRENT_DATE THEN 1 END) as con_hieu_luc,
            COUNT(CASE WHEN dcnd.thoidiemketthuc < CURRENT_DATE THEN 1 END) as het_han
        FROM diachinguoidung dcnd
        WHERE dcnd.loaidiachi IN ('TamVang', 'TamTru')
    """
    stats_params = []
    if user_role == 'NguoiDan':
        stats_query += " AND dcnd.cccd = %s"
        stats_params.append(user_cccd)
    stats_query += " GROUP BY dcnd.loaidiachi"
    
    stats_result = execute_query(stats_query, tuple(stats_params), fetch_all=True)
    stats = {'TamVang': 0, 'TamTru': 0, 'TamVang_hieuluc': 0, 'TamTru_hieuluc': 0}
    if stats_result:
        for row in stats_result:
            if row[0] == 'TamVang':
                stats['TamVang'] = row[1]
                stats['TamVang_hieuluc'] = row[2]
            elif row[0] == 'TamTru':
                stats['TamTru'] = row[1]
                stats['TamTru_hieuluc'] = row[2]
    
    # Pagination
    offset = (page - 1) * per_page
    query += f" LIMIT {per_page} OFFSET {offset}"
    
    records = execute_query(query, tuple(params), fetch_all=True)
    records = records if records else []
    
    total_pages = (total + per_page - 1) // per_page if total > 0 else 1
    
    from datetime import datetime
    return render_template('tam_vang_tru.html',
                         records=records,
                         page=page,
                         total_pages=total_pages,
                         total=total,
                         loai=loai,
                         xaphuong=xaphuong,
                         trangthai=trangthai,
                         stats=stats,
                         now=datetime.now())


@app.route('/cu-tru')
@login_required
def cu_tru_list():
    """Danh sách cư trú với thống kê, tìm kiếm và lọc"""
    
    user_role = session['user'].get('vaitro')
    user_cccd = session['user'].get('cccd')
    
    # Lấy tham số
    search = request.args.get('search', '').strip()
    xaphuong = request.args.get('xaphuong', '').strip()
    hieuluc = request.args.get('hieuluc', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    # Build query
    query = """
        SELECT 
            n.cccd,
            n.name,
            n.ngaysinh,
            n.gioitinh,
            dcnd.thoidiemxacnhan,
            dcnd.thoidiemketthuc,
            dc.chitiet,
            dc.xaphuong,
            dc.tinh
        FROM diachinguoidung dcnd
        JOIN nguoidung n ON dcnd.cccd = n.cccd
        JOIN diachi dc ON dcnd.madiachi = dc.madiachi
        WHERE dcnd.loaidiachi = 'CuTru'
    """
    
    params = []
    
    # Filter by status - hieuluc (hiệu lực)
    if hieuluc == 'con':
        query += " AND (dcnd.thoidiemketthuc IS NULL OR dcnd.thoidiemketthuc >= CURRENT_DATE)"
    elif hieuluc == 'het':
        query += " AND dcnd.thoidiemketthuc < CURRENT_DATE"
    else:
        # Default: show all
        pass
    
    # Nếu là Người dân, chỉ xem của mình
    if user_role == 'NguoiDan':
        query += " AND dcnd.cccd = %s"
        params.append(user_cccd)
    
    if search:
        query += " AND (n.name ILIKE %s OR n.cccd ILIKE %s)"
        params.extend([f'%{search}%', f'%{search}%'])
    
    if xaphuong:
        query += " AND dc.xaphuong = %s"
        params.append(xaphuong)
    
    # Count total
    count_query = f"SELECT COUNT(*) FROM ({query}) AS subquery"
    total_result = execute_query(count_query, tuple(params), fetch_one=True)
    total = total_result[0] if total_result else 0
    
    # Statistics - Tổng số và số còn hiệu lực (thoidiemketthuc IS NULL OR >= CURRENT_DATE)
    stats_query = """
        SELECT 
            COUNT(*) as tong_cutru,
            COUNT(CASE WHEN dcnd.thoidiemketthuc IS NULL OR dcnd.thoidiemketthuc >= CURRENT_DATE THEN 1 END) as con_hieu_luc,
            COUNT(CASE WHEN dcnd.thoidiemketthuc < CURRENT_DATE THEN 1 END) as het_hieu_luc
        FROM diachinguoidung dcnd
        JOIN nguoidung n ON dcnd.cccd = n.cccd
        JOIN diachi dc ON dcnd.madiachi = dc.madiachi
        WHERE dcnd.loaidiachi = 'CuTru'
    """
    stats_params = []
    if user_role == 'NguoiDan':
        stats_query += " AND dcnd.cccd = %s"
        stats_params.append(user_cccd)
    if search:
        stats_query += " AND (n.name ILIKE %s OR n.cccd ILIKE %s)"
        stats_params.extend([f'%{search}%', f'%{search}%'])
    if xaphuong:
        stats_query += " AND dc.xaphuong = %s"
        stats_params.append(xaphuong)
    
    stats_result = execute_query(stats_query, tuple(stats_params), fetch_one=True)
    stats = {
        'tong_cutru': stats_result[0] if stats_result else 0,
        'con_hieu_luc': stats_result[1] if stats_result else 0,
        'het_hieu_luc': stats_result[2] if stats_result else 0
    }
    
    # Pagination
    query += " ORDER BY dcnd.thoidiemxacnhan DESC"
    offset = (page - 1) * per_page
    query += f" LIMIT {per_page} OFFSET {offset}"
    
    records = execute_query(query, tuple(params), fetch_all=True)
    records = records if records else []
    
    total_pages = (total + per_page - 1) // per_page if total > 0 else 1
    
    from datetime import datetime
    return render_template('cu_tru.html',
                         records=records,
                         page=page,
                         total_pages=total_pages,
                         total=total,
                         stats=stats,
                         search=search,
                         xaphuong=xaphuong,
                         hieuluc=hieuluc,
                         now=datetime.now())


@app.route('/cu-tru/add', methods=['GET', 'POST'])
@login_required
def cu_tru_add():
    """Đăng ký cư trú mới"""
    
    user_role = session['user'].get('vaitro')
    user_cccd = session['user'].get('cccd')
    
    if request.method == 'POST':
        cccd = request.form.get('cccd', '').strip()
        ngaybatdau = request.form.get('ngaybatdau', '').strip()
        tinh = request.form.get('tinh', '').strip()
        xaphuong = request.form.get('xaphuong', '').strip()
        diachi_chitiet = request.form.get('diachi_chitiet', '').strip()
        ghichu = request.form.get('ghichu', '').strip()
        
        # Validation
        if not cccd or not ngaybatdau or not tinh or not xaphuong or not diachi_chitiet:
            flash('Vui lòng điền đầy đủ thông tin bắt buộc!', 'danger')
            return redirect(url_for('cu_tru_add'))
        
        # Nếu là Người dân, chỉ được đăng ký cho chính mình
        if user_role == 'NguoiDan' and cccd != user_cccd:
            flash('Bạn chỉ được đăng ký cư trú cho chính mình!', 'danger')
            return redirect(url_for('cu_tru_add'))
        
        # Kiểm tra người dùng có tồn tại không
        check_user = execute_query("SELECT cccd FROM nguoidung WHERE cccd = %s", (cccd,), fetch_one=True)
        if not check_user:
            flash('Không tìm thấy người dùng với CCCD này!', 'danger')
            return redirect(url_for('cu_tru_add'))
        
        # Kiểm tra xem đã có cư trú active chưa
        check_duplicate = execute_query(
            "SELECT 1 FROM diachinguoidung WHERE cccd = %s AND loaidiachi = 'CuTru' AND (thoidiemketthuc IS NULL OR thoidiemketthuc >= CURRENT_DATE)",
            (cccd,),
            fetch_one=True
        )
        if check_duplicate:
            flash('Người dùng này đã có địa chỉ cư trú đang hiệu lực! Vui lòng chỉnh sửa hoặc kết thúc trước.', 'warning')
            return redirect(url_for('cu_tru_list'))
        
        connection = get_db_connection()
        if not connection:
            flash('Lỗi kết nối database!', 'danger')
            return redirect(url_for('cu_tru_add'))
        
        try:
            cursor = connection.cursor()
            
            # Tạo địa chỉ mới
            cursor.execute(
                "INSERT INTO diachi (tinh, xaphuong, chitiet) VALUES (%s, %s, %s) RETURNING madiachi",
                (tinh, xaphuong, diachi_chitiet)
            )
            madiachi = cursor.fetchone()[0]
            
            # Tạo bản ghi cư trú
            cursor.execute(
                """INSERT INTO diachinguoidung (madiachi, cccd, loaidiachi, thoidiemxacnhan, ghichu) 
                   VALUES (%s, %s, 'CuTru', %s, %s)""",
                (madiachi, cccd, ngaybatdau, ghichu)
            )
            
            connection.commit()
            cursor.close()
            connection.close()
            
            flash('Đăng ký cư trú thành công!', 'success')
            return redirect(url_for('cu_tru_list'))
            
        except Error as e:
            connection.rollback()
            cursor.close()
            connection.close()
            flash(f'Lỗi khi đăng ký cư trú: {str(e)}', 'danger')
            return redirect(url_for('cu_tru_add'))
    
    # GET - Lấy danh sách người dùng
    if user_role == 'NguoiDan':
        # Người dân chỉ thấy chính mình
        query_nguoidung = "SELECT cccd, name, ngaysinh FROM nguoidung WHERE cccd = %s"
        nguoidung_list = execute_query(query_nguoidung, (user_cccd,), fetch_all=True)
    else:
        # Cán bộ/Quản lý xem tất cả người dùng đang active trong hộ khẩu
        query_nguoidung = """
            SELECT DISTINCT n.cccd, n.name, n.ngaysinh
            FROM nguoidung n
            WHERE n.cccd IN (
                SELECT tv.cccd FROM thanhvienhokhau tv WHERE tv.ngayketthuc IS NULL
            )
            ORDER BY n.name
        """
        nguoidung_list = execute_query(query_nguoidung, fetch_all=True)
    
    nguoidung_list = nguoidung_list if nguoidung_list else []
    
    from datetime import datetime
    return render_template('cu_tru_add.html', nguoidung_list=nguoidung_list, now=datetime.now())


@app.route('/cu-tru/edit/<string:cccd>', methods=['GET', 'POST'])
@login_required
def cu_tru_edit(cccd):
    """Chỉnh sửa/Thay đổi địa chỉ cư trú"""
    
    user_role = session['user'].get('vaitro')
    user_cccd = session['user'].get('cccd')
    
    # Kiểm tra quyền
    if user_role == 'NguoiDan' and cccd != user_cccd:
        flash('Bạn chỉ được chỉnh sửa cư trú của chính mình!', 'danger')
        return redirect(url_for('cu_tru_list'))
    
    if request.method == 'POST':
        ngayketthuc_cu = request.form.get('ngayketthuc_cu', '').strip()
        ngaybatdau_moi = request.form.get('ngaybatdau_moi', '').strip()
        tinh = request.form.get('tinh', '').strip()
        xaphuong = request.form.get('xaphuong', '').strip()
        diachi_chitiet = request.form.get('diachi_chitiet', '').strip()
        ghichu = request.form.get('ghichu', '').strip()
        
        # Validation
        if not ngayketthuc_cu or not ngaybatdau_moi or not tinh or not xaphuong or not diachi_chitiet:
            flash('Vui lòng điền đầy đủ thông tin bắt buộc!', 'danger')
            return redirect(url_for('cu_tru_edit', cccd=cccd))
        
        connection = get_db_connection()
        if not connection:
            flash('Lỗi kết nối database!', 'danger')
            return redirect(url_for('cu_tru_edit', cccd=cccd))
        
        try:
            cursor = connection.cursor()
            
            # Kết thúc địa chỉ cũ
            cursor.execute(
                """UPDATE diachinguoidung 
                   SET thoidiemketthuc = %s 
                   WHERE cccd = %s AND loaidiachi = 'CuTru' 
                   AND (thoidiemketthuc IS NULL OR thoidiemketthuc >= CURRENT_DATE)""",
                (ngayketthuc_cu, cccd)
            )
            
            # Tạo địa chỉ mới
            cursor.execute(
                "INSERT INTO diachi (tinh, xaphuong, chitiet) VALUES (%s, %s, %s) RETURNING madiachi",
                (tinh, xaphuong, diachi_chitiet)
            )
            madiachi = cursor.fetchone()[0]
            
            # Tạo bản ghi cư trú mới
            cursor.execute(
                """INSERT INTO diachinguoidung (madiachi, cccd, loaidiachi, thoidiemxacnhan, ghichu) 
                   VALUES (%s, %s, 'CuTru', %s, %s)""",
                (madiachi, cccd, ngaybatdau_moi, ghichu)
            )
            
            connection.commit()
            cursor.close()
            connection.close()
            
            flash('Thay đổi địa chỉ cư trú thành công!', 'success')
            return redirect(url_for('cu_tru_list'))
            
        except Error as e:
            connection.rollback()
            cursor.close()
            connection.close()
            flash(f'Lỗi khi thay đổi cư trú: {str(e)}', 'danger')
            return redirect(url_for('cu_tru_edit', cccd=cccd))
    
    # GET - Lấy thông tin cư trú hiện tại
    query = """
        SELECT 
            n.cccd,
            n.name,
            n.ngaysinh,
            dcnd.thoidiemxacnhan,
            dc.chitiet,
            dc.xaphuong,
            dc.tinh,
            dcnd.ghichu
        FROM diachinguoidung dcnd
        JOIN nguoidung n ON dcnd.cccd = n.cccd
        JOIN diachi dc ON dcnd.madiachi = dc.madiachi
        WHERE dcnd.cccd = %s 
        AND dcnd.loaidiachi = 'CuTru'
        AND (dcnd.thoidiemketthuc IS NULL OR dcnd.thoidiemketthuc >= CURRENT_DATE)
    """
    record = execute_query(query, (cccd,), fetch_one=True)
    
    if not record:
        # Kiểm tra xem có bản ghi cư trú nào không (kể cả hết hạn)
        check_query = """
            SELECT COUNT(*) FROM diachinguoidung 
            WHERE cccd = %s AND loaidiachi = 'CuTru'
        """
        has_any_record = execute_query(check_query, (cccd,), fetch_one=True)
        
        if has_any_record and has_any_record[0] > 0:
            flash('Tất cả bản ghi cư trú của người này đã hết hạn! Vui lòng đăng ký cư trú mới.', 'warning')
        else:
            flash('Người này chưa có thông tin cư trú! Vui lòng đăng ký cư trú trước.', 'info')
        return redirect(url_for('cu_tru_list'))
    
    from datetime import datetime
    return render_template('cu_tru_edit.html', record=record, now=datetime.now())


@role_required(['CanBo', 'QuanLy'])
@app.route('/hokhau/<int:mahokhau>/lich-su')
@login_required
def hokhau_lich_su(mahokhau):
    """Lịch sử biến động nhân khẩu của hộ"""
    
    # Lấy thông tin hộ khẩu
    query_hk = """
        SELECT hk.mahokhau, hk.ghichu, dc.chitiet, dc.xaphuong, hk.ngaycap
        FROM hokhau hk
        LEFT JOIN diachi dc ON hk.madiachi = dc.madiachi
        WHERE hk.mahokhau = %s
    """
    hokhau = execute_query(query_hk, (mahokhau,), fetch_one=True)
    
    if not hokhau:
        flash('Không tìm thấy hộ khẩu!', 'danger')
        return redirect(url_for('hokhau_list'))
    
    # Lấy thành viên hiện tại
    query_hientai = """
        SELECT 
            n.cccd,
            n.name,
            n.ngaysinh,
            n.gioitinh,
            tv.quanhechuho,
            tv.ngaybatdau
        FROM thanhvienhokhau tv
        JOIN nguoidung n ON tv.cccd = n.cccd
        WHERE tv.mahokhau = %s AND tv.ngayketthuc IS NULL
        ORDER BY 
            CASE WHEN tv.quanhechuho = 'Chủ hộ' THEN 0 ELSE 1 END,
            tv.ngaybatdau
    """
    thanhvien_hientai = execute_query(query_hientai, (mahokhau,), fetch_all=True)
    thanhvien_hientai = thanhvien_hientai if thanhvien_hientai else []
    
    # Lấy lịch sử biến động (người đã rời khỏi hộ)
    query_lichsu = """
        SELECT 
            n.cccd,
            n.name,
            n.ngaysinh,
            tv.quanhechuho,
            tv.ngaybatdau,
            tv.ngayketthuc,
            tv.lydochuyen,
            tv.noichuyenden,
            tv.ghichu
        FROM thanhvienhokhau tv
        JOIN nguoidung n ON tv.cccd = n.cccd
        WHERE tv.mahokhau = %s AND tv.ngayketthuc IS NOT NULL
        ORDER BY tv.ngayketthuc DESC, tv.ngaybatdau DESC
    """
    lichsu = execute_query(query_lichsu, (mahokhau,), fetch_all=True)
    lichsu = lichsu if lichsu else []
    
    # Lấy lịch sử thay đổi chủ hộ
    query_doichuho = """
        SELECT 
            ls.ngaythaydoi,
            n1.name as ten_cu,
            n2.name as ten_moi,
            ls.lydothaydoi,
            ls.noidung,
            ls.nguoithuchien
        FROM lichsuthaydoichuho ls
        LEFT JOIN nguoidung n1 ON ls.cccd_cu = n1.cccd
        LEFT JOIN nguoidung n2 ON ls.cccd_moi = n2.cccd
        WHERE ls.mahokhau = %s
        ORDER BY ls.ngaythaydoi DESC
    """
    lichsu_doichuho = execute_query(query_doichuho, (mahokhau,), fetch_all=True)
    lichsu_doichuho = lichsu_doichuho if lichsu_doichuho else []
    
    return render_template('hokhau_lich_su.html',
                         hokhau=hokhau,
                         thanhvien_hientai=thanhvien_hientai,
                         lichsu=lichsu,
                         lichsu_doichuho=lichsu_doichuho,
                         mahokhau=mahokhau)


@app.route('/thong-ke/dan-so')
@login_required
def thongke_danso():
    """Thống kê dân số theo giới tính, độ tuổi và địa bàn"""
    
    # Lấy tham số filter
    xaphuong = request.args.get('xaphuong', '').strip()
    
    # Query tổng quan - Lấy toàn bộ người dùng hoặc lọc theo xã/phường nếu có
    if xaphuong:
        # Có filter theo xã/phường - phải JOIN với hộ khẩu
        query_total = """
            SELECT 
                COUNT(DISTINCT n.cccd) as tong,
                COUNT(DISTINCT CASE WHEN LOWER(n.gioitinh) = 'nam' THEN n.cccd END) as nam,
                COUNT(DISTINCT CASE WHEN LOWER(n.gioitinh) = 'nu' THEN n.cccd END) as nu
            FROM nguoidung n
            LEFT JOIN thanhvienhokhau tv ON n.cccd = tv.cccd AND tv.ngayketthuc IS NULL
            LEFT JOIN hokhau hk ON tv.mahokhau = hk.mahokhau
            LEFT JOIN diachi dc ON hk.madiachi = dc.madiachi
            WHERE LOWER(dc.xaphuong) LIKE LOWER(%s)
        """
        params_total = [f'%{xaphuong}%']
    else:
        # Không filter - lấy toàn bộ người dùng
        query_total = """
            SELECT 
                COUNT(cccd) as tong,
                COUNT(CASE WHEN LOWER(gioitinh) = 'nam' THEN cccd END) as nam,
                COUNT(CASE WHEN LOWER(gioitinh) = 'nu' THEN cccd END) as nu
            FROM nguoidung
        """
        params_total = []
    
    stats_total = execute_query(query_total, tuple(params_total) if params_total else None, fetch_one=True)
    stats_total = stats_total or (0, 0, 0)
    
    # Query phân nhóm tuổi (0-5, 6-10, 11-14, 15-17, 18-59, 60+)
    if xaphuong:
        query_age = """
            SELECT 
                CASE 
                    WHEN DATE_PART('year', AGE(CURRENT_DATE, n.ngaysinh)) BETWEEN 0 AND 5 THEN '0-5'
                    WHEN DATE_PART('year', AGE(CURRENT_DATE, n.ngaysinh)) BETWEEN 6 AND 10 THEN '6-10'
                    WHEN DATE_PART('year', AGE(CURRENT_DATE, n.ngaysinh)) BETWEEN 11 AND 14 THEN '11-14'
                    WHEN DATE_PART('year', AGE(CURRENT_DATE, n.ngaysinh)) BETWEEN 15 AND 17 THEN '15-17'
                    WHEN DATE_PART('year', AGE(CURRENT_DATE, n.ngaysinh)) BETWEEN 18 AND 59 THEN '18-59'
                    ELSE '60+'
                END as nhom_tuoi,
                COUNT(n.cccd) as soluong
            FROM nguoidung n
            LEFT JOIN thanhvienhokhau tv ON n.cccd = tv.cccd AND tv.ngayketthuc IS NULL
            LEFT JOIN hokhau hk ON tv.mahokhau = hk.mahokhau
            LEFT JOIN diachi dc ON hk.madiachi = dc.madiachi
            WHERE n.ngaysinh IS NOT NULL AND LOWER(dc.xaphuong) LIKE LOWER(%s)
            GROUP BY 1
            ORDER BY 1
        """
        params_age = [f'%{xaphuong}%']
    else:
        query_age = """
            SELECT 
                CASE 
                    WHEN DATE_PART('year', AGE(CURRENT_DATE, ngaysinh)) BETWEEN 0 AND 5 THEN '0-5'
                    WHEN DATE_PART('year', AGE(CURRENT_DATE, ngaysinh)) BETWEEN 6 AND 10 THEN '6-10'
                    WHEN DATE_PART('year', AGE(CURRENT_DATE, ngaysinh)) BETWEEN 11 AND 14 THEN '11-14'
                    WHEN DATE_PART('year', AGE(CURRENT_DATE, ngaysinh)) BETWEEN 15 AND 17 THEN '15-17'
                    WHEN DATE_PART('year', AGE(CURRENT_DATE, ngaysinh)) BETWEEN 18 AND 59 THEN '18-59'
                    ELSE '60+'
                END as nhom_tuoi,
                COUNT(cccd) as soluong
            FROM nguoidung
            WHERE ngaysinh IS NOT NULL
            GROUP BY 1
            ORDER BY 1
        """
        params_age = []
    
    stats_age = execute_query(query_age, tuple(params_age) if params_age else None, fetch_all=True)
    # Đảm bảo tất cả nhóm tuổi đều có trong kết quả
    age_order = {'0-5': 1, '6-10': 2, '11-14': 3, '15-17': 4, '18-59': 5, '60+': 6}
    all_age_groups = ['0-5', '6-10', '11-14', '15-17', '18-59', '60+']
    
    # Tạo dict từ kết quả query
    age_dict = {row[0]: row[1] for row in (stats_age or [])}
    
    # Đảm bảo tất cả nhóm tuổi đều có (với giá trị 0 nếu không có)
    stats_age = [(age_group, age_dict.get(age_group, 0)) for age_group in all_age_groups]
    
    # Debug - In ra để kiểm tra
    print("="*80)
    print("DEBUG - KIỂM TRA STATS_AGE:")
    print(f"Query trả về {len(age_dict)} nhóm tuổi: {list(age_dict.keys())}")
    print(f"Sau khi xử lý có {len(stats_age)} nhóm tuổi:")
    for age_group, count in stats_age:
        print(f"  - {age_group}: {count} người")
    print("="*80)
    
    # Query phân nhóm tuổi theo giới tính
    if xaphuong:
        query_age_gender = """
            SELECT 
                CASE 
                    WHEN DATE_PART('year', AGE(CURRENT_DATE, n.ngaysinh)) BETWEEN 0 AND 5 THEN '0-5'
                    WHEN DATE_PART('year', AGE(CURRENT_DATE, n.ngaysinh)) BETWEEN 6 AND 10 THEN '6-10'
                    WHEN DATE_PART('year', AGE(CURRENT_DATE, n.ngaysinh)) BETWEEN 11 AND 14 THEN '11-14'
                    WHEN DATE_PART('year', AGE(CURRENT_DATE, n.ngaysinh)) BETWEEN 15 AND 17 THEN '15-17'
                    WHEN DATE_PART('year', AGE(CURRENT_DATE, n.ngaysinh)) BETWEEN 18 AND 59 THEN '18-59'
                    ELSE '60+'
                END as nhom_tuoi,
                n.gioitinh,
                COUNT(n.cccd) as soluong
            FROM nguoidung n
            LEFT JOIN thanhvienhokhau tv ON n.cccd = tv.cccd AND tv.ngayketthuc IS NULL
            LEFT JOIN hokhau hk ON tv.mahokhau = hk.mahokhau
            LEFT JOIN diachi dc ON hk.madiachi = dc.madiachi
            WHERE n.ngaysinh IS NOT NULL AND LOWER(dc.xaphuong) LIKE LOWER(%s)
            GROUP BY 1, 2
            ORDER BY 1, 2
        """
        params_age_gender = [f'%{xaphuong}%']
    else:
        query_age_gender = """
            SELECT 
                CASE 
                    WHEN DATE_PART('year', AGE(CURRENT_DATE, ngaysinh)) BETWEEN 0 AND 5 THEN '0-5'
                    WHEN DATE_PART('year', AGE(CURRENT_DATE, ngaysinh)) BETWEEN 6 AND 10 THEN '6-10'
                    WHEN DATE_PART('year', AGE(CURRENT_DATE, ngaysinh)) BETWEEN 11 AND 14 THEN '11-14'
                    WHEN DATE_PART('year', AGE(CURRENT_DATE, ngaysinh)) BETWEEN 15 AND 17 THEN '15-17'
                    WHEN DATE_PART('year', AGE(CURRENT_DATE, ngaysinh)) BETWEEN 18 AND 59 THEN '18-59'
                    ELSE '60+'
                END as nhom_tuoi,
                gioitinh,
                COUNT(cccd) as soluong
            FROM nguoidung
            WHERE ngaysinh IS NOT NULL
            GROUP BY 1, 2
            ORDER BY 1, 2
        """
        params_age_gender = []
    
    stats_age_gender = execute_query(query_age_gender, tuple(params_age_gender) if params_age_gender else None, fetch_all=True)
    
    # Đảm bảo tất cả nhóm tuổi và giới tính đều có trong kết quả
    all_genders = ['Nam', 'Nu']
    
    # Tạo dict từ kết quả query
    age_gender_dict = {(row[0], row[1]): row[2] for row in (stats_age_gender or [])}
    
    # Đảm bảo tất cả tổ hợp nhóm tuổi x giới tính đều có
    stats_age_gender = []
    for age_group in all_age_groups:
        for gender in all_genders:
            count = age_gender_dict.get((age_group, gender), 0)
            stats_age_gender.append((age_group, gender, count))
    
    # Debug - In ra để kiểm tra
    print("="*80)
    print("DEBUG - KIỂM TRA STATS_AGE_GENDER:")
    print(f"Query trả về {len(age_gender_dict)} tổ hợp: {list(age_gender_dict.keys())}")
    print(f"Sau khi xử lý có {len(stats_age_gender)} tổ hợp:")
    for age_group, gender, count in stats_age_gender:
        print(f"  - {age_group} - {gender}: {count} người")
    print("="*80)
    print(f"DEBUG stats_total: {stats_total}")
    print("="*80)
    
    # Query thống kê theo địa bàn (top 10)
    query_diaban = """
        SELECT 
            dc.xaphuong,
            COUNT(DISTINCT n.cccd) as tong,
            COUNT(DISTINCT CASE WHEN LOWER(n.gioitinh) = 'nam' THEN n.cccd END) as nam,
            COUNT(DISTINCT CASE WHEN LOWER(n.gioitinh) = 'nu' THEN n.cccd END) as nu
        FROM nguoidung n
        LEFT JOIN thanhvienhokhau tv ON n.cccd = tv.cccd AND tv.ngayketthuc IS NULL
        LEFT JOIN hokhau hk ON tv.mahokhau = hk.mahokhau
        LEFT JOIN diachi dc ON hk.madiachi = dc.madiachi
        WHERE dc.xaphuong IS NOT NULL
    """
    
    params_diaban = []
    if xaphuong:
        query_diaban += " AND LOWER(dc.xaphuong) LIKE LOWER(%s)"
        params_diaban.append(f'%{xaphuong}%')
    
    query_diaban += """
        GROUP BY dc.xaphuong
        ORDER BY tong DESC
        LIMIT 10
    """
    
    stats_diaban = execute_query(query_diaban, tuple(params_diaban) if params_diaban else None, fetch_all=True)
    stats_diaban = stats_diaban or []
    
    return render_template('thongke_danso.html',
                         stats_total=stats_total,
                         stats_age=stats_age,
                         stats_age_gender=stats_age_gender,
                         stats_diaban=stats_diaban,
                         xaphuong=xaphuong)

@app.route('/thong-ke/tam-vang-tru')
@login_required
def thongke_tamvangtru():
    """Báo cáo thống kê tạm vắng/tạm trú"""
    
    # Lấy tham số filter
    thang = request.args.get('thang', '')
    nam = request.args.get('nam', '')
    loai = request.args.get('loai', '')
    xaphuong = request.args.get('xaphuong', '').strip()
    
    # Build query
    query = """
        SELECT 
            dn.loaidiachi,
            dc.xaphuong,
            COUNT(*) as soluong,
            COUNT(CASE WHEN dn.thoidiemketthuc >= CURRENT_DATE THEN 1 END) as dang_hieuluc,
            COUNT(CASE WHEN dn.thoidiemketthuc < CURRENT_DATE THEN 1 END) as da_hethan
        FROM diachinguoidung dn
        INNER JOIN diachi dc ON dn.madiachi = dc.madiachi
        WHERE dn.loaidiachi IN ('TamVang', 'TamTru')
    """
    
    params = []
    
    # Filter by month/year
    if thang and nam:
        query += " AND EXTRACT(MONTH FROM dn.ngaybatdau) = %s AND EXTRACT(YEAR FROM dn.ngaybatdau) = %s"
        params.extend([int(thang), int(nam)])
    elif nam:
        query += " AND EXTRACT(YEAR FROM dn.ngaybatdau) = %s"
        params.append(int(nam))
    
    # Filter by type
    if loai:
        query += " AND dn.loaidiachi = %s"
        params.append(loai)
    
    # Filter by district
    if xaphuong:
        query += " AND LOWER(dc.xaphuong) LIKE LOWER(%s)"
        params.append(f'%{xaphuong}%')
    
    query += """
        GROUP BY dn.loaidiachi, dc.xaphuong
        ORDER BY dn.loaidiachi, dc.xaphuong
    """
    
    stats = execute_query(query, tuple(params) if params else None, fetch_all=True)
    
    # Query chi tiết
    query_detail = """
        SELECT 
            n.cccd,
            n.hoten,
            n.ngaysinh,
            n.gioitinh,
            dn.loaidiachi,
            dc.xaphuong,
            dc.chitiet,
            dn.ngaybatdau,
            dn.ngayketthuc,
            dn.lydo
        FROM diachinguoidung dn
        INNER JOIN diachi dc ON dn.madiachi = dc.madiachi
        INNER JOIN nguoidung n ON dn.cccd = n.cccd
        WHERE dn.loaidiachi IN ('TamVang', 'TamTru')
    """
    
    params_detail = []
    
    if thang and nam:
        query_detail += " AND EXTRACT(MONTH FROM dn.ngaybatdau) = %s AND EXTRACT(YEAR FROM dn.ngaybatdau) = %s"
        params_detail.extend([int(thang), int(nam)])
    elif nam:
        query_detail += " AND EXTRACT(YEAR FROM dn.ngaybatdau) = %s"
        params_detail.append(int(nam))
    
    if loai:
        query_detail += " AND dn.loaidiachi = %s"
        params_detail.append(loai)
    
    if xaphuong:
        query_detail += " AND LOWER(dc.xaphuong) LIKE LOWER(%s)"
        params_detail.append(f'%{xaphuong}%')
    
    query_detail += " ORDER BY dn.ngaybatdau DESC"
    
    details = execute_query(query_detail, tuple(params_detail) if params_detail else None, fetch_all=True)
    
    # Đảm bảo stats không phải None để tránh lỗi for qua NoneType
    stats = stats or []
    tong_tamvang = sum(row[2] for row in stats if row[0] == 'TamVang')
    tong_tamtru = sum(row[2] for row in stats if row[0] == 'TamTru')
    
    # Đảm bảo details không phải None để tránh lỗi len(None)
    details = details or []
    from datetime import date
    return render_template('thongke_tamvangtru.html',
                         stats=stats,
                         details=details,
                         tong_tamvang=tong_tamvang,
                         tong_tamtru=tong_tamtru,
                         thang=thang,
                         nam=nam,
                         loai=loai,
                         xaphuong=xaphuong,
                         today=date.today(),
                         openpyxl_available=OPENPYXL_AVAILABLE)


@app.route('/thong-ke/tam-vang-tru/export')
@login_required
def export_tamvangtru():
    """Export báo cáo tạm vắng/tạm trú ra Excel"""
    
    if not OPENPYXL_AVAILABLE:
        flash('Vui lòng cài đặt thư viện openpyxl để sử dụng tính năng này', 'error')
        return redirect(url_for('thongke_tamvangtru'))
    
    # Lấy tham số filter (same as report view)
    thang = request.args.get('thang', '')
    nam = request.args.get('nam', '')
    loai = request.args.get('loai', '')
    xaphuong = request.args.get('xaphuong', '').strip()
    
    # Query chi tiết
    query_detail = """
        SELECT 
            n.cccd,
            n.name,
            n.ngaysinh,
            n.gioitinh,
            dn.loaidiachi,
            dc.xaphuong,
            dc.chitiet,
            dn.thoidiemxacnhan,
            dn.thoidiemketthuc
        FROM diachinguoidung dn
        INNER JOIN diachi dc ON dn.madiachi = dc.madiachi
        INNER JOIN nguoidung n ON dn.cccd = n.cccd
        WHERE dn.loaidiachi IN ('TamVang', 'TamTru')
    """
    
    params = []
    
    if thang and nam:
        query_detail += " AND EXTRACT(MONTH FROM dn.thoidiemxacnhan) = %s AND EXTRACT(YEAR FROM dn.thoidiemxacnhan) = %s"
        params.extend([int(thang), int(nam)])
    elif nam:
        query_detail += " AND EXTRACT(YEAR FROM dn.thoidiemxacnhan) = %s"
        params.append(int(nam))
    
    if loai:
        query_detail += " AND dn.loaidiachi = %s"
        params.append(loai)
    
    if xaphuong:
        query_detail += " AND LOWER(dc.xaphuong) LIKE LOWER(%s)"
        params.append(f'%{xaphuong}%')
    
    query_detail += " ORDER BY dn.loaidiachi, dn.thoidiemxacnhan DESC"
    
    details = execute_query(query_detail, tuple(params) if params else None, fetch_all=True)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo Tạm vắng Tạm trú"
    
    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = "BÁO CÁO TẠM VẮNG/TẠM TRÚ"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Filter info
    filter_info = f"Thời gian: {thang}/{nam}" if thang and nam else (f"Năm: {nam}" if nam else "Tất cả")
    if loai:
        filter_info += f" - Loại: {loai}"
    if xaphuong:
        filter_info += f" - Xã/Phường: {xaphuong}"
    
    ws.merge_cells('A2:J2')
    filter_cell = ws['A2']
    filter_cell.value = filter_info
    filter_cell.alignment = Alignment(horizontal="center")
    
    # Headers
    headers = ['STT', 'CCCD', 'Họ tên', 'Ngày sinh', 'Giới tính', 'Loại', 'Xã/Phường', 
               'Địa chỉ', 'Ngày bắt đầu', 'Ngày kết thúc']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    for idx, row in enumerate(details, start=1):
        ws_row = idx + 4
        ws.cell(row=ws_row, column=1, value=idx).border = border
        ws.cell(row=ws_row, column=2, value=row[0]).border = border
        ws.cell(row=ws_row, column=3, value=row[1]).border = border
        ws.cell(row=ws_row, column=4, value=row[2].strftime('%d/%m/%Y') if row[2] else '').border = border
        ws.cell(row=ws_row, column=5, value=row[3]).border = border
        ws.cell(row=ws_row, column=6, value='Tạm vắng' if row[4] == 'TamVang' else 'Tạm trú').border = border
        ws.cell(row=ws_row, column=7, value=row[5]).border = border
        ws.cell(row=ws_row, column=8, value=row[6]).border = border
        ws.cell(row=ws_row, column=9, value=row[7].strftime('%d/%m/%Y') if row[7] else '').border = border
        ws.cell(row=ws_row, column=10, value=row[8].strftime('%d/%m/%Y') if row[8] else '').border = border
        # ws.cell(row=ws_row, column=11, value=row[9]).border = border
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 13
    ws.column_dimensions['J'].width = 13
    # ws.column_dimensions['K'].width = 25
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with timestamp
    filename = f"BaoCao_TamVangTamTru_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@role_required(['CanBo', 'QuanLy'])
@app.route('/nguoidung')
@login_required  
def nguoidung_list():
    """Danh sách người dùng/nhân khẩu với phân trang, tìm kiếm và lọc"""
    
    user_role = session['user'].get('vaitro')
    user_cccd = session['user'].get('cccd')
    
    # ========== CẤU HÌNH PHÂN TRANG ==========
    per_page = 20  # Số bản ghi mỗi trang
    
    # Lấy số trang từ URL, mặc định là 1
    try:
        page = int(request.args.get('page', 1))
        if page < 1:
            page = 1
    except (ValueError, TypeError):
        page = 1
    
    # Tính offset
    offset = (page - 1) * per_page
    
    # ========== LẤY THAM SỐ TÌM KIẾM VÀ LỌC ==========
    search = request.args.get('search', '').strip()
    gioitinh = request.args.get('gioitinh', '').strip()
    vaitro = request.args.get('vaitro', '').strip()
    dantoc = request.args.get('dantoc', '').strip()
    
    # ========== XÂY DỰNG ĐIỀU KIỆN WHERE ==========
    where_conditions = []
    params = []
    
    # Nếu là Người dân, chỉ xem những người trong hộ khẩu của mình
    if user_role == 'NguoiDan':
        where_conditions.append("""
            cccd IN (
                SELECT DISTINCT tv.cccd 
                FROM thanhvienhokhau tv
                WHERE tv.mahokhau IN (
                    SELECT mahokhau FROM thanhvienhokhau WHERE cccd = %s
                )
            )
        """)
        params.append(user_cccd)
    
    if search:
        where_conditions.append("(cccd LIKE %s OR name ILIKE %s OR sdt LIKE %s)")
        search_pattern = f"%{search}%"
        params.extend([search_pattern, search_pattern, search_pattern])
    
    if gioitinh:
        where_conditions.append("LOWER(gioitinh) = LOWER(%s)")
        params.append(gioitinh)
    
    if vaitro:
        where_conditions.append("vaitro = %s")
        params.append(vaitro)
    
    if dantoc:
        where_conditions.append("dantoc ILIKE %s")
        params.append(dantoc)
    
    where_clause = " AND ".join(where_conditions) if where_conditions else "1=1"
    
    # ========== ĐẾM TỔNG SỐ NGƯỜI DÙNG ==========
    query_count = f"SELECT COUNT(*) FROM nguoidung WHERE {where_clause}"
    total_count = execute_query(query_count, tuple(params), fetch_one=True)
    total_count = total_count[0] if total_count else 0
    
    # Tính tổng số trang
    import math
    total_pages = math.ceil(total_count / per_page) if total_count > 0 else 1
    
    # Đảm bảo page không vượt quá total_pages
    if page > total_pages and total_pages > 0:
        page = total_pages
    
    # ========== QUERY DANH SÁCH NGƯỜI DÙNG VỚI PHÂN TRANG ==========
    query = f"""
        SELECT 
            cccd,
            name,
            sdt,
            ngaysinh,
            gioitinh,
            dantoc,
            vaitro
        FROM nguoidung
        WHERE {where_clause}
        ORDER BY name
        LIMIT %s OFFSET %s
    """
    params.extend([per_page, offset])
    ds_nguoidung = execute_query(query, tuple(params), fetch_all=True)
    ds_nguoidung = ds_nguoidung if ds_nguoidung else []
    
    return render_template('nguoidung.html', 
                         ds_nguoidung=ds_nguoidung,
                         page=page,
                         per_page=per_page,
                         total_pages=total_pages,
                         total_count=total_count,
                         search=search,
                         gioitinh=gioitinh,
                         vaitro=vaitro,
                         dantoc=dantoc)


@role_required(['CanBo', 'QuanLy'])
@app.route('/nguoidung/add', methods=['GET', 'POST'])
@login_required
def nguoidung_add():
    """Thêm nhân khẩu mới"""
    if request.method == 'POST':
        # Lấy dữ liệu từ form - Thông tin cơ bản
        cccd = request.form.get('cccd', '').strip()
        name = request.form.get('name', '').strip()
        user_name = request.form.get('user_name', '').strip()
        matkhau = request.form.get('matkhau', '').strip()
        bidanh = request.form.get('bidanh', '').strip()
        sdt = request.form.get('sdt', '').strip()
        ngaysinh = request.form.get('ngaysinh', '').strip()
        gioitinh = request.form.get('gioitinh', '').strip()
        dantoc = request.form.get('dantoc', '').strip()
        vaitro = request.form.get('vaitro', '').strip()
        nghenghiep = request.form.get('nghenghiep', '').strip()
        noilamviec = request.form.get('noilamviec', '').strip()
        noisinh = request.form.get('noisinh', '').strip()
        nguyenquan = request.form.get('nguyenquan', '').strip()
        
        # Thông tin CCCD
        ngaycapcccd = request.form.get('ngaycapcccd', '').strip()
        noicapcccd = request.form.get('noicapcccd', '').strip()
        
        # Địa chỉ thường trú
        ngaydangkythuongtru = request.form.get('ngaydangkythuongtru', '').strip()
        diachitruoc = request.form.get('diachitruoc', '').strip()
        
        # Validate
        if not cccd or not name or not user_name or not matkhau:
            flash('Vui lòng nhập đầy đủ thông tin bắt buộc (CCCD, Họ tên, Tên đăng nhập, Mật khẩu)!', 'danger')
            return render_template('nguoidung_add.html')
        
        # Kiểm tra CCCD đã tồn tại chưa
        check_query = "SELECT cccd FROM nguoidung WHERE cccd = %s"
        existing = execute_query(check_query, (cccd,), fetch_one=True)
        if existing:
            flash('CCCD đã tồn tại trong hệ thống!', 'danger')
            return render_template('nguoidung_add.html')
        
        # Kiểm tra username đã tồn tại chưa
        check_username = "SELECT user_name FROM nguoidung WHERE user_name = %s"
        existing_username = execute_query(check_username, (user_name,), fetch_one=True)
        if existing_username:
            flash('Tên đăng nhập đã tồn tại!', 'danger')
            return render_template('nguoidung_add.html')
        
        # Insert vào database
        insert_query = """
            INSERT INTO nguoidung (
                cccd, name, user_name, matkhau, sdt, ngaysinh, gioitinh, dantoc, vaitro, nghenghiep, noilamviec
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        try:
            # Log params để debug
            params_list = (
                cccd, name, user_name, matkhau,
                sdt or None, ngaysinh or None, gioitinh or None, dantoc or None,
                vaitro if vaitro in ['NguoiDan', 'CanBo', 'QuanLy'] else 'NguoiDan', nghenghiep or None, noilamviec or None
            )
            print(f"DEBUG - Insert params: {params_list}")

            result = execute_query(insert_query, params_list)
            print(f"DEBUG - Execute result: {result}, type: {type(result)}")

            if result is None or result == 0:
                print(f"DEBUG - Result is None or 0: {result}")
                # Try to get more info from DB
                connection = get_db_connection()
                if connection:
                    try:
                        cursor = connection.cursor()
                        cursor.execute('ROLLBACK')
                        print("DEBUG - DB rollback done.")
                    except Exception as e2:
                        print(f"DEBUG - Rollback error: {e2}")
                    finally:
                        connection.close()
                flash(f'Có lỗi xảy ra khi thêm nhân khẩu! (result={result})', 'danger')
            else:
                flash(f'Đã thêm nhân khẩu {name} thành công!', 'success')
                return redirect(url_for('nguoidung_list'))
        except Exception as e:
            flash(f'Lỗi: {str(e)}', 'danger')
            import traceback
            print("=== TRACEBACK ===")
            print(traceback.format_exc())
            print(f"Error type: {type(e).__name__}")
            print(f"Error: {str(e)}")
    
    return render_template('nguoidung_add.html')


@app.route('/nguoidung/<string:cccd>')
@login_required
def nguoidung_detail(cccd):
    """Xem chi tiết nhân khẩu"""
    
    user_role = session['user'].get('vaitro')
    user_cccd = session['user'].get('cccd')
    
    # Lấy thông tin nhân khẩu
    query = """
        SELECT cccd, name, user_name, sdt, ngaysinh, gioitinh, dantoc, vaitro, nghenghiep
        FROM nguoidung WHERE cccd = %s
    """
    nguoidung = execute_query(query, (cccd,), fetch_one=True)
    
    if not nguoidung:
        flash('Không tìm thấy nhân khẩu!', 'danger')
        return redirect(url_for('nguoidung_list'))
    
    # Nếu là Người dân, kiểm tra xem người này có trong hộ khẩu của mình không
    if user_role == 'NguoiDan':
        query_check = """
            SELECT 1 FROM thanhvienhokhau tv1
            WHERE tv1.cccd = %s 
            AND tv1.mahokhau IN (
                SELECT mahokhau FROM thanhvienhokhau WHERE cccd = %s
            )
        """
        is_in_same_hokhau = execute_query(query_check, (cccd, user_cccd), fetch_one=True)
        if not is_in_same_hokhau:
            flash('Bạn không có quyền xem thông tin người này!', 'danger')
            return redirect(url_for('nguoidung_list'))
    
    # Lấy danh sách hộ khẩu mà người này tham gia (hiện tại và quá khứ)
    query_hokhau = """
        SELECT 
            hk.mahokhau,
            dc.xaphuong,
            dc.chitiet,
            tv.quanhechuho,
            tv.ngaybatdau,
            tv.ngayketthuc,
            hk.ghichu
        FROM thanhvienhokhau tv
        INNER JOIN hokhau hk ON tv.mahokhau = hk.mahokhau
        LEFT JOIN diachi dc ON hk.madiachi = dc.madiachi
        WHERE tv.cccd = %s
        ORDER BY tv.ngaybatdau DESC
    """
    ds_hokhau = execute_query(query_hokhau, (cccd,), fetch_all=True)
    ds_hokhau = ds_hokhau or []
    
    # Lấy lịch sử tạm vắng/tạm trú
    query_tamvangtru = """
        SELECT 
            dn.loaidiachi,
            dc.xaphuong,
            dc.chitiet,
            dn.ngaybatdau,
            dn.ngayketthuc,
            dn.lydo
        FROM diachinguoidung dn
        LEFT JOIN diachi dc ON dn.madiachi = dc.madiachi
        WHERE dn.cccd = %s AND dn.loaidiachi IN ('TamVang', 'TamTru')
        ORDER BY dn.ngaybatdau DESC
    """
    ds_tamvangtru = execute_query(query_tamvangtru, (cccd,), fetch_all=True)
    ds_tamvangtru = ds_tamvangtru or []
    
    return render_template('nguoidung_detail.html', 
                         nguoidung=nguoidung,
                         ds_hokhau=ds_hokhau,
                         ds_tamvangtru=ds_tamvangtru)

@role_required(['CanBo', 'QuanLy'])

@app.route('/nguoidung/edit/<string:cccd>', methods=['GET', 'POST'])
@login_required
def nguoidung_edit(cccd):
    """Sửa thông tin nhân khẩu"""
    
    if request.method == 'POST':
        # Lấy dữ liệu từ form
        name = request.form.get('name', '').strip()
        user_name = request.form.get('user_name', '').strip()
        bidanh = request.form.get('bidanh', '').strip()
        sdt = request.form.get('sdt', '').strip()
        ngaysinh = request.form.get('ngaysinh', '').strip()
        gioitinh = request.form.get('gioitinh', '').strip()
        dantoc = request.form.get('dantoc', '').strip()
        vaitro = request.form.get('vaitro', '').strip()
        nghenghiep = request.form.get('nghenghiep', '').strip()
        noilamviec = request.form.get('noilamviec', '').strip()
        noisinh = request.form.get('noisinh', '').strip()
        nguyenquan = request.form.get('nguyenquan', '').strip()
        ngaycapcccd = request.form.get('ngaycapcccd', '').strip()
        noicapcccd = request.form.get('noicapcccd', '').strip()
        ngaydangkythuongtru = request.form.get('ngaydangkythuongtru', '').strip()
        diachitruoc = request.form.get('diachitruoc', '').strip()
        
        # Validate
        if not name or not user_name:
            flash('Vui lòng nhập đầy đủ họ tên và tên đăng nhập!', 'danger')
            return redirect(url_for('nguoidung_edit', cccd=cccd))
        
        # Kiểm tra username đã tồn tại chưa (trừ chính nó)
        check_username = "SELECT user_name FROM nguoidung WHERE user_name = %s AND cccd != %s"
        existing_username = execute_query(check_username, (user_name, cccd), fetch_one=True)
        if existing_username:
            flash('Tên đăng nhập đã tồn tại!', 'danger')
            return redirect(url_for('nguoidung_edit', cccd=cccd))
        
        # Update database
        update_query = """
            UPDATE nguoidung 
            SET name = %s, user_name = %s, sdt = %s, ngaysinh = %s, gioitinh = %s, 
                dantoc = %s, vaitro = %s, nghenghiep = %s, bidanh = %s, noilamviec = %s,
                noisinh = %s, nguyenquan = %s, ngaycapcccd = %s, noicapcccd = %s,
                ngaydangkythuongtru = %s, diachitruoc = %s
            WHERE cccd = %s
        """
        result = execute_query(update_query, (
            name, user_name, sdt or None, ngaysinh or None, gioitinh or None,
            dantoc or None, vaitro or 'NguoiDan', nghenghiep or None, bidanh or None, noilamviec or None,
            noisinh or None, nguyenquan or None, ngaycapcccd or None, noicapcccd or None,
            ngaydangkythuongtru or None, diachitruoc or None, cccd
        ))
        
        if result:
            flash(f'Đã cập nhật thông tin nhân khẩu {name} thành công!', 'success')
            return redirect(url_for('nguoidung_list'))
        else:
            flash('Có lỗi xảy ra khi cập nhật!', 'danger')
    
    # GET: Load dữ liệu hiện tại
    query = """
        SELECT cccd, name, user_name, sdt, ngaysinh, gioitinh, dantoc, vaitro, nghenghiep
        FROM nguoidung WHERE cccd = %s
    """
    nguoidung = execute_query(query, (cccd,), fetch_one=True)
    
    if not nguoidung:
        flash('Không tìm thấy nhân khẩu!', 'danger')
        return redirect(url_for('nguoidung_list'))
    
    # Tạo tuple mở rộng với các giá trị mặc định cho các cột chưa có trong DB
    # Thứ tự: cccd, name, user_name, sdt, ngaysinh, gioitinh, dantoc, vaitro, nghenghiep,
    #         bidanh, noilamviec, noisinh, nguyenquan, ngaycapcccd, noicapcccd,
    #         ngaydangkythuongtru, diachitruoc
    nguoidung_extended = nguoidung + ('', '', '', '', None, '', None, '')  # Thêm 8 cột mặc định
    
    return render_template('nguoidung_edit.html', nguoidung=nguoidung_extended)

@role_required(['CanBo', 'QuanLy'])

@app.route('/nguoidung/delete/<string:cccd>', methods=['POST'])
@login_required
def nguoidung_delete(cccd):
    """Xóa nhân khẩu"""
    
    # Kiểm tra xem người này có trong hộ khẩu nào không
    check_hokhau = "SELECT COUNT(*) FROM thanhvienhokhau WHERE cccd = %s AND ngayketthuc IS NULL"
    count = execute_query(check_hokhau, (cccd,), fetch_one=True)
    
    if count and count[0] > 0:
        flash('Không thể xóa! Nhân khẩu này đang thuộc hộ khẩu. Vui lòng xóa khỏi hộ khẩu trước.', 'danger')
        return redirect(url_for('nguoidung_list'))
    
    # Lấy tên để hiển thị thông báo
    query_name = "SELECT name FROM nguoidung WHERE cccd = %s"
    nguoidung = execute_query(query_name, (cccd,), fetch_one=True)
    name = nguoidung[0] if nguoidung else cccd
    
    # Xóa
    delete_query = "DELETE FROM nguoidung WHERE cccd = %s"
    result = execute_query(delete_query, (cccd,))
    
    if result:
        flash(f'Đã xóa nhân khẩu {name} thành công!', 'success')
    else:
        flash('Có lỗi xảy ra khi xóa!', 'danger')
    
    return redirect(url_for('nguoidung_list'))


@app.route('/phananh')
@login_required
def phananh_list():
    """Danh sách phản ánh - Phân quyền theo vai trò"""
    
    # Người dân chỉ xem phản ánh của mình
    # Cán bộ/Quản lý xem tất cả
    user_role = session['user'].get('vaitro')
    user_cccd = session['user'].get('cccd')
    
    if user_role == 'NguoiDan':
        # Người dân chỉ xem phản ánh của mình
        query = """
            SELECT 
                p.maphananh,
                p.tieude,
                p.loaiphananh,
                p.trangthaiphananh,
                p.mota,
                p.thoigiantao,
                p.is_public,
                p.like_count,
                p.comment_count,
                p.view_count,
                v.tenvande,
                v.trangthai as trangthai_vande
            FROM phananh p
            LEFT JOIN vande v ON p.mavande = v.mavande
            WHERE p.cccd = %s
            ORDER BY p.thoigiantao DESC
        """
        ds_phananh = execute_query(query, (user_cccd,), fetch_all=True)
    else:
        # Cán bộ/Quản lý xem tất cả
        query = """
            SELECT 
                p.maphananh,
                n.name as nguoiphan,
                p.tieude,
                p.loaiphananh,
                p.trangthaiphananh,
                p.mota,
                p.thoigiantao,
                p.is_public,
                p.like_count,
                p.comment_count,
                p.view_count,
                v.tenvande,
                v.trangthai as trangthai_vande,
                d.xaphuong,
                d.chitiet
            FROM phananh p
            LEFT JOIN nguoidung n ON p.cccd = n.cccd
            LEFT JOIN vande v ON p.mavande = v.mavande
            LEFT JOIN diachi d ON p.madiachi = d.madiachi
            ORDER BY p.thoigiantao DESC
        """
        ds_phananh = execute_query(query, fetch_all=True)
    
    ds_phananh = ds_phananh if ds_phananh else []
    
    return render_template('phananh.html', ds_phananh=ds_phananh)


@app.route('/phananh/add', methods=['GET', 'POST'])
@login_required
def phananh_add():
    """Tạo phản ánh mới - Người dân có thể tạo phản ánh"""
    
    # Lấy thông tin địa chỉ hộ khẩu của user để có thể tự động điền (dùng cho cả GET và POST)
    user_cccd = session['user']['cccd']
    query_hokhau_diachi = """
        SELECT d.tinh, d.xaphuong, d.chitiet
        FROM thanhvienhokhau tv
        JOIN hokhau h ON tv.mahokhau = h.mahokhau
        LEFT JOIN diachi d ON h.madiachi = d.madiachi
        WHERE tv.cccd = %s AND tv.ngayketthuc IS NULL
        LIMIT 1
    """
    hokhau_diachi = execute_query(query_hokhau_diachi, (user_cccd,), fetch_one=True)
    
    if request.method == 'POST':
        tieude = request.form.get('tieude', '').strip()
        mota = request.form.get('mota', '').strip()
        loaiphananh = request.form.get('loaiphananh', '').strip()
        is_public = request.form.get('is_public') == 'on'
        allow_comment = request.form.get('allow_comment') == 'on'
        
        # Địa chỉ (optional)
        tinh = request.form.get('tinh', '').strip()
        xaphuong = request.form.get('xaphuong', '').strip()
        chitiet = request.form.get('chitiet', '').strip()
        
        # Xử lý file đính kèm
        matepdinhkem = None
        if 'file' in request.files:
            file = request.files['file']
            if file and file.filename:
                matepdinhkem = upload_file(file)
                if matepdinhkem:
                    print(f"DEBUG: Đã upload file với matepdinhkem = {matepdinhkem}")
                else:
                    flash('File không hợp lệ hoặc quá lớn. Cho phép: png, jpg, pdf, doc, xls (max 16MB)', 'warning')
        
        # DEBUG - In ra thông tin form
        print("="*80)
        print("DEBUG PHANANH_ADD - FORM DATA:")
        print(f"  - Tiêu đề: {tieude}")
        print(f"  - Mô tả: {mota[:100]}..." if len(mota) > 100 else f"  - Mô tả: {mota}")
        print(f"  - Loại phản ánh: {loaiphananh}")
        print(f"  - Is public: {is_public}")
        print(f"  - Allow comment: {allow_comment}")
        print(f"  - Địa chỉ: {tinh} - {xaphuong} - {chitiet}")
        print(f"  - CCCD user: {session['user']['cccd']}")
        print("="*80)
        
        # Validate
        if not tieude or not mota:
            print("ERROR: Thiếu tiêu đề hoặc mô tả!")
            flash('Vui lòng nhập đầy đủ tiêu đề và mô tả!', 'warning')
            return render_template('phananh_add.html', hokhau_diachi=hokhau_diachi)
        
        try:
            connection = get_db_connection()
            if not connection:
                flash('Không thể kết nối database!', 'danger')
                return render_template('phananh_add.html', hokhau_diachi=hokhau_diachi)
            
            cursor = connection.cursor()
            
            # Tạo địa chỉ nếu có
            madiachi = None
            if tinh or xaphuong or chitiet:
                query_diachi = """
                    INSERT INTO diachi (tinh, xaphuong, chitiet)
                    VALUES (%s, %s, %s)
                    RETURNING madiachi
                """
                print(f"DEBUG: Đang tạo địa chỉ: {tinh}, {xaphuong}, {chitiet}")
                cursor.execute(query_diachi, (tinh, xaphuong, chitiet))
                result = cursor.fetchone()
                if result:
                    madiachi = result[0]
                    print(f"DEBUG: Đã tạo địa chỉ với madiachi = {madiachi}")
                else:
                    print("WARNING: Không tạo được địa chỉ (result = None)")
            
            # Tạo phản ánh
            query_phananh = """
                INSERT INTO phananh (
                    cccd, madiachi, loaiphananh, trangthaiphananh,
                    mota, tieude, is_public, allow_comment,
                    matepdinhkem, thoigiantao
                )
                VALUES (%s, %s, %s, 'ChuaXuLy', %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
                RETURNING maphananh
            """
            
            print(f"DEBUG: Đang tạo phản ánh với params: cccd={session['user']['cccd']}, madiachi={madiachi}, loaiphananh={loaiphananh}, matepdinhkem={matepdinhkem}")
            cursor.execute(query_phananh, (session['user']['cccd'], madiachi, loaiphananh, mota, tieude, is_public, allow_comment, matepdinhkem))
            result = cursor.fetchone()
            
            if result:
                maphananh = result[0]
                print(f"SUCCESS: Đã tạo phản ánh với maphananh = {maphananh}")
                
                # Tự động tạo boxchat nếu phản ánh riêng tư
                if not is_public:
                    query_boxchat = """
                        INSERT INTO boxchat (maphananh, cccd_nguoidan)
                        VALUES (%s, %s)
                    """
                    print(f"DEBUG: Tạo boxchat cho phản ánh riêng tư {maphananh}")
                    cursor.execute(query_boxchat, (maphananh, session['user']['cccd']))
                
                # Commit tất cả thay đổi
                connection.commit()
                cursor.close()
                connection.close()
                
                print("="*80)
                flash(f'Đã tạo phản ánh thành công! Mã phản ánh: {maphananh}', 'success')
                return redirect(url_for('phananh_detail', maphananh=maphananh))
            else:
                print("ERROR: execute_query trả về None khi tạo phản ánh!")
                connection.rollback()
                cursor.close()
                connection.close()
                flash('Có lỗi xảy ra khi tạo phản ánh!', 'danger')
                return render_template('phananh_add.html', hokhau_diachi=hokhau_diachi)
        
        except Exception as e:
            print(f"EXCEPTION trong phananh_add: {type(e).__name__}: {str(e)}")
            import traceback
            traceback.print_exc()
            if connection:
                connection.rollback()
                connection.close()
            flash(f'Có lỗi xảy ra: {str(e)}', 'danger')
            return render_template('phananh_add.html', hokhau_diachi=hokhau_diachi)
    
    return render_template('phananh_add.html', hokhau_diachi=hokhau_diachi)


@app.route('/phananh/<int:maphananh>')
@login_required
def phananh_detail(maphananh):
    """Xem chi tiết phản ánh - Tăng view count"""
    
    # Query thông tin phản ánh đầy đủ
    query = """
        SELECT 
            p.maphananh,
            p.cccd,
            n.name AS nguoi_tao,
            n.sdt AS sdt_nguoi_tao,
            p.tieude,
            p.mota,
            p.loaiphananh,
            p.trangthaiphananh,
            p.is_public,
            p.allow_comment,
            p.like_count,
            p.comment_count,
            p.view_count,
            p.thoigiantao,
            p.thoigianxuly,
            p.mavande,
            v.tenvande,
            v.phanloai AS phanloai_vande,
            v.trangthai AS trangthai_vande,
            v.ketqua AS ketqua_vande,
            d.tinh,
            d.xaphuong,
            d.chitiet AS diachi_chitiet,
            t.duongdan AS hinh_anh,
            n.avatar_url
        FROM phananh p
        LEFT JOIN nguoidung n ON p.cccd = n.cccd
        LEFT JOIN vande v ON p.mavande = v.mavande
        LEFT JOIN diachi d ON p.madiachi = d.madiachi
        LEFT JOIN tepdinhkem t ON p.matepdinhkem = t.matepdinhkem
        WHERE p.maphananh = %s
    """
    
    phananh = execute_query(query, (maphananh,), fetch_one=True)
    
    if not phananh:
        flash('Không tìm thấy phản ánh!', 'danger')
        return redirect(url_for('phananh_list'))
    
    # Kiểm tra quyền xem (Phản ánh riêng tư chỉ chủ nhân và Cán bộ/Quản lý xem được)
    user_role = session['user'].get('vaitro')
    user_cccd = session['user'].get('cccd')
    phananh_cccd = phananh[1]  # cccd của người tạo
    is_public = phananh[8]  # is_public
    
    if not is_public and user_role == 'NguoiDan' and user_cccd != phananh_cccd:
        flash('Bạn không có quyền xem phản ánh này!', 'danger')
        return redirect(url_for('phananh_list'))
    
    # Tăng view count - Chỉ tăng 1 lần cho mỗi user
    viewed_posts = session.get('viewed_posts', [])
    if maphananh not in viewed_posts:
        update_view = "UPDATE phananh SET view_count = view_count + 1 WHERE maphananh = %s"
        execute_query(update_view, (maphananh,))
        # Lưu vào session
        viewed_posts.append(maphananh)
        session['viewed_posts'] = viewed_posts
        session.modified = True
    
    # Lấy danh sách bình luận (nếu cho phép)
    comments = []
    if phananh[9]:  # allow_comment
        query_comments = """
            SELECT 
                b.id,
                b.cccd_nguoidung,
                n.name AS ten_nguoi_binh_luan,
                n.avatar_url,
                b.noidung,
                b.thoigian,
                b.parent_id,
                b.is_hidden,
                t.duongdan AS file_dinh_kem,
                b.matepdinhkem
            FROM binhluan b
            LEFT JOIN nguoidung n ON b.cccd_nguoidung = n.cccd
            LEFT JOIN tepdinhkem t ON b.matepdinhkem = t.matepdinhkem
            WHERE b.maphananh = %s AND b.is_hidden = FALSE
            ORDER BY b.thoigian DESC
        """
        comments = execute_query(query_comments, (maphananh,), fetch_all=True)
        comments = comments if comments else []
    
    # Kiểm tra user đã like chưa
    user_liked = False
    if is_public:
        query_like = "SELECT 1 FROM like_post WHERE maphananh = %s AND cccd = %s"
        like_result = execute_query(query_like, (maphananh, user_cccd), fetch_one=True)
        user_liked = like_result is not None
    
    # Lấy boxchat_id nếu có
    query_boxchat = "SELECT maboxchat FROM boxchat WHERE maphananh = %s LIMIT 1"
    boxchat_result = execute_query(query_boxchat, (maphananh,), fetch_one=True)
    boxchat_id = boxchat_result[0] if boxchat_result else None
    
    # Lấy tên người dùng và tên vấn đề từ phananh tuple
    nguoidung_name = phananh[2]  # nguoi_tao từ query
    vande_name = phananh[16] if phananh[16] else None  # tenvande từ LEFT JOIN
    
    return render_template('phananh_detail.html', 
                         phananh=phananh, 
                         comments=comments,
                         user_liked=user_liked,
                         nguoidung_name=nguoidung_name,
                         vande_name=vande_name,
                         boxchat_id=boxchat_id)


@app.route('/phananh/edit/<int:maphananh>', methods=['GET', 'POST'])
@login_required
def phananh_edit(maphananh):
    """Chỉnh sửa phản ánh - Ownership check"""
    
    # Lấy thông tin phản ánh
    query_check = "SELECT cccd, trangthaiphananh FROM phananh WHERE maphananh = %s"
    phananh_check = execute_query(query_check, (maphananh,), fetch_one=True)
    
    if not phananh_check:
        flash('Không tìm thấy phản ánh!', 'danger')
        return redirect(url_for('phananh_list'))
    
    phananh_cccd = phananh_check[0]
    user_role = session['user'].get('vaitro')
    user_cccd = session['user'].get('cccd')
    
    # Kiểm tra quyền sở hữu
    if user_role == 'NguoiDan' and phananh_cccd != user_cccd:
        flash('Bạn chỉ có thể sửa phản ánh của chính mình!', 'danger')
        return redirect(url_for('phananh_detail', maphananh=maphananh))
    
    if request.method == 'POST':
        tieude = request.form.get('tieude', '').strip()
        mota = request.form.get('mota', '').strip()
        loaiphananh = request.form.get('loaiphananh', '').strip()
        
        # Cán bộ có thể cập nhật trạng thái
        if user_role in ['CanBo', 'QuanLy']:
            trangthaiphananh = request.form.get('trangthaiphananh', 'ChuaXuLy')
            query_update = """
                UPDATE phananh 
                SET tieude = %s, mota = %s, loaiphananh = %s, 
                    trangthaiphananh = %s, thoigianxuly = CURRENT_TIMESTAMP
                WHERE maphananh = %s
            """
            result = execute_query(query_update, (tieude, mota, loaiphananh, trangthaiphananh, maphananh))
        else:
            # Người dân chỉ sửa nội dung
            query_update = """
                UPDATE phananh 
                SET tieude = %s, mota = %s, loaiphananh = %s
                WHERE maphananh = %s
            """
            result = execute_query(query_update, (tieude, mota, loaiphananh, maphananh))
        
        if result:
            flash('Đã cập nhật phản ánh thành công!', 'success')
            return redirect(url_for('phananh_detail', maphananh=maphananh))
        else:
            flash('Có lỗi xảy ra khi cập nhật!', 'danger')
    
    # GET: Load dữ liệu
    query = """
        SELECT 
            p.maphananh,
            p.tieude,
            p.mota,
            p.loaiphananh,
            p.trangthaiphananh,
            p.is_public,
            p.allow_comment,
            d.tinh,
            d.xaphuong,
            d.chitiet
        FROM phananh p
        LEFT JOIN diachi d ON p.madiachi = d.madiachi
        WHERE p.maphananh = %s
    """
    phananh = execute_query(query, (maphananh,), fetch_one=True)
    
    return render_template('phananh_edit.html', phananh=phananh)


@app.route('/phananh/<int:maphananh>/chat')
@login_required
def phananh_chat(maphananh):
    """
    Chuyển trực tiếp đến trang chat của phản ánh
    Tự động tìm hoặc tạo boxchat nếu chưa có
    """
    user = session.get('user')
    cccd = user['cccd']
    user_role = user['vaitro']
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Kiểm tra phản ánh có tồn tại không
        cur.execute("SELECT maphananh, cccd, is_public FROM phananh WHERE maphananh = %s", (maphananh,))
        phananh = cur.fetchone()
        
        if not phananh:
            conn.close()
            flash('Phản ánh không tồn tại!', 'danger')
            return redirect(url_for('phananh_list'))
        
        phananh_cccd = phananh[1]
        is_public = phananh[2]
        
        # Kiểm tra quyền truy cập (nếu là phản ánh riêng tư)
        if not is_public and user_role == 'NguoiDan' and cccd != phananh_cccd:
            conn.close()
            flash('Bạn không có quyền truy cập phản ánh này!', 'danger')
            return redirect(url_for('phananh_list'))
        
        # Tìm boxchat của phản ánh này
        cur.execute("SELECT maboxchat FROM boxchat WHERE maphananh = %s", (maphananh,))
        boxchat = cur.fetchone()
        
        if boxchat:
            # Boxchat đã tồn tại
            maboxchat = boxchat[0]
        else:
            # Tạo boxchat mới
            cur.execute("""
                INSERT INTO boxchat (maphananh, cccd_nguoidan, cccd_canbo)
                VALUES (%s, %s, %s)
                RETURNING maboxchat
            """, (maphananh, phananh_cccd, cccd if user_role in ['CanBo', 'QuanLy'] else None))
            maboxchat = cur.fetchone()[0]
            conn.commit()
            print(f"DEBUG: Đã tạo boxchat mới {maboxchat} cho phản ánh {maphananh}")
        
        conn.close()
        
        # Chuyển đến trang chat
        return redirect(url_for('chat_detail', maboxchat=maboxchat))
    
    except Exception as e:
        conn.rollback()
        conn.close()
        print(f"ERROR: Lỗi khi tạo/tìm boxchat: {str(e)}")
        flash(f'Lỗi khi mở chat: {str(e)}', 'danger')
        return redirect(url_for('phananh_detail', maphananh=maphananh))


@app.route('/phananh/delete/<int:maphananh>', methods=['POST'])
@login_required
def phananh_delete(maphananh):
    """Xóa phản ánh - Ownership check"""
    
    # Lấy thông tin phản ánh
    query_check = "SELECT cccd, tieude FROM phananh WHERE maphananh = %s"
    phananh = execute_query(query_check, (maphananh,), fetch_one=True)
    
    if not phananh:
        flash('Không tìm thấy phản ánh!', 'danger')
        return redirect(url_for('phananh_list'))
    
    phananh_cccd = phananh[0]
    tieude = phananh[1]
    user_role = session['user'].get('vaitro')
    user_cccd = session['user'].get('cccd')
    
    # Kiểm tra quyền sở hữu
    if user_role == 'NguoiDan' and phananh_cccd != user_cccd:
        flash('Bạn chỉ có thể xóa phản ánh của chính mình!', 'danger')
        return redirect(url_for('phananh_list'))
    
    # Xóa phản ánh
    delete_query = "DELETE FROM phananh WHERE maphananh = %s"
    result = execute_query(delete_query, (maphananh,))
    
    if result:
        flash(f'Đã xóa phản ánh "{tieude}" thành công!', 'success')
    else:
        flash('Có lỗi xảy ra khi xóa!', 'danger')
    
    return redirect(url_for('phananh_list'))


# ========== PHASE 3: QUẢN LÝ VẤN ĐỀ ==========

@app.route('/vande')
@login_required
@role_required(['CanBo', 'QuanLy'])
def vande_list():
    """Danh sách vấn đề - Chỉ Cán bộ/Quản lý"""
    
    # Lấy tham số filter
    trangthai = request.args.get('trangthai', '').strip()
    phanloai = request.args.get('phanloai', '').strip()
    search = request.args.get('search', '').strip()
    
    # Pagination
    per_page = 20
    try:
        page = int(request.args.get('page', 1))
        if page < 1:
            page = 1
    except (ValueError, TypeError):
        page = 1
    
    offset = (page - 1) * per_page
    
    # Xây dựng điều kiện WHERE
    where_conditions = []
    params = []
    
    if trangthai:
        where_conditions.append("v.trangthai = %s")
        params.append(trangthai)
    
    if phanloai:
        where_conditions.append("v.phanloai = %s")
        params.append(phanloai)
    
    if search:
        where_conditions.append("v.tenvande ILIKE %s")
        params.append(f'%{search}%')
    
    where_clause = " AND ".join(where_conditions) if where_conditions else "1=1"
    
    # Đếm tổng số
    query_count = f"""
        SELECT COUNT(*) 
        FROM vande v
        WHERE {where_clause}
    """
    total_count = execute_query(query_count, tuple(params), fetch_one=True)
    total_count = total_count[0] if total_count else 0
    total_pages = math.ceil(total_count / per_page) if total_count > 0 else 1
    
    # Query danh sách vấn đề với thống kê
    query_list = f"""
        SELECT 
            v.mavande,
            v.tenvande,
            v.phanloai,
            v.trangthai,
            v.ngaytao,
            v.ngaycapnhat,
            n.name AS ten_canbo,
            COUNT(p.maphananh) AS so_phananh,
            COUNT(DISTINCT p.cccd) AS so_nguoi,
            COALESCE(SUM(p.like_count), 0) AS tong_like,
            COALESCE(SUM(p.comment_count), 0) AS tong_comment
        FROM vande v
        LEFT JOIN nguoidung n ON v.cccd_canbo_xuly = n.cccd
        LEFT JOIN phananh p ON v.mavande = p.mavande
        WHERE {where_clause}
        GROUP BY v.mavande, v.tenvande, v.phanloai, v.trangthai, 
                 v.ngaytao, v.ngaycapnhat, n.name
        ORDER BY v.ngaytao DESC
        LIMIT %s OFFSET %s
    """
    
    params_list = params + [per_page, offset]
    ds_vande = execute_query(query_list, tuple(params_list), fetch_all=True)
    ds_vande = ds_vande if ds_vande else []
    
    return render_template('vande_list.html',
                         ds_vande=ds_vande,
                         page=page,
                         per_page=per_page,
                         total_pages=total_pages,
                         trangthai=trangthai,
                         phanloai=phanloai,
                         search=search)


@app.route('/vande/add', methods=['GET', 'POST'])
@login_required
@role_required(['CanBo', 'QuanLy'])
def vande_add():
    """Tạo vấn đề mới và gộp phản ánh - Chỉ Cán bộ/Quản lý"""
    
    if request.method == 'POST':
        tenvande = request.form.get('tenvande', '').strip()
        phanloai = request.form.get('phanloai', 'Khac')
        # Chuyển giá trị tiếng Việt sang mã hợp lệ
        phanloai_map = {
            'Hạ tầng': 'HaTang',
            'Môi trường': 'MoiTruong',
            'An ninh': 'AnNinh',
            'Giao thông': 'GiaoThong',
            'Y tế': 'YTe',
            'Giáo dục': 'GiaoDuc',
            'Văn hóa': 'VanHoa',
            'Dân sinh': 'Khac',
            'Khác': 'Khac'
        }
        phanloai = phanloai_map.get(phanloai, 'Khac')
        
        # Lấy danh sách maphananh[] từ form
        list_maphananh = request.form.getlist('maphananh[]')
        
        # Validate
        if not tenvande:
            flash('Vui lòng nhập tên vấn đề!', 'warning')
            return redirect(url_for('vande_add'))
        
        if not list_maphananh:
            flash('Vui lòng chọn ít nhất 1 phản ánh để gộp!', 'warning')
            return redirect(url_for('vande_add'))
        
        # Kiểm tra các phản ánh chưa thuộc vấn đề nào
        placeholders = ','.join(['%s'] * len(list_maphananh))
        query_check = f"""
            SELECT maphananh, tieude 
            FROM phananh 
            WHERE maphananh IN ({placeholders}) AND mavande IS NOT NULL
        """
        da_gop = execute_query(query_check, tuple(list_maphananh), fetch_all=True)
        
        if da_gop:
            tieude_da_gop = ', '.join([str(row[1]) for row in da_gop])
            flash(f'Các phản ánh sau đã thuộc vấn đề khác: {tieude_da_gop}', 'danger')
            return redirect(url_for('vande_add'))
        
        # Tạo vấn đề mới
        cccd_canbo_xuly = session['user']['cccd']
        query_vande = """
            INSERT INTO vande (tenvande, phanloai, trangthai, cccd_canbo_xuly)
            VALUES (%s, %s, 'Moi', %s)
            RETURNING mavande
        """
        result = execute_query(
            query_vande,
            (tenvande, phanloai, cccd_canbo_xuly),
            fetch_one=True
        )
        
        if result:
            mavande = result[0]
            
            # Gộp các phản ánh vào vấn đề
            query_update = f"""
                UPDATE phananh 
                SET mavande = %s 
                WHERE maphananh IN ({placeholders})
            """
            params_update = [mavande] + list_maphananh
            execute_query(query_update, tuple(params_update))
            
            flash(f'Đã tạo vấn đề "{tenvande}" và gộp {len(list_maphananh)} phản ánh thành công!', 'success')
            return redirect(url_for('vande_detail', mavande=mavande))
        else:
            flash('Có lỗi xảy ra khi tạo vấn đề!', 'danger')
    
    # GET: Lấy danh sách phản ánh chưa gộp
    query_phananh = """
        SELECT 
            p.maphananh,
            p.tieude,
            p.loaiphananh,
            n.name AS nguoi_tao,
            p.thoigiantao,
            d.xaphuong,
            p.like_count,
            p.comment_count
        FROM phananh p
        LEFT JOIN nguoidung n ON p.cccd = n.cccd
        LEFT JOIN diachi d ON p.madiachi = d.madiachi
        WHERE p.mavande IS NULL
        ORDER BY p.thoigiantao DESC
        LIMIT 100
    """
    ds_phananh_chua_gop = execute_query(query_phananh, fetch_all=True)
    ds_phananh_chua_gop = ds_phananh_chua_gop if ds_phananh_chua_gop else []
    
    return render_template('vande_add.html', ds_phananh_chua_gop=ds_phananh_chua_gop)


@app.route('/vande/<int:mavande>')
@login_required
def vande_detail(mavande):
    """Chi tiết vấn đề - Cán bộ xem tất cả, Người dân chỉ xem vấn đề của phản ánh mình"""
    
    # Query thông tin vấn đề
    query_vande = """
        SELECT 
            v.mavande,
            v.tenvande,
            v.phanloai,
            v.trangthai,
            v.ketqua,
            v.ngaytao,
            v.ngaycapnhat,
            v.cccd_canbo_xuly,
            n.name AS ten_canbo,
            n.sdt AS sdt_canbo
        FROM vande v
        LEFT JOIN nguoidung n ON v.cccd_canbo_xuly = n.cccd
        WHERE v.mavande = %s
    """
    vande = execute_query(query_vande, (mavande,), fetch_one=True)
    
    if not vande:
        flash('Không tìm thấy vấn đề!', 'danger')
        return redirect(url_for('vande_list'))
    
    # Kiểm tra quyền xem
    user_role = session['user'].get('vaitro')
    user_cccd = session['user'].get('cccd')
    
    if user_role == 'NguoiDan':
        # Người dân chỉ xem được vấn đề của phản ánh mình
        query_check = """
            SELECT 1 FROM phananh 
            WHERE mavande = %s AND cccd = %s
        """
        has_permission = execute_query(query_check, (mavande, user_cccd), fetch_one=True)
        
        if not has_permission:
            flash('Bạn không có quyền xem vấn đề này!', 'danger')
            return redirect(url_for('phananh_list'))
    
    # Query danh sách phản ánh thuộc vấn đề
    query_phananh = """
        SELECT 
            p.maphananh,
            p.cccd,
            n.name AS nguoi_tao,
            p.tieude,
            p.loaiphananh,
            p.trangthaiphananh,
            p.thoigiantao,
            p.like_count,
            p.comment_count,
            p.view_count,
            d.xaphuong,
            d.chitiet
        FROM phananh p
        LEFT JOIN nguoidung n ON p.cccd = n.cccd
        LEFT JOIN diachi d ON p.madiachi = d.madiachi
        WHERE p.mavande = %s
        ORDER BY p.thoigiantao DESC
    """
    ds_phananh = execute_query(query_phananh, (mavande,), fetch_all=True)
    ds_phananh = ds_phananh if ds_phananh else []
    
    return render_template('vande_detail.html', vande=vande, ds_phananh=ds_phananh)


@app.route('/vande/<int:mavande>/update-status', methods=['POST'])
@login_required
@role_required(['CanBo', 'QuanLy'])
def vande_update_status(mavande):
    """
    Cập nhật trạng thái vấn đề và tự động:
    - Đồng bộ trạng thái sang tất cả phản ánh con
    - Gửi thông báo đến người dân
    - Gửi tin nhắn vào boxchat
    """
    
    trangthai_moi = request.form.get('trangthai', '').strip()
    ketqua = request.form.get('ketqua', '').strip()
    
    if not trangthai_moi:
        flash('Vui lòng chọn trạng thái!', 'warning')
        return redirect(url_for('vande_detail', mavande=mavande))
    
    # 1. Cập nhật vấn đề
    query_update_vande = """
        UPDATE vande 
        SET trangthai = %s, 
            ketqua = %s, 
            ngaycapnhat = CURRENT_TIMESTAMP 
        WHERE mavande = %s
        RETURNING tenvande
    """
    result = execute_query(query_update_vande, (trangthai_moi, ketqua, mavande), fetch_one=True)
    
    if not result:
        flash('Có lỗi xảy ra khi cập nhật vấn đề!', 'danger')
        return redirect(url_for('vande_detail', mavande=mavande))
    
    tenvande = result[0]
    
    # 2. Đồng bộ trạng thái sang tất cả phản ánh con
    query_update_phananh = """
        UPDATE phananh 
        SET trangthaiphananh = %s,
            thoigianxuly = CURRENT_TIMESTAMP
        WHERE mavande = %s
    """
    execute_query(query_update_phananh, (trangthai_moi, mavande))
    
    # 3. Lấy danh sách người dân bị ảnh hưởng
    query_affected_users = """
        SELECT DISTINCT p.cccd, n.name 
        FROM phananh p
        JOIN nguoidung n ON p.cccd = n.cccd
        WHERE p.mavande = %s
    """
    affected_users = execute_query(query_affected_users, (mavande,), fetch_all=True)
    affected_users = affected_users if affected_users else []
    
    # 4. Tạo nội dung thông báo
    noidung_thongbao = f"""
Vấn đề "{tenvande}" đã được cập nhật trạng thái: {trangthai_moi}.
"""
    
    if ketqua:
        noidung_thongbao += f"\nKết quả: {ketqua}"
    
    # 5. Gửi thông báo đến từng người
    for user in affected_users:
        cccd = user[0]
        
        # a) Insert vào bảng thông báo cá nhân
        query_notification = """
            INSERT INTO thongbao_nguoidung (cccd, noidung, loai, mavande)
            VALUES (%s, %s, 'VanDe', %s)
        """
        execute_query(query_notification, (cccd, noidung_thongbao, mavande))
        
        # b) Gửi tin nhắn vào boxchat (nếu tồn tại)
        query_boxchat = """
            SELECT b.maboxchat 
            FROM boxchat b
            JOIN phananh p ON b.maphananh = p.maphananh
            WHERE p.mavande = %s AND p.cccd = %s
            LIMIT 1
        """
        boxchat = execute_query(query_boxchat, (mavande, cccd), fetch_one=True)
        
        if boxchat:
            maboxchat = boxchat[0]
            query_insert_message = """
                INSERT INTO tinnhan (maboxchat, nguoigui, noidung, thoigiangui)
                VALUES (%s, %s, %s, CURRENT_TIMESTAMP)
            """
            execute_query(query_insert_message, (maboxchat, session['user']['cccd'], noidung_thongbao))
    
    flash(f'Đã cập nhật vấn đề và gửi thông báo đến {len(affected_users)} người dân!', 'success')
    return redirect(url_for('vande_detail', mavande=mavande))


@app.route('/vande/<int:mavande>/gop-them', methods=['POST'])
@login_required
@role_required(['CanBo', 'QuanLy'])
def vande_gop_them(mavande):
    """Gộp thêm phản ánh vào vấn đề đã có"""
    
    # Lấy danh sách maphananh[] từ form
    list_maphananh = request.form.getlist('maphananh[]')
    
    if not list_maphananh:
        flash('Vui lòng chọn ít nhất 1 phản ánh để gộp!', 'warning')
        return redirect(url_for('vande_detail', mavande=mavande))
    
    # Kiểm tra vấn đề tồn tại
    query_check_vande = "SELECT tenvande FROM vande WHERE mavande = %s"
    vande = execute_query(query_check_vande, (mavande,), fetch_one=True)
    
    if not vande:
        flash('Không tìm thấy vấn đề!', 'danger')
        return redirect(url_for('vande_list'))
    
    tenvande = vande[0]
    
    # Kiểm tra các phản ánh chưa thuộc vấn đề nào
    placeholders = ','.join(['%s'] * len(list_maphananh))
    query_check = f"""
        SELECT maphananh, tieude 
        FROM phananh 
        WHERE maphananh IN ({placeholders}) AND mavande IS NOT NULL
    """
    da_gop = execute_query(query_check, tuple(list_maphananh), fetch_all=True)
    
    if da_gop:
        tieude_da_gop = ', '.join([str(row[1]) for row in da_gop])
        flash(f'Các phản ánh sau đã thuộc vấn đề khác: {tieude_da_gop}', 'danger')
        return redirect(url_for('vande_detail', mavande=mavande))
    
    # Gộp các phản ánh vào vấn đề
    query_update = f"""
        UPDATE phananh 
        SET mavande = %s 
        WHERE maphananh IN ({placeholders})
    """
    params_update = [mavande] + list_maphananh
    result = execute_query(query_update, tuple(params_update))
    
    if result:
        flash(f'Đã gộp thêm {len(list_maphananh)} phản ánh vào vấn đề "{tenvande}"!', 'success')
    else:
        flash('Có lỗi xảy ra khi gộp phản ánh!', 'danger')
    
    return redirect(url_for('vande_detail', mavande=mavande))


@app.route('/vande/<int:mavande>/delete', methods=['POST'])
@login_required
@role_required(['CanBo', 'QuanLy'])
def vande_delete(mavande):
    """Xóa vấn đề - Chỉ Cán bộ/Quản lý"""
    
    # Lấy thông tin vấn đề
    query_vande = "SELECT tenvande FROM vande WHERE mavande = %s"
    vande = execute_query(query_vande, (mavande,), fetch_one=True)
    
    if not vande:
        flash('Không tìm thấy vấn đề!', 'danger')
        return redirect(url_for('vande_list'))
    
    tenvande = vande[0]
    
    # Hủy gộp các phản ánh (set mavande = NULL)
    query_ungop = "UPDATE phananh SET mavande = NULL WHERE mavande = %s"
    execute_query(query_ungop, (mavande,))
    
    # Xóa vấn đề
    query_delete = "DELETE FROM vande WHERE mavande = %s"
    result = execute_query(query_delete, (mavande,))
    
    if result:
        flash(f'Đã xóa vấn đề "{tenvande}" và hủy gộp các phản ánh!', 'success')
    else:
        flash('Có lỗi xảy ra khi xóa vấn đề!', 'danger')
    
    return redirect(url_for('vande_list'))


# ========== PHASE 4: SOCIAL FEATURES (LIKE, COMMENT, NEWS FEED) ==========

@app.route('/newsfeed')
@login_required
def newsfeed():
    """
    Hiển thị newsfeed - danh sách phản ánh công khai
    Sắp xếp theo hot_score (engagement) hoặc thoigiantao mới nhất
    """
    user = session.get('user')
    user_role = user['vaitro']
    
    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = 10
    offset = (page - 1) * per_page
    
    # Sort options: 'hot' (mặc định) hoặc 'new'
    sort_by = request.args.get('sort', 'hot')
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Query phản ánh công khai với thông tin đầy đủ
        if sort_by == 'new':
            order_clause = "ORDER BY p.thoigiantao DESC"
        else:
            # Hot score: like_count*2 + comment_count + view_count*0.1
            order_clause = "ORDER BY (p.like_count*2 + p.comment_count + p.view_count*0.1) DESC, p.thoigiantao DESC"
        
        # Lấy danh sách phản ánh
        cur.execute(f"""
            SELECT 
                p.maphananh,
                p.cccd,
                n.name AS nguoi_tao,
                p.tieude,
                p.mota,
                p.loaiphananh,
                p.trangthaiphananh,
                p.thoigiantao,
                p.like_count,
                p.comment_count,
                p.view_count,
                d.tinh,
                d.xaphuong,
                d.chitiet,
                n.avatar_url
            FROM phananh p
            JOIN nguoidung n ON p.cccd = n.cccd
            LEFT JOIN diachi d ON p.madiachi = d.madiachi
            WHERE p.is_public = TRUE
            {order_clause}
            LIMIT %s OFFSET %s
        """, (per_page, offset))
        
        phananh_list = cur.fetchall()
        
        # Đếm tổng số phản ánh công khai
        cur.execute("SELECT COUNT(*) FROM phananh WHERE is_public = TRUE")
        total = cur.fetchone()[0]
        total_pages = (total + per_page - 1) // per_page
        
        # Kiểm tra xem user đã like các post nào chưa
        if phananh_list:
            maphananh_list = [p[0] for p in phananh_list]  # Assuming maphananh is first column
            placeholders = ','.join(['%s'] * len(maphananh_list))
            cur.execute(f"""
                SELECT maphananh FROM like_post
                WHERE maphananh IN ({placeholders}) AND cccd = %s
            """, (*maphananh_list, user['cccd']))
            
            liked_posts = set(row[0] for row in cur.fetchall())
        else:
            liked_posts = set()
        
        conn.close()
        
        return render_template('newsfeed.html',
                             phananh_list=phananh_list,
                             liked_posts=liked_posts,
                             sort_by=sort_by,
                             page=page,
                             total_pages=total_pages)
    
    except Exception as e:
        conn.close()
        flash(f'Lỗi khi tải newsfeed: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))


@app.route('/phananh/<int:maphananh>/like', methods=['POST'])
@login_required
def phananh_like(maphananh):
    """
    Like một phản ánh
    """
    user = session.get('user')
    cccd = user['cccd']
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Kiểm tra phản ánh tồn tại
        cur.execute("SELECT maphananh FROM phananh WHERE maphananh = %s", (maphananh,))
        if not cur.fetchone():
            conn.close()
            flash('Phản ánh không tồn tại!', 'danger')
            return redirect(url_for('newsfeed'))
        
        # Kiểm tra đã like chưa
        cur.execute("""
            SELECT maphananh FROM like_post
            WHERE maphananh = %s AND cccd = %s
        """, (maphananh, cccd))
        
        if cur.fetchone():
            conn.close()
            flash('Bạn đã like phản ánh này rồi!', 'warning')
            return redirect(request.referrer or url_for('newsfeed'))
        
        # Thêm like (trigger sẽ tự động tăng like_count)
        cur.execute("""
            INSERT INTO like_post (maphananh, cccd, thoigian)
            VALUES (%s, %s, NOW())
        """, (maphananh, cccd))
        
        conn.commit()
        conn.close()
        
        flash('Đã like phản ánh!', 'success')
        return redirect(request.referrer or url_for('newsfeed'))
    
    except Exception as e:
        conn.rollback()
        conn.close()
        flash(f'Lỗi khi like phản ánh: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('newsfeed'))


@app.route('/phananh/<int:maphananh>/unlike', methods=['POST'])
@login_required
def phananh_unlike(maphananh):
    """
    Unlike một phản ánh
    """
    user = session.get('user')
    cccd = user['cccd']
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Kiểm tra đã like chưa
        cur.execute("""
            SELECT maphananh FROM like_post
            WHERE maphananh = %s AND cccd = %s
        """, (maphananh, cccd))
        
        if not cur.fetchone():
            conn.close()
            flash('Bạn chưa like phản ánh này!', 'warning')
            return redirect(request.referrer or url_for('newsfeed'))
        
        # Xóa like (trigger sẽ tự động giảm like_count)
        cur.execute("""
            DELETE FROM like_post
            WHERE maphananh = %s AND cccd = %s
        """, (maphananh, cccd))
        
        conn.commit()
        conn.close()
        
        flash('Đã bỏ like phản ánh!', 'success')
        return redirect(request.referrer or url_for('newsfeed'))
    
    except Exception as e:
        conn.rollback()
        conn.close()
        flash(f'Lỗi khi unlike phản ánh: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('newsfeed'))


@app.route('/phananh/<int:maphananh>/comment', methods=['POST'])
@login_required
def comment_add(maphananh):
    """
    Thêm bình luận cho phản ánh
    Support nested comments (parent_id)
    """
    user = session.get('user')
    cccd = user['cccd']
    noidung = request.form.get('noidung', '').strip()
    parent_id = request.form.get('parent_id', None)  # Cho nested reply
    
    if not noidung:
        flash('Nội dung bình luận không được để trống!', 'warning')
        return redirect(request.referrer or url_for('newsfeed'))
    
    # Xử lý file đính kèm
    matepdinhkem = None
    if 'file' in request.files:
        file = request.files['file']
        if file and file.filename:
            matepdinhkem = upload_file(file)
            if not matepdinhkem:
                flash('File không hợp lệ hoặc quá lớn. Cho phép: png, jpg, pdf, doc, xls (max 16MB)', 'warning')
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Kiểm tra phản ánh tồn tại
        cur.execute("SELECT maphananh FROM phananh WHERE maphananh = %s", (maphananh,))
        if not cur.fetchone():
            conn.close()
            flash('Phản ánh không tồn tại!', 'danger')
            return redirect(url_for('newsfeed'))
        
        # Nếu có parent_id, kiểm tra comment cha tồn tại
        if parent_id:
            cur.execute("""
                SELECT id FROM binhluan
                WHERE id = %s AND maphananh = %s
            """, (parent_id, maphananh))
            if not cur.fetchone():
                conn.close()
                flash('Bình luận cha không tồn tại!', 'danger')
                return redirect(request.referrer or url_for('newsfeed'))
        
        # Thêm bình luận
        if parent_id:
            cur.execute("""
                INSERT INTO binhluan (maphananh, cccd_nguoidung, noidung, thoigian, parent_id, matepdinhkem)
                VALUES (%s, %s, %s, NOW(), %s, %s)
            """, (maphananh, cccd, noidung, parent_id, matepdinhkem))
        else:
            cur.execute("""
                INSERT INTO binhluan (maphananh, cccd_nguoidung, noidung, thoigian, matepdinhkem)
                VALUES (%s, %s, %s, NOW(), %s)
            """, (maphananh, cccd, noidung, matepdinhkem))
        
        # Cập nhật comment_count = đếm lại số comment không bị ẩn
        cur.execute("""
            UPDATE phananh
            SET comment_count = (
                SELECT COUNT(*) FROM binhluan 
                WHERE maphananh = %s AND is_hidden = FALSE
            )
            WHERE maphananh = %s
        """, (maphananh, maphananh))
        
        conn.commit()
        conn.close()
        
        flash('Đã thêm bình luận!', 'success')
        return redirect(request.referrer or url_for('phananh_detail', maphananh=maphananh))
    
    except Exception as e:
        conn.rollback()
        conn.close()
        flash(f'Lỗi khi thêm bình luận: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('newsfeed'))


@app.route('/phananh/<int:maphananh>/comments')
@login_required
def comment_list(maphananh):
    """
    Lấy danh sách bình luận của phản ánh
    Trả về JSON cho AJAX loading
    """
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Lấy comments (không bao gồm hidden comments)
        cur.execute("""
            SELECT 
                b.mabinhluan,
                b.noidung,
                b.thoigian,
                b.parent_id,
                b.is_hidden,
                n.cccd,
                n.hovaten
            FROM binhluan b
            JOIN nguoidung n ON b.cccd = n.cccd
            WHERE b.maphananh = %s AND b.is_hidden = FALSE
            ORDER BY b.thoigian ASC
        """, (maphananh,))
        
        comments = cur.fetchall()
        conn.close()
        
        # Format response
        comments_list = []
        for c in comments:
            comments_list.append({
                'mabinhluan': c[0],
                'noidung': c[1],
                'thoigian': c[2].strftime('%d/%m/%Y %H:%M') if c[2] else '',
                'parent_id': c[3],
                'cccd': c[5],
                'hovaten': c[6]
            })
        
        return jsonify({'success': True, 'comments': comments_list})
    
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': str(e)})


@app.route('/comment/<int:mabinhluan>/hide', methods=['POST'])
@login_required
@role_required(['CanBo', 'QuanLy'])
def comment_hide(mabinhluan):
    """
    Ẩn bình luận vi phạm (chỉ CanBo/QuanLy)
    """
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Kiểm tra comment tồn tại
        cur.execute("""
            SELECT maphananh FROM binhluan
            WHERE id = %s
        """, (mabinhluan,))
        
        result = cur.fetchone()
        if not result:
            conn.close()
            flash('Bình luận không tồn tại!', 'danger')
            return redirect(request.referrer or url_for('newsfeed'))
        
        maphananh = result[0]
        
        # Ẩn comment
        cur.execute("""
            UPDATE binhluan
            SET is_hidden = TRUE
            WHERE id = %s
        """, (mabinhluan,))
        
        # Cập nhật comment_count = đếm lại số comment không bị ẩn
        cur.execute("""
            UPDATE phananh
            SET comment_count = (
                SELECT COUNT(*) FROM binhluan 
                WHERE maphananh = %s AND is_hidden = FALSE
            )
            WHERE maphananh = %s
        """, (maphananh, maphananh))
        
        conn.commit()
        conn.close()
        
        flash('Đã ẩn bình luận vi phạm!', 'success')
        return redirect(request.referrer or url_for('phananh_detail', maphananh=maphananh))
    
    except Exception as e:
        conn.rollback()
        conn.close()
        flash(f'Lỗi khi ẩn bình luận: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('newsfeed'))


@app.route('/comment/<int:mabinhluan>/edit', methods=['GET', 'POST'])
@login_required
def comment_edit(mabinhluan):
    """
    Chỉnh sửa bình luận của chính mình
    """
    user_cccd = session['user']['cccd']
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Lấy thông tin comment
        cur.execute("""
            SELECT id, maphananh, cccd_nguoidung, noidung 
            FROM binhluan
            WHERE id = %s AND is_hidden = FALSE
        """, (mabinhluan,))
        
        comment = cur.fetchone()
        
        if not comment:
            conn.close()
            flash('Bình luận không tồn tại!', 'danger')
            return redirect(request.referrer or url_for('newsfeed'))
        
        comment_cccd = comment[2]
        maphananh = comment[1]
        
        # Kiểm tra ownership
        if user_cccd != comment_cccd:
            conn.close()
            flash('Bạn không có quyền chỉnh sửa bình luận này!', 'danger')
            return redirect(url_for('phananh_detail', maphananh=maphananh))
        
        if request.method == 'POST':
            noidung = request.form.get('noidung', '').strip()
            
            if not noidung:
                flash('Nội dung bình luận không được để trống!', 'warning')
                return redirect(url_for('phananh_detail', maphananh=maphananh))
            
            # Cập nhật bình luận
            cur.execute("""
                UPDATE binhluan
                SET noidung = %s, thoigian = NOW()
                WHERE id = %s
            """, (noidung, mabinhluan))
            
            conn.commit()
            conn.close()
            
            flash('Đã cập nhật bình luận!', 'success')
            return redirect(url_for('phananh_detail', maphananh=maphananh))
        
        conn.close()
        # GET request - return to detail page (edit inline in template)
        return redirect(url_for('phananh_detail', maphananh=maphananh))
    
    except Exception as e:
        conn.rollback()
        conn.close()
        flash(f'Lỗi khi chỉnh sửa bình luận: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('newsfeed'))


@app.route('/comment/<int:mabinhluan>/delete', methods=['POST'])
@login_required
def comment_delete(mabinhluan):
    """
    Xóa bình luận của chính mình
    """
    user_cccd = session['user']['cccd']
    user_role = session['user']['vaitro']
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Lấy thông tin comment
        cur.execute("""
            SELECT id, maphananh, cccd_nguoidung 
            FROM binhluan
            WHERE id = %s
        """, (mabinhluan,))
        
        comment = cur.fetchone()
        
        if not comment:
            conn.close()
            flash('Bình luận không tồn tại!', 'danger')
            return redirect(request.referrer or url_for('newsfeed'))
        
        comment_cccd = comment[2]
        maphananh = comment[1]
        
        # Kiểm tra ownership (chủ comment hoặc CanBo/QuanLy)
        if user_cccd != comment_cccd and user_role not in ['CanBo', 'QuanLy']:
            conn.close()
            flash('Bạn không có quyền xóa bình luận này!', 'danger')
            return redirect(url_for('phananh_detail', maphananh=maphananh))
        
        # Xóa bình luận (CASCADE sẽ xóa các reply)
        cur.execute("DELETE FROM binhluan WHERE id = %s", (mabinhluan,))
        
        # Cập nhật comment_count
        cur.execute("""
            UPDATE phananh
            SET comment_count = (
                SELECT COUNT(*) FROM binhluan 
                WHERE maphananh = %s AND is_hidden = FALSE
            )
            WHERE maphananh = %s
        """, (maphananh, maphananh))
        
        conn.commit()
        conn.close()
        
        flash('Đã xóa bình luận!', 'success')
        return redirect(url_for('phananh_detail', maphananh=maphananh))
    
    except Exception as e:
        conn.rollback()
        conn.close()
        flash(f'Lỗi khi xóa bình luận: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('newsfeed'))


# ========== PHASE 5: CHAT & NOTIFICATION SYSTEM ==========

@app.route('/chat')
@login_required
def chat_list():
    """
    Danh sách các boxchat của user
    Hiển thị tin nhắn cuối cùng và số tin chưa đọc
    """
    user = session.get('user')
    cccd = user['cccd']
    user_role = user['vaitro']
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Lấy danh sách boxchat mà user tham gia
        # JOIN với tin nhắn cuối cùng và đếm tin chưa đọc
        if user_role in ['CanBo', 'QuanLy']:
            # Cán bộ/Quản lý thấy tất cả boxchat
            cur.execute("""
                WITH latest_messages AS (
                    SELECT DISTINCT ON (maboxchat)
                        maboxchat,
                        noidung as last_message,
                        thoigiangui as last_time,
                        nguoigui as last_sender
                    FROM tinnhan
                    ORDER BY maboxchat, thoigiangui DESC
                ),
                unread_counts AS (
                    SELECT 
                        maboxchat,
                        COUNT(*) as unread_count
                    FROM tinnhan
                    WHERE dadoc = FALSE AND nguoigui != %s
                    GROUP BY maboxchat
                )
                SELECT 
                    b.maboxchat,
                    b.maphananh,
                    p.tieude as phananh_title,
                    lm.last_message,
                    lm.last_time,
                    lm.last_sender,
                    n.name as last_sender_name,
                    COALESCE(uc.unread_count, 0) as unread_count,
                    b.cccd_canbo,
                    nd.name as canbo_name
                FROM boxchat b
                LEFT JOIN phananh p ON b.maphananh = p.maphananh
                LEFT JOIN latest_messages lm ON b.maboxchat = lm.maboxchat
                LEFT JOIN nguoidung n ON lm.last_sender = n.cccd
                LEFT JOIN unread_counts uc ON b.maboxchat = uc.maboxchat
                LEFT JOIN nguoidung nd ON b.cccd_canbo = nd.cccd
                ORDER BY lm.last_time DESC NULLS LAST
            """, (cccd,))
        else:
            # Người dân chỉ thấy boxchat của mình
            cur.execute("""
                WITH latest_messages AS (
                    SELECT DISTINCT ON (maboxchat)
                        maboxchat,
                        noidung as last_message,
                        thoigiangui as last_time,
                        nguoigui as last_sender
                    FROM tinnhan
                    ORDER BY maboxchat, thoigiangui DESC
                ),
                unread_counts AS (
                    SELECT 
                        maboxchat,
                        COUNT(*) as unread_count
                    FROM tinnhan
                    WHERE dadoc = FALSE AND nguoigui != %s
                    GROUP BY maboxchat
                )
                SELECT 
                    b.maboxchat,
                    b.maphananh,
                    p.tieude as phananh_title,
                    lm.last_message,
                    lm.last_time,
                    lm.last_sender,
                    n.name as last_sender_name,
                    COALESCE(uc.unread_count, 0) as unread_count,
                    b.cccd_canbo,
                    nd.name as canbo_name
                FROM boxchat b
                LEFT JOIN phananh p ON b.maphananh = p.maphananh
                LEFT JOIN latest_messages lm ON b.maboxchat = lm.maboxchat
                LEFT JOIN nguoidung n ON lm.last_sender = n.cccd
                LEFT JOIN unread_counts uc ON b.maboxchat = uc.maboxchat
                LEFT JOIN nguoidung nd ON b.cccd_canbo = nd.cccd
                WHERE b.cccd_nguoidan = %s OR b.cccd_canbo = %s
                ORDER BY lm.last_time DESC NULLS LAST
            """, (cccd, cccd, cccd))
        
        boxchat_list = cur.fetchall()
        conn.close()
        
        return render_template('chat_list.html', boxchat_list=boxchat_list)
    
    except Exception as e:
        conn.close()
        flash(f'Lỗi khi tải danh sách chat: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))


@app.route('/chat/<int:maboxchat>')
@login_required
def chat_detail(maboxchat):
    """
    Chi tiết boxchat - hiển thị tất cả tin nhắn
    """
    user = session.get('user')
    cccd = user['cccd']
    user_role = user['vaitro']
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Kiểm tra quyền truy cập boxchat
        cur.execute("""
            SELECT 
                b.maboxchat,
                b.maphananh,
                b.cccd_nguoidan,
                b.cccd_canbo,
                p.tieude,
                n1.name as nguoidung_name,
                n2.name as canbo_name
            FROM boxchat b
            LEFT JOIN phananh p ON b.maphananh = p.maphananh
            LEFT JOIN nguoidung n1 ON b.cccd_nguoidan = n1.cccd
            LEFT JOIN nguoidung n2 ON b.cccd_canbo = n2.cccd
            WHERE b.maboxchat = %s
        """, (maboxchat,))
        
        boxchat = cur.fetchone()
        
        if not boxchat:
            conn.close()
            flash('Boxchat không tồn tại!', 'danger')
            return redirect(url_for('chat_list'))
        
        # Kiểm tra quyền truy cập
        cccd_nguoidung = boxchat[2]
        cccd_canbo = boxchat[3]
        
        # Cán bộ/Quản lý có thể truy cập mọi boxchat
        if user_role in ['CanBo', 'QuanLy']:
            # Nếu chưa có cán bộ được assign, tự động assign cán bộ này vào
            if not cccd_canbo:
                cur.execute("""
                    UPDATE boxchat
                    SET cccd_canbo = %s
                    WHERE maboxchat = %s
                """, (cccd, maboxchat))
                conn.commit()
                print(f"DEBUG: Đã assign cán bộ {cccd} vào boxchat {maboxchat}")
                # Reload boxchat data
                cur.execute("""
                    SELECT 
                        b.maboxchat,
                        b.maphananh,
                        b.cccd_nguoidan,
                        b.cccd_canbo,
                        p.tieude,
                        n1.name as nguoidung_name,
                        n2.name as canbo_name
                    FROM boxchat b
                    LEFT JOIN phananh p ON b.maphananh = p.maphananh
                    LEFT JOIN nguoidung n1 ON b.cccd_nguoidan = n1.cccd
                    LEFT JOIN nguoidung n2 ON b.cccd_canbo = n2.cccd
                    WHERE b.maboxchat = %s
                """, (maboxchat,))
                boxchat = cur.fetchone()
        elif cccd not in [cccd_nguoidung, cccd_canbo]:
            # Người dân chỉ được xem boxchat của mình
            conn.close()
            flash('Bạn không có quyền truy cập boxchat này!', 'danger')
            return redirect(url_for('chat_list'))
        
        # Lấy tất cả tin nhắn
        cur.execute("""
            SELECT 
                t.tinnhanid,
                t.nguoigui,
                t.noidung,
                t.thoigiangui,
                t.dadoc,
                n.name as sender_name,
                td.duongdan as file_dinh_kem,
                t.matepdinhkem
            FROM tinnhan t
            JOIN nguoidung n ON t.nguoigui = n.cccd
            LEFT JOIN tepdinhkem td ON t.matepdinhkem = td.matepdinhkem
            WHERE t.maboxchat = %s
            ORDER BY t.thoigiangui ASC
        """, (maboxchat,))
        
        messages = cur.fetchall()
        
        # Đánh dấu tất cả tin nhắn là đã đọc (trừ tin nhắn do mình gửi)
        cur.execute("""
            UPDATE tinnhan
            SET dadoc = TRUE
            WHERE maboxchat = %s AND nguoigui != %s AND dadoc = FALSE
        """, (maboxchat, cccd))
        
        conn.commit()
        conn.close()
        
        return render_template('chat_detail.html',
                             boxchat=boxchat,
                             messages=messages,
                             current_cccd=cccd)
    
    except Exception as e:
        conn.close()
        flash(f'Lỗi khi tải chat: {str(e)}', 'danger')
        return redirect(url_for('chat_list'))


@app.route('/chat/<int:maboxchat>/send', methods=['POST'])
@login_required
def chat_send_message(maboxchat):
    """
    Gửi tin nhắn trong boxchat
    """
    user = session.get('user')
    cccd = user['cccd']
    noidung = request.form.get('noidung', '').strip()
    
    if not noidung:
        flash('Nội dung tin nhắn không được để trống!', 'warning')
        return redirect(url_for('chat_detail', maboxchat=maboxchat))
    
    # Xử lý file đính kèm
    matepdinhkem = None
    if 'file' in request.files:
        file = request.files['file']
        if file and file.filename:
            matepdinhkem = upload_file(file)
            if not matepdinhkem:
                flash('File không hợp lệ hoặc quá lớn. Cho phép: png, jpg, pdf, doc, xls (max 16MB)', 'warning')
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Kiểm tra quyền gửi tin nhắn
        cur.execute("""
            SELECT cccd_nguoidan, cccd_canbo
            FROM boxchat
            WHERE maboxchat = %s
        """, (maboxchat,))
        
        boxchat = cur.fetchone()
        
        if not boxchat:
            conn.close()
            flash('Boxchat không tồn tại!', 'danger')
            return redirect(url_for('chat_list'))
        
        cccd_nguoidung = boxchat[0]
        cccd_canbo = boxchat[1]
        
        # Cán bộ/Quản lý có thể gửi tin nhắn trong mọi boxchat
        if user['vaitro'] not in ['CanBo', 'QuanLy']:
            # Người dân chỉ gửi được trong boxchat của mình
            if cccd not in [cccd_nguoidung, cccd_canbo]:
                conn.close()
                flash('Bạn không có quyền gửi tin nhắn trong boxchat này!', 'danger')
                return redirect(url_for('chat_list'))
        
        # Thêm tin nhắn
        cur.execute("""
            INSERT INTO tinnhan (maboxchat, nguoigui, noidung, thoigiangui, dadoc, matepdinhkem)
            VALUES (%s, %s, %s, NOW(), FALSE, %s)
        """, (maboxchat, cccd, noidung, matepdinhkem))
        
        conn.commit()
        conn.close()
        
        flash('Đã gửi tin nhắn!', 'success')
        return redirect(url_for('chat_detail', maboxchat=maboxchat))
    
    except Exception as e:
        conn.rollback()
        conn.close()
        flash(f'Lỗi khi gửi tin nhắn: {str(e)}', 'danger')
        return redirect(url_for('chat_detail', maboxchat=maboxchat))


@app.route('/notifications')
@login_required
def notifications():
    """
    Danh sách thông báo của user
    """
    user = session.get('user')
    cccd = user['cccd']
    
    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = 20
    offset = (page - 1) * per_page
    
    # Filter: 'all' hoặc 'unread'
    filter_type = request.args.get('filter', 'all')
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Build WHERE clause
        where_clause = "WHERE tn.cccd = %s"
        params = [cccd]
        
        if filter_type == 'unread':
            where_clause += " AND tn.trangthai_doc = FALSE"
        
        # Lấy danh sách thông báo
        cur.execute(f"""
            SELECT 
                tn.mathongbao_nguoidung,
                tn.noidung,
                tn.loai,
                tn.thoigian,
                tn.trangthai_doc,
                tn.mavande,
                tn.maphananh,
                v.tenvande,
                p.tieude as phananh_title
            FROM thongbao_nguoidung tn
            LEFT JOIN vande v ON tn.mavande = v.mavande
            LEFT JOIN phananh p ON tn.maphananh = p.maphananh
            {where_clause}
            ORDER BY tn.thoigian DESC
            LIMIT %s OFFSET %s
        """, params + [per_page, offset])
        
        notification_list = cur.fetchall()
        
        # Đếm tổng số thông báo
        cur.execute(f"""
            SELECT COUNT(*)
            FROM thongbao_nguoidung tn
            {where_clause}
        """, params)
        
        total = cur.fetchone()[0]
        total_pages = (total + per_page - 1) // per_page
        
        # Đếm số thông báo chưa đọc
        cur.execute("""
            SELECT COUNT(*)
            FROM thongbao_nguoidung
            WHERE cccd = %s AND trangthai_doc = FALSE
        """, (cccd,))
        
        unread_count = cur.fetchone()[0]
        
        conn.close()
        
        return render_template('notifications.html',
                             notification_list=notification_list,
                             unread_count=unread_count,
                             filter_type=filter_type,
                             page=page,
                             total_pages=total_pages)
    
    except Exception as e:
        conn.close()
        flash(f'Lỗi khi tải thông báo: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))


@app.route('/notification/<int:mathongbao_nguoidung>/read', methods=['POST'])
@login_required
def notification_mark_read(mathongbao_nguoidung):
    """
    Đánh dấu một thông báo là đã đọc
    """
    user = session.get('user')
    cccd = user['cccd']
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Kiểm tra quyền sở hữu
        cur.execute("""
            SELECT mathongbao_nguoidung
            FROM thongbao_nguoidung
            WHERE mathongbao_nguoidung = %s AND cccd = %s
        """, (mathongbao_nguoidung, cccd))
        
        if not cur.fetchone():
            conn.close()
            return jsonify({'success': False, 'message': 'Thông báo không tồn tại hoặc không thuộc về bạn!'})
        
        # Đánh dấu đã đọc
        cur.execute("""
            UPDATE thongbao_nguoidung
            SET trangthai_doc = TRUE
            WHERE mathongbao_nguoidung = %s
        """, (mathongbao_nguoidung,))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Đã đánh dấu đã đọc!'})
    
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': str(e)})


@app.route('/notifications/mark_all_read', methods=['POST'])
@login_required
def notifications_mark_all_read():
    """
    Đánh dấu tất cả thông báo là đã đọc
    """
    user = session.get('user')
    cccd = user['cccd']
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Đánh dấu tất cả thông báo chưa đọc
        cur.execute("""
            UPDATE thongbao_nguoidung
            SET trangthai_doc = TRUE
            WHERE cccd = %s AND trangthai_doc = FALSE
        """, (cccd,))
        
        rows_affected = cur.rowcount
        
        conn.commit()
        conn.close()
        
        flash(f'Đã đánh dấu {rows_affected} thông báo là đã đọc!', 'success')
        return redirect(url_for('notifications'))
    
    except Exception as e:
        conn.rollback()
        conn.close()
        flash(f'Lỗi khi đánh dấu thông báo: {str(e)}', 'danger')
        return redirect(url_for('notifications'))


# ========== QUẢN LÝ THÔNG BÁO CHUNG ==========

@app.route('/thongbao-chung')
@login_required
def thongbao_chung_list():
    """
    Danh sách thông báo chung (bảng tin công khai)
    """
    page = request.args.get('page', 1, type=int)
    per_page = 10
    offset = (page - 1) * per_page
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Lấy danh sách thông báo đang hiển thị
        cur.execute("""
            SELECT 
                tb.mathongbao,
                tb.tieude,
                tb.noidung,
                tb.thoigiandang,
                tb.nguoidang,
                n.name as ten_nguoidang,
                tb.trangthai
            FROM thongbao tb
            LEFT JOIN nguoidung n ON tb.nguoidang = n.cccd
            WHERE tb.trangthai = 'HienThi'
            ORDER BY tb.thoigiandang DESC
            LIMIT %s OFFSET %s
        """, (per_page, offset))
        
        thongbao_list = cur.fetchall()
        
        # Đếm tổng số
        cur.execute("""
            SELECT COUNT(*)
            FROM thongbao
            WHERE trangthai = 'HienThi'
        """)
        
        total = cur.fetchone()[0]
        total_pages = (total + per_page - 1) // per_page
        
        conn.close()
        
        return render_template('thongbao_chung.html',
                             thongbao_list=thongbao_list,
                             page=page,
                             total_pages=total_pages)
    
    except Exception as e:
        conn.close()
        flash(f'Lỗi khi tải thông báo: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))


@app.route('/thongbao-chung/tao', methods=['GET', 'POST'])
@login_required
@role_required(['CanBo', 'QuanLy'])
def thongbao_chung_tao():
    """
    Tạo thông báo chung (chỉ cán bộ/quản lý)
    """
    if request.method == 'GET':
        return render_template('thongbao_chung_tao.html')
    
    user = session.get('user')
    cccd = user['cccd']
    tieude = request.form.get('tieude', '').strip()
    noidung = request.form.get('noidung', '').strip()
    gui_cho_tat_ca = request.form.get('gui_cho_tat_ca') == 'on'
    
    if not tieude or not noidung:
        flash('Vui lòng nhập đầy đủ tiêu đề và nội dung!', 'warning')
        return redirect(url_for('thongbao_chung_tao'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # 1. Tạo thông báo chung
        cur.execute("""
            INSERT INTO thongbao (tieude, noidung, nguoidang, trangthai)
            VALUES (%s, %s, %s, 'HienThi')
            RETURNING mathongbao
        """, (tieude, noidung, cccd))
        
        mathongbao = cur.fetchone()[0]
        
        # 2. Nếu chọn gửi cho tất cả → tạo thông báo cá nhân
        if gui_cho_tat_ca:
            cur.execute("SELECT cccd FROM nguoidung WHERE vaitro = 'NguoiDan'")
            nguoidan_list = cur.fetchall()
            
            for (cccd_nguoidan,) in nguoidan_list:
                cur.execute("""
                    INSERT INTO thongbao_nguoidung 
                    (cccd, mathongbao, noidung, loai)
                    VALUES (%s, %s, %s, 'General')
                """, (cccd_nguoidan, mathongbao, f"{tieude}: {noidung}"))
        
        conn.commit()
        conn.close()
        
        if gui_cho_tat_ca:
            flash(f'Đã tạo thông báo và gửi cho tất cả người dân!', 'success')
        else:
            flash('Đã tạo thông báo chung!', 'success')
        
        return redirect(url_for('thongbao_chung_list'))
    
    except Exception as e:
        conn.rollback()
        conn.close()
        flash(f'Lỗi khi tạo thông báo: {str(e)}', 'danger')
        return redirect(url_for('thongbao_chung_tao'))


@app.route('/thongbao-chung/<int:mathongbao>/an', methods=['POST'])
@login_required
@role_required(['QuanLy'])
def thongbao_chung_an(mathongbao):
    """
    Ẩn thông báo chung (chỉ quản lý)
    """
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            UPDATE thongbao
            SET trangthai = 'An'
            WHERE mathongbao = %s
        """, (mathongbao,))
        
        conn.commit()
        conn.close()
        
        flash('Đã ẩn thông báo!', 'success')
        return redirect(url_for('thongbao_chung_list'))
    
    except Exception as e:
        conn.rollback()
        conn.close()
        flash(f'Lỗi: {str(e)}', 'danger')
        return redirect(url_for('thongbao_chung_list'))


@app.route('/api/unread-count')
@login_required
def api_unread_count():
    """
    API lấy số lượng thông báo chưa đọc (dùng cho badge)
    """
    user = session.get('user')
    cccd = user['cccd']
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT COUNT(*)
            FROM thongbao_nguoidung
            WHERE cccd = %s AND trangthai_doc = FALSE
        """, (cccd,))
        
        count = cur.fetchone()[0]
        conn.close()
        
        return jsonify({'success': True, 'unread_count': count})
    
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({'success': False, 'message': str(e)})


# ========== PHASE 6: REPORTS & STATISTICS ==========

@app.route('/reports/overview')
@login_required
@role_required(['CanBo', 'QuanLy'])
def reports_overview():
    """
    Tổng quan thống kê về phản ánh và vấn đề
    """
    # Tham số thời gian
    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Build WHERE clause cho filter thời gian
        where_clause_phananh = ""
        where_clause_vande = ""
        params = []
        
        if from_date and to_date:
            where_clause_phananh = "WHERE thoigiantao BETWEEN %s AND %s"
            where_clause_vande = "WHERE ngaytao BETWEEN %s AND %s"
            params = [from_date, to_date]
        
        # 1. Thống kê phản ánh
        cur.execute(f"""
            SELECT 
                COUNT(*) as total,
                COUNT(CASE WHEN trangthaiphananh = 'Mới' THEN 1 END) as moi,
                COUNT(CASE WHEN trangthaiphananh = 'Đang xử lý' THEN 1 END) as dangxuly,
                COUNT(CASE WHEN trangthaiphananh = 'Đã xử lý' THEN 1 END) as daxuly,
                COUNT(CASE WHEN trangthaiphananh = 'Đã từ chối' THEN 1 END) as tuchoi,
                COUNT(CASE WHEN is_public = TRUE THEN 1 END) as public_count,
                SUM(like_count) as total_likes,
                SUM(comment_count) as total_comments,
                SUM(view_count) as total_views
            FROM phananh
            {where_clause_phananh}
        """, params)
        
        phananh_stats = cur.fetchone()
        
        # 2. Thống kê vấn đề
        cur.execute(f"""
            SELECT 
                COUNT(*) as total,
                COUNT(CASE WHEN trangthai = 'Mới' THEN 1 END) as moi,
                COUNT(CASE WHEN trangthai = 'Đang xử lý' THEN 1 END) as dangxuly,
                COUNT(CASE WHEN trangthai = 'Đã giải quyết' THEN 1 END) as dagiaiquyet,
                COUNT(CASE WHEN trangthai = 'Đóng' THEN 1 END) as dong
            FROM vande
            {where_clause_vande}
        """, params)
        
        vande_stats = cur.fetchone()
        
        # 3. Thống kê theo phân loại vấn đề
        cur.execute(f"""
            SELECT 
                phanloai,
                COUNT(*) as count
            FROM vande
            {where_clause_vande}
            GROUP BY phanloai
            ORDER BY count DESC
        """, params)
        
        phanloai_stats = cur.fetchall()
        
        # 4. Xu hướng theo tháng (12 tháng gần nhất)
        cur.execute("""
            SELECT 
                TO_CHAR(thoigiantao, 'YYYY-MM') as month,
                COUNT(*) as count
            FROM phananh
            WHERE thoigiantao >= NOW() - INTERVAL '12 months'
            GROUP BY month
            ORDER BY month DESC
            LIMIT 12
        """)
        
        monthly_trend = cur.fetchall()
        
        # 5. Top 5 cán bộ xử lý nhiều vấn đề nhất
        cur.execute(f"""
            SELECT 
                n.name,
                n.cccd,
                COUNT(v.mavande) as solved_count
            FROM vande v
            JOIN nguoidung n ON v.cccd_canbo_xuly = n.cccd
            {where_clause_vande}
            GROUP BY n.name, n.cccd
            ORDER BY solved_count DESC
            LIMIT 5
        """, params)
        
        top_canbo = cur.fetchall()
        
        conn.close()
        
        return render_template('reports_overview.html',
                             phananh_stats=phananh_stats,
                             vande_stats=vande_stats,
                             phanloai_stats=phanloai_stats,
                             monthly_trend=monthly_trend,
                             top_canbo=top_canbo,
                             from_date=from_date,
                             to_date=to_date)
    
    except Exception as e:
        conn.close()
        flash(f'Lỗi khi tải báo cáo: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))


@app.route('/reports/phananh')
@login_required
@role_required(['CanBo', 'QuanLy'])
def reports_phananh():
    """
    Thống kê chi tiết về phản ánh
    """
    # Tham số
    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')
    trangthai = request.args.get('trangthai', '')
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Build WHERE clause
        where_parts = []
        params = []
        
        if from_date and to_date:
            where_parts.append("thoigiantao BETWEEN %s AND %s")
            params.extend([from_date, to_date])
        
        if trangthai:
            where_parts.append("trangthaiphananh = %s")
            params.append(trangthai)
        
        where_clause = "WHERE " + " AND ".join(where_parts) if where_parts else ""
        
        # 1. Thống kê theo trạng thái
        cur.execute(f"""
            SELECT 
                trangthaiphananh,
                COUNT(*) as count,
                ROUND(COUNT(*) * 100.0 / SUM(COUNT(*)) OVER(), 2) as percentage
            FROM phananh
            {where_clause}
            GROUP BY trangthaiphananh
            ORDER BY count DESC
        """, params)
        
        status_stats = cur.fetchall()
        
        # 2. Thống kê theo ngày (7 ngày gần nhất)
        cur.execute(f"""
            SELECT 
                DATE(thoigiantao) as date,
                COUNT(*) as count,
                COUNT(CASE WHEN is_public = TRUE THEN 1 END) as public_count
            FROM phananh
            {where_clause}
            GROUP BY DATE(thoigiantao)
            ORDER BY date DESC
            LIMIT 30
        """, params)
        
        daily_stats = cur.fetchall()
        
        # 3. Top 10 phản ánh có engagement cao nhất
        cur.execute(f"""
            SELECT 
                maphananh,
                tieude,
                like_count,
                comment_count,
                view_count,
                (like_count * 2 + comment_count + view_count * 0.1) as engagement_score
            FROM phananh
            {where_clause}
            ORDER BY engagement_score DESC
            LIMIT 10
        """, params)
        
        top_engagement = cur.fetchall()
        
        # 4. Thống kê phản ánh công khai vs riêng tư
        cur.execute(f"""
            SELECT 
                CASE WHEN is_public THEN 'Công khai' ELSE 'Riêng tư' END as type,
                COUNT(*) as count
            FROM phananh
            {where_clause}
            GROUP BY is_public
        """, params)
        
        privacy_stats = cur.fetchall()
        
        conn.close()
        
        return render_template('reports_phananh.html',
                             status_stats=status_stats,
                             daily_stats=daily_stats,
                             top_engagement=top_engagement,
                             privacy_stats=privacy_stats,
                             from_date=from_date,
                             to_date=to_date,
                             trangthai=trangthai)
    
    except Exception as e:
        conn.close()
        flash(f'Lỗi khi tải báo cáo phản ánh: {str(e)}', 'danger')
        return redirect(url_for('reports_overview'))


@app.route('/reports/vande')
@login_required
@role_required(['CanBo', 'QuanLy'])
def reports_vande():
    """
    Thống kê hiệu suất xử lý vấn đề
    """
    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Build WHERE clause
        where_clause = ""
        params = []
        
        if from_date and to_date:
            where_clause = "WHERE ngaytao BETWEEN %s AND %s"
            params = [from_date, to_date]
        
        # 1. Tỷ lệ giải quyết
        cur.execute(f"""
            SELECT 
                trangthai,
                COUNT(*) as count,
                ROUND(COUNT(*) * 100.0 / SUM(COUNT(*)) OVER(), 2) as percentage
            FROM vande
            {where_clause}
            GROUP BY trangthai
        """, params)
        
        resolution_rate = cur.fetchall()
        
        # 2. Thời gian xử lý trung bình (ngày)
        cur.execute(f"""
            SELECT 
                phanloai,
                COUNT(*) as total_issues,
                ROUND(AVG(EXTRACT(EPOCH FROM (ngaycapnhat - ngaytao)) / 86400), 1) as avg_days
            FROM vande
            {where_clause if where_clause else 'WHERE'} 
            {' AND ' if where_clause else ''} ngaycapnhat IS NOT NULL
            GROUP BY phanloai
            ORDER BY avg_days DESC
        """, params)
        
        avg_resolution_time = cur.fetchall()
        
        # 3. Thống kê theo cán bộ xử lý
        cur.execute(f"""
            SELECT 
                n.hovaten,
                n.cccd,
                COUNT(v.mavande) as total_handled,
                COUNT(CASE WHEN v.trangthai = 'Đã giải quyết' THEN 1 END) as resolved,
                COUNT(CASE WHEN v.trangthai = 'Đóng' THEN 1 END) as closed,
                ROUND(AVG(EXTRACT(EPOCH FROM (v.ngaycapnhat - v.ngaytao)) / 86400), 1) as avg_days
            FROM vande v
            JOIN nguoidung n ON v.cccd_canbo_xuly = n.cccd
            {where_clause}
            GROUP BY n.hovaten, n.cccd
            ORDER BY total_handled DESC
        """, params)
        
        canbo_performance = cur.fetchall()
        
        # 4. Thống kê theo phân loại
        cur.execute(f"""
            SELECT 
                phanloai,
                COUNT(*) as total,
                COUNT(CASE WHEN trangthai IN ('Đã giải quyết', 'Đóng') THEN 1 END) as resolved
            FROM vande
            {where_clause}
            GROUP BY phanloai
            ORDER BY total DESC
        """, params)
        
        category_stats = cur.fetchall()
        
        conn.close()
        
        return render_template('reports_vande.html',
                             resolution_rate=resolution_rate,
                             avg_resolution_time=avg_resolution_time,
                             canbo_performance=canbo_performance,
                             category_stats=category_stats,
                             from_date=from_date,
                             to_date=to_date)
    
    except Exception as e:
        conn.close()
        flash(f'Lỗi khi tải báo cáo vấn đề: {str(e)}', 'danger')
        return redirect(url_for('reports_overview'))


@app.route('/reports/engagement')
@login_required
@role_required(['CanBo', 'QuanLy'])
def reports_engagement():
    """
    Thống kê tương tác người dùng
    """
    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Build WHERE clause
        where_clause = ""
        params = []
        
        if from_date and to_date:
            where_clause = "WHERE thoigiantao BETWEEN %s AND %s"
            params = [from_date, to_date]
        
        # 1. Tổng quan tương tác
        cur.execute(f"""
            SELECT 
                SUM(like_count) as total_likes,
                SUM(comment_count) as total_comments,
                SUM(view_count) as total_views,
                AVG(like_count) as avg_likes,
                AVG(comment_count) as avg_comments,
                AVG(view_count) as avg_views
            FROM phananh
            {where_clause}
        """, params)
        
        engagement_overview = cur.fetchone()
        
        # 2. Top 10 bài viết hot nhất
        cur.execute(f"""
            SELECT 
                p.maphananh,
                p.tieude,
                p.like_count,
                p.comment_count,
                p.view_count,
                n.name,
                (p.like_count * 2 + p.comment_count + p.view_count * 0.1) as hot_score
            FROM phananh p
            JOIN nguoidung n ON p.cccd = n.cccd
            {where_clause}
            ORDER BY hot_score DESC
            LIMIT 10
        """, params)
        
        top_posts = cur.fetchall()
        
        # 3. Top 10 người dùng tích cực nhất (nhiều phản ánh + like + comment)
        cur.execute(f"""
            SELECT 
                n.name,
                n.cccd,
                COUNT(DISTINCT p.maphananh) as post_count,
                COUNT(DISTINCT l.malike) as like_count,
                COUNT(DISTINCT b.mabinhluan) as comment_count,
                (COUNT(DISTINCT p.maphananh) * 3 + 
                 COUNT(DISTINCT l.malike) + 
                 COUNT(DISTINCT b.mabinhluan) * 2) as activity_score
            FROM nguoidung n
            LEFT JOIN phananh p ON n.cccd = p.cccd {' AND p.thoigiantao BETWEEN %s AND %s' if params else ''}
            LEFT JOIN like_post l ON n.cccd = l.cccd
            LEFT JOIN binhluan b ON n.cccd = b.cccd_nguoidung
            WHERE n.vaitro = 'NguoiDan'
            GROUP BY n.name, n.cccd
            ORDER BY activity_score DESC
            LIMIT 10
        """, params if params else [])
        
        top_users = cur.fetchall()
        
        # 4. Thống kê tương tác theo ngày
        cur.execute(f"""
            SELECT 
                DATE(thoigiantao) as date,
                SUM(like_count) as daily_likes,
                SUM(comment_count) as daily_comments,
                SUM(view_count) as daily_views
            FROM phananh
            {where_clause}
            GROUP BY DATE(thoigiantao)
            ORDER BY date DESC
            LIMIT 30
        """, params)
        
        daily_engagement = cur.fetchall()
        
        conn.close()
        
        return render_template('reports_engagement.html',
                             engagement_overview=engagement_overview,
                             top_posts=top_posts,
                             top_users=top_users,
                             daily_engagement=daily_engagement,
                             from_date=from_date,
                             to_date=to_date)
    
    except Exception as e:
        conn.close()
        flash(f'Lỗi khi tải báo cáo tương tác: {str(e)}', 'danger')
        return redirect(url_for('reports_overview'))


@app.route('/reports/export')
@login_required
@role_required(['CanBo', 'QuanLy'])
def reports_export():
    """
    Xuất báo cáo ra file Excel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from datetime import datetime
    
    report_type = request.args.get('type', 'overview')  # overview, phananh, vande
    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        wb = Workbook()
        ws = wb.active
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        if report_type == 'phananh':
            ws.title = "Báo cáo Phản ánh"
            
            # Build WHERE
            where_parts = []
            params = []
            if from_date and to_date:
                where_parts.append("thoigiantao BETWEEN %s AND %s")
                params.extend([from_date, to_date])
            where_clause = "WHERE " + " AND ".join(where_parts) if where_parts else ""
            
            # Headers
            headers = ["Mã PA", "Tiêu đề", "Người gửi", "Trạng thái", "Công khai", 
                      "Likes", "Comments", "Views", "Ngày tạo"]
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            cur.execute(f"""
                SELECT 
                    p.maphananh,
                    p.tieude,
                    n.name,
                    p.trangthaiphananh,
                    CASE WHEN p.is_public THEN 'Có' ELSE 'Không' END,
                    p.like_count,
                    p.comment_count,
                    p.view_count,
                    p.thoigiantao
                FROM phananh p
                JOIN nguoidung n ON p.cccd = n.cccd
                {where_clause}
                ORDER BY p.thoigiantao DESC
            """, params)
            
            for row in cur.fetchall():
                ws.append(row)
        
        elif report_type == 'vande':
            ws.title = "Báo cáo Vấn đề"
            
            where_clause = ""
            params = []
            if from_date and to_date:
                where_clause = "WHERE v.ngaytao BETWEEN %s AND %s"
                params = [from_date, to_date]
            
            headers = ["Mã VĐ", "Tên vấn đề", "Phân loại", "Trạng thái", 
                      "Cán bộ xử lý", "Kết quả", "Ngày tạo", "Ngày cập nhật"]
            ws.append(headers)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            cur.execute(f"""
                SELECT 
                    v.mavande,
                    v.tenvande,
                    v.phanloai,
                    v.trangthai,
                    n.name,
                    v.ketqua,
                    v.ngaytao,
                    v.ngaycapnhat
                FROM vande v
                LEFT JOIN nguoidung n ON v.cccd_canbo_xuly = n.cccd
                {where_clause}
                ORDER BY v.ngaytao DESC
            """, params)
            
            for row in cur.fetchall():
                ws.append(row)
        
        else:  # overview
            ws.title = "Tổng quan"
            
            # Thống kê phản ánh
            ws.append(["BÁO CÁO TỔNG QUAN HỆ THỐNG PHẢN ÁNH & KIẾN NGHỊ"])
            ws.merge_cells('A1:E1')
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws.append([])
            ws.append(["THỐNG KÊ PHẢN ÁNH"])
            ws['A3'].font = Font(bold=True, size=12)
            
            cur.execute("""
                SELECT 
                    COUNT(*) as total,
                    COUNT(CASE WHEN trangthaiphananh = 'Mới' THEN 1 END) as moi,
                    COUNT(CASE WHEN trangthaiphananh = 'Đang xử lý' THEN 1 END) as dangxuly,
                    COUNT(CASE WHEN trangthaiphananh = 'Đã xử lý' THEN 1 END) as daxuly
                FROM phananh
            """)
            
            stats = cur.fetchone()
            ws.append(["Tổng số phản ánh:", stats[0]])
            ws.append(["Phản ánh mới:", stats[1]])
            ws.append(["Đang xử lý:", stats[2]])
            ws.append(["Đã xử lý:", stats[3]])
            
            ws.append([])
            ws.append(["THỐNG KÊ VẤN ĐỀ"])
            ws['A9'].font = Font(bold=True, size=12)
            
            cur.execute("""
                SELECT 
                    COUNT(*) as total,
                    COUNT(CASE WHEN trangthai = 'Đã giải quyết' THEN 1 END) as resolved
                FROM vande
            """)
            
            vande_stats = cur.fetchone()
            ws.append(["Tổng số vấn đề:", vande_stats[0]])
            ws.append(["Đã giải quyết:", vande_stats[1]])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        conn.close()
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"BaoCao_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        conn.close()
        flash(f'Lỗi khi xuất báo cáo: {str(e)}', 'danger')
        return redirect(url_for('reports_overview'))


# ========== CHẠY ỨNG DỤNG ==========
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
